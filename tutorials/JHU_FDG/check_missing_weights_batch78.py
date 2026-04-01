from __future__ import annotations

import argparse
import csv
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Set, Tuple

from openpyxl import load_workbook


DEFAULT_MERGED_CSV = Path(
    "/scratch/jchen/python_projects/clinical_reports/batch78_img_clinical_os_merged_annotated.csv"
)
DEFAULT_WEIGHT_XLSX = Path("/scratch2/jchen/DATA/JHU_FDG/CCDA10838.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Count missing-weight cases for JHU_FDG batches 7/8 directly from the merged "
            "clinical CSV MRN/study dates against the CCDA workbook."
        )
    )
    parser.add_argument("--merged-csv", type=Path, default=DEFAULT_MERGED_CSV)
    parser.add_argument("--weights-xlsx", type=Path, default=DEFAULT_WEIGHT_XLSX)
    parser.add_argument(
        "--allow-missing-mrn",
        action="store_true",
        help=(
            "Include rows with blank MRN in the missing-weight count instead of treating them as a separate category."
        ),
    )
    parser.add_argument(
        "--show-sample",
        type=int,
        default=0,
        help="Print the first N merged rows without an exact same-day CCDA weight match.",
    )
    return parser.parse_args()


def normalize_study_date(row: Dict[str, str]) -> Optional[str]:
    study_date = (row.get("StudyDate") or "").strip()
    if len(study_date) == 8 and study_date.isdigit():
        return f"{study_date[:4]}-{study_date[4:6]}-{study_date[6:8]}"

    slash_date = (row.get("study_date") or "").strip()
    if slash_date:
        try:
            return datetime.strptime(slash_date, "%m/%d/%Y").date().isoformat()
        except ValueError:
            return None
    return None


def load_weight_keys(weights_xlsx: Path) -> Tuple[Set[Tuple[str, str]], Dict[str, Set[str]]]:
    workbook = load_workbook(weights_xlsx, read_only=True, data_only=True)
    worksheet = workbook["Weight"]

    weight_keys: Set[Tuple[str, str]] = set()
    weight_dates_by_mrn: Dict[str, Set[str]] = {}
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_index == 1:
            continue

        mrn = row[0]
        exam_started = row[1]
        if mrn is None or exam_started is None:
            continue
        if not isinstance(exam_started, datetime):
            continue

        mrn_str = str(mrn).strip()
        exam_date = exam_started.date().isoformat()
        weight_keys.add((mrn_str, exam_date))
        weight_dates_by_mrn.setdefault(mrn_str, set()).add(exam_date)

    return weight_keys, weight_dates_by_mrn


def summarize_direct_xlsx_missing(
    merged_csv: Path,
    weight_keys: Set[Tuple[str, str]],
    weight_dates_by_mrn: Dict[str, Set[str]],
    allow_missing_mrn: bool = False,
    show_sample: int = 0,
) -> Tuple[int, Counter, Counter, Counter, Counter, list]:
    total_images = 0
    missing_count = 0
    total_by_batch: Counter = Counter()
    same_day_match_by_batch: Counter = Counter()
    missing_by_batch: Counter = Counter()
    reason_counts: Counter = Counter()
    sample_rows = []

    with merged_csv.open(newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            session_id = (row.get("XNATSessionID") or "").strip()
            mrn = (row.get("MRN") or "").strip()
            study_date = normalize_study_date(row)
            batch = (row.get("batch") or "").strip()

            total_images += 1
            total_by_batch[batch] += 1

            reason: Optional[str] = None
            if not mrn:
                reason = "missing_mrn"
            elif not study_date:
                reason = "missing_study_date"
            elif (mrn, study_date) not in weight_keys:
                if mrn in weight_dates_by_mrn:
                    reason = "ccda_weight_not_same_day"
                else:
                    reason = "no_ccda_weight_for_mrn"

            if reason is None:
                same_day_match_by_batch[batch] += 1
                continue

            reason_counts[reason] += 1

            if reason == "missing_mrn" and not allow_missing_mrn:
                continue

            if show_sample > 0 and len(sample_rows) < show_sample:
                sample_rows.append(
                    {
                        "XNATSessionID": session_id,
                        "MRN": mrn,
                        "batch": batch,
                        "StudyDate": (row.get("StudyDate") or "").strip(),
                        "study_date": (row.get("study_date") or "").strip(),
                        "reason": reason,
                    }
                )

            missing_count += 1
            missing_by_batch[batch] += 1

    return total_images, total_by_batch, same_day_match_by_batch, missing_by_batch, reason_counts, sample_rows


def print_counter(title: str, counter: Counter) -> None:
    print(title)
    for key in sorted(counter):
        print(f"  {key}: {counter[key]}")


def main() -> None:
    args = parse_args()

    weight_keys, weight_dates_by_mrn = load_weight_keys(args.weights_xlsx)
    total_images, total_by_batch, same_day_match_by_batch, direct_missing_by_batch, reason_counts, sample_rows = summarize_direct_xlsx_missing(
        args.merged_csv,
        weight_keys,
        weight_dates_by_mrn,
        allow_missing_mrn=args.allow_missing_mrn,
        show_sample=args.show_sample,
    )
    direct_missing_count = sum(direct_missing_by_batch.values())
    same_day_match_count = sum(same_day_match_by_batch.values())
    no_same_day_but_has_weight = reason_counts.get("ccda_weight_not_same_day", 0)
    no_weight_for_mrn = reason_counts.get("no_ccda_weight_for_mrn", 0)
    missing_mrn_count = reason_counts.get("missing_mrn", 0)
    missing_study_date_count = reason_counts.get("missing_study_date", 0)

    print("CCDA weight summary computed from merged CSV MRN/study date against the CCDA Weight sheet")
    print(f"  unique (MRN, exam_date) keys in workbook: {len(weight_keys)}")
    print(f"  total images in merged batch7+batch8 CSV: {total_images}")
    print_counter("  total images by batch:", total_by_batch)
    print(f"  images with a same-day CCDA weight: {same_day_match_count}")
    print_counter("  same-day CCDA weight matches by batch:", same_day_match_by_batch)
    print(f"  images missing a same-day CCDA weight: {direct_missing_count}")
    print_counter("  missing same-day CCDA weight by batch:", direct_missing_by_batch)
    print(f"  of those missing same-day matches, images with a CCDA weight on a different day: {no_same_day_but_has_weight}")
    print(f"  of those missing same-day matches, images with no CCDA weight for that MRN: {no_weight_for_mrn}")

    if missing_mrn_count:
        print(f"  rows with blank MRN in merged CSV: {missing_mrn_count}")
    if missing_study_date_count:
        print(f"  rows with unreadable or missing study date in merged CSV: {missing_study_date_count}")

    if not args.allow_missing_mrn and reason_counts.get("missing_mrn", 0):
        print("  note: rows with blank MRN are reported above but excluded from the main missing-weight total")

    if sample_rows:
        print()
        print("Sample merged rows without exact same-day CCDA weight match")
        for row in sample_rows:
            print(
                "  "
                f"{row['XNATSessionID']} | batch={row['batch']} | MRN={row['MRN']} | "
                f"StudyDate={row['StudyDate']} | study_date={row['study_date']} | reason={row['reason']}"
            )


if __name__ == "__main__":
    main()