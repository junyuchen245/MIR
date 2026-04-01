#!/usr/bin/env python3
import argparse
import csv
import glob
import json
import math
import os
import re
import statistics
from collections import Counter, defaultdict
from datetime import datetime
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Summarize PSMA clinical JSON files into a demographics table. "
            "By default, this uses one record per patient (the earliest scan)."
        )
    )
    parser.add_argument(
        "--clinical-dir",
        default="/scratch2/jchen/DATA/PSMA_JHU/Preprocessed/clinical_jsons",
        help="Folder containing scan-level clinical JSON files.",
    )
    parser.add_argument(
        "--unit",
        choices=["patient", "scan"],
        default="patient",
        help="Summarize unique patients or all scan JSONs.",
    )
    parser.add_argument(
        "--dicom-dir",
        default="/scratch2/jchen/DATA/PSMA_JHU/Preprocessed/JHU_dicom_info",
        help="Folder containing scan-level DICOM metadata JSON files for tracer statistics.",
    )
    parser.add_argument(
        "--image-dir",
        default="/scratch2/jchen/DATA/PSMA_JHU/Preprocessed/JHU",
        help="Folder containing processed image files used to define the available imaging cohort.",
    )
    parser.add_argument(
        "--image-filter",
        choices=["ct-suv", "any", "none"],
        default="ct-suv",
        help=(
            "Restrict the cohort to scans with actual processed image files. "
            "'ct-suv' requires both CT and SUV; 'any' requires at least one image; 'none' keeps all clinical records."
        ),
    )
    parser.add_argument(
        "--assignment-xlsx",
        default="/scratch2/jchen/DATA/PSMA_JHU/Preprocessed/follow-up_assignment8(AutoRecovered).xlsx",
        help="Workbook containing the Recip patient to XNATSessionID crosswalk for old PSMA aliases.",
    )
    parser.add_argument(
        "--scan-choice",
        choices=["earliest", "latest"],
        default="earliest",
        help="Which scan to keep when --unit patient is used.",
    )
    parser.add_argument(
        "--sex-mode",
        choices=["infer-male", "unknown"],
        default="infer-male",
        help=(
            "How to handle sex when no sex field exists in the JSONs. "
            "'infer-male' treats this prostate cohort as men; 'unknown' keeps it missing."
        ),
    )
    parser.add_argument(
        "--therapy-mode",
        choices=["dominant", "combination"],
        default="dominant",
        help=(
            "How to summarize post-PSMA therapy. 'dominant' collapses combinations into a "
            "single main category; 'combination' keeps exact combinations."
        ),
    )
    parser.add_argument(
        "--format",
        choices=["text", "markdown", "csv"],
        default="text",
        help="Output format.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Optional output file. If omitted, prints to stdout.",
    )
    return parser.parse_args()


def get_value(record: Dict, key: str) -> Optional[str]:
    value = record.get(key, None)
    if isinstance(value, dict):
        value = value.get("data", None)
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def parse_float(value: Optional[str]) -> float:
    if value is None:
        return math.nan
    try:
        return float(value)
    except Exception:
        return math.nan


def parse_scan_date(scan_name: str) -> Optional[datetime]:
    try:
        return datetime.strptime(scan_name.split("_", 1)[1], "%Y-%m-%d")
    except Exception:
        return None


def parse_any_date(text: Optional[str]) -> Optional[datetime]:
    if text is None:
        return None
    for fmt in ("%Y-%m-%d", "%m.%d.%Y", "%m.%d.%y", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass
    return None


def parse_psa(text: Optional[str]) -> float:
    if text is None:
        return math.nan
    match = re.match(r"^<\s*([0-9.]+)\s*$", text)
    if match:
        return 0.5 * float(match.group(1))
    try:
        return float(text)
    except Exception:
        return math.nan


def parse_binary_12(text: Optional[str]) -> int:
    try:
        return 1 if int(float(text)) == 1 else 0
    except Exception:
        return 0


def parse_gleason_bucket(text: Optional[str]) -> str:
    if text is None:
        return "NA"
    compact = text.replace(" ", "")
    match = re.match(r"^(\d)\+(\d)(?:=(\d+))?$", compact)
    if match:
        gsum = int(match.group(3)) if match.group(3) else int(match.group(1)) + int(match.group(2))
    elif compact.isdigit():
        gsum = int(compact)
    else:
        return "NA"
    if gsum <= 6:
        return "≤6"
    if gsum == 7:
        return "7"
    if gsum == 8:
        return "8"
    if gsum == 9:
        return "9"
    if gsum == 10:
        return "10"
    return "NA"


def parse_race_bucket(text: Optional[str]) -> str:
    try:
        value = int(float(text))
    except Exception:
        return "Unknown"
    return {
        1: "White",
        2: "Black or African American",
        3: "Asian",
        4: "Other",
    }.get(value, "Other")


def parse_sex_bucket(record: Dict, sex_mode: str) -> str:
    sex_keys = ["Sex", "sex", "Gender", "gender"]
    raw = None
    for key in sex_keys:
        raw = get_value(record, key)
        if raw is not None:
            break
    if raw is None:
        return "Men" if sex_mode == "infer-male" else "Unknown"
    normalized = raw.strip().lower()
    if normalized in {"m", "male", "1", "man", "men"}:
        return "Men"
    if normalized in {"f", "female", "2", "woman", "women"}:
        return "Women"
    return "Unknown"


def dominant_therapy_bucket(record: Dict, prefix: str) -> str:
    keys = [
        f"{prefix} Local",
        f"{prefix} Focal",
        f"{prefix} Systemic Androgen Targeted",
        f"{prefix} Systemic and cytotxic",
    ]
    values = [get_value(record, key) for key in keys]
    if all(value is None for value in values):
        return "NA"
    local = parse_binary_12(values[0]) or parse_binary_12(values[1])
    androgen_targeted = parse_binary_12(values[2])
    cytotoxic = parse_binary_12(values[3])
    if local == 0 and androgen_targeted == 0 and cytotoxic == 0:
        return "None"
    if cytotoxic:
        return "Systemic and cytotoxic"
    if androgen_targeted:
        return "Systemic androgen-targeted"
    if local:
        return "Local"
    return "NA"


def combination_therapy_bucket(record: Dict, prefix: str) -> str:
    keys = [
        f"{prefix} Local",
        f"{prefix} Focal",
        f"{prefix} Systemic Androgen Targeted",
        f"{prefix} Systemic and cytotxic",
    ]
    values = [get_value(record, key) for key in keys]
    if all(value is None for value in values):
        return "NA"
    labels: List[str] = []
    if parse_binary_12(values[0]) or parse_binary_12(values[1]):
        labels.append("Local")
    if parse_binary_12(values[2]):
        labels.append("Systemic androgen-targeted")
    if parse_binary_12(values[3]):
        labels.append("Systemic and cytotoxic")
    if not labels:
        return "None"
    return " + ".join(labels)


def load_records(clinical_dir: str, image_dir: Optional[str] = None) -> List[Dict]:
    records: List[Dict] = []
    pattern = os.path.join(clinical_dir, "*.json")
    for path in sorted(glob.glob(pattern)):
        with open(path, "r") as handle:
            payload = json.load(handle)
        scan_name = os.path.splitext(os.path.basename(path))[0]
        patient_id = scan_name.split("_", 1)[0]
        image_flags = {
            "ct": False,
            "suv": False,
            "ct_seg": False,
            "suv_seg": False,
        }
        if image_dir:
            image_flags = {
                "ct": os.path.exists(os.path.join(image_dir, f"{scan_name}_CT.nii.gz")),
                "suv": os.path.exists(os.path.join(image_dir, f"{scan_name}_SUV.nii.gz")),
                "ct_seg": os.path.exists(os.path.join(image_dir, f"{scan_name}_CT_seg.nii.gz")),
                "suv_seg": os.path.exists(os.path.join(image_dir, f"{scan_name}_SUV_seg.nii.gz")),
            }
        records.append(
            {
                "path": path,
                "scan_name": scan_name,
                "patient_id": patient_id,
                "scan_date": parse_scan_date(scan_name),
                "image_flags": image_flags,
                "record": payload,
            }
        )
    return records


def normalize_tracer_name(text: Optional[str]) -> str:
    value = (text or "").strip()
    if not value:
        return "Missing"
    key = value.lower().replace(" ", "")
    if any(token in key for token in ["dcfpyl", "pylarify", "piflufolastat", "pilfufolastat"]):
        return "DCFPyL / Pylarify"
    if "fluciclovine" in key:
        return "Fluciclovine"
    if key in {"psma", "psma11"}:
        return "68Ga-PSMA"
    if "naf" in key:
        return "NaF"
    if "fluorodeoxyglucose" in key or "fdg" in key:
        return "FDG"
    if "solution" in key:
        return "Solution / unspecified"
    return value


def normalize_radionuclide(text: Optional[str]) -> str:
    value = (text or "").strip()
    if not value:
        return "Missing"
    lower = value.lower()
    if "18" in lower and ("fluor" in lower or lower == "18f"):
        return "18F"
    if "68" in lower and ("gall" in lower or lower == "68ga"):
        return "68Ga"
    return value


def infer_radionuclide_from_tracer(tracer_name: str) -> str:
    tracer = normalize_tracer_name(tracer_name)
    if tracer in {"DCFPyL / Pylarify", "Fluciclovine", "NaF", "FDG"}:
        return "18F"
    if tracer == "68Ga-PSMA":
        return "68Ga"
    return "Missing"


def normalize_scanner_model(text: Optional[str]) -> str:
    value = (text or "").strip()
    return value if value else "Missing"


def load_dicom_records(dicom_dir: str) -> Tuple[Dict[str, Dict], Dict[str, List[Dict]]]:
    records: Dict[str, Dict] = {}
    records_by_date: Dict[str, List[Dict]] = defaultdict(list)
    pattern = os.path.join(dicom_dir, "*.json")
    for path in sorted(glob.glob(pattern)):
        scan_name = os.path.splitext(os.path.basename(path))[0]
        with open(path, "r") as handle:
            payload = json.load(handle)
        records[scan_name] = payload
        date_part = scan_name.split("_", 1)[1] if "_" in scan_name else ""
        records_by_date[date_part].append(payload)
    return records, records_by_date


def normalize_session_id(session_id: str) -> str:
    text = str(session_id).strip()
    if not text or "_" not in text:
        return text
    patient_id, date_part = text.split("_", 1)
    if len(date_part) == 8 and date_part.isdigit():
        date_part = f"{date_part[:4]}-{date_part[4:6]}-{date_part[6:8]}"
    return f"{patient_id}_{date_part}"


def parse_recip_alias_number(value: object) -> Optional[int]:
    if value is None:
        return None
    match = re.match(r"^(\d+)", str(value).strip())
    if not match:
        return None
    return int(match.group(1))


def load_assignment_crosswalk(workbook_path: str) -> Dict[str, str]:
    if not workbook_path or not os.path.exists(workbook_path):
        return {}
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    crosswalk: Dict[str, str] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        recip_value = row[1] if len(row) > 1 else None
        xnat_session_id = row[2] if len(row) > 2 else None
        recip_number = parse_recip_alias_number(recip_value)
        if recip_number is None or xnat_session_id is None:
            continue
        normalized_xnat = normalize_session_id(str(xnat_session_id))
        if "_" not in normalized_xnat:
            continue
        _, date_part = normalized_xnat.split("_", 1)
        alias = f"PSMA{recip_number:04d}_{date_part}"
        crosswalk[normalized_xnat] = alias
    return crosswalk


def match_dicom_record(
    scan_name: str,
    dicom_records: Dict[str, Dict],
    dicom_records_by_date: Dict[str, List[Dict]],
    assignment_crosswalk: Dict[str, str],
) -> Tuple[Optional[Dict], str]:
    exact = dicom_records.get(scan_name)
    if exact is not None:
        return exact, "exact"
    alias_scan_name = assignment_crosswalk.get(scan_name)
    if alias_scan_name:
        alias_match = dicom_records.get(alias_scan_name)
        if alias_match is not None:
            return alias_match, "assignment-crosswalk"
    date_part = scan_name.split("_", 1)[1] if "_" in scan_name else ""
    date_matches = dicom_records_by_date.get(date_part, [])
    if len(date_matches) == 1:
        return date_matches[0], "date-fallback"
    return None, "missing"


def extract_tracer_stats(dicom_record: Optional[Dict]) -> Dict[str, object]:
    if not dicom_record:
        return {
            "tracer": "Missing",
            "radionuclide": "Missing",
            "scanner_model": "Missing",
            "dose_mbq": math.nan,
        }
    sequence = dicom_record.get("RadiopharmaceuticalInformationSequence") or []
    tracer_name = "Missing"
    radionuclide = "Missing"
    dose_mbq = math.nan
    if sequence:
        item = sequence[0] if isinstance(sequence[0], dict) else {}
        tracer_name = normalize_tracer_name(str(item.get("Radiopharmaceutical", "")).strip())
        code_sequence = item.get("RadionuclideCodeSequence") or []
        if code_sequence and isinstance(code_sequence[0], dict):
            radionuclide = normalize_radionuclide(str(code_sequence[0].get("CodeMeaning", "")).strip())
        else:
            radionuclide = normalize_radionuclide(str(item.get("Radionuclide", "")).strip())
        if radionuclide == "Missing":
            radionuclide = infer_radionuclide_from_tracer(tracer_name)
        try:
            dose_mbq = float(item.get("RadionuclideTotalDose")) / 1e6
        except Exception:
            dose_mbq = math.nan
    return {
        "tracer": tracer_name,
        "radionuclide": radionuclide,
        "scanner_model": normalize_scanner_model(dicom_record.get("ManufacturerModelName")),
        "dose_mbq": dose_mbq,
    }


def choose_records(records: Sequence[Dict], unit: str, scan_choice: str) -> List[Dict]:
    if unit == "scan":
        return list(records)
    by_patient: Dict[str, List[Dict]] = defaultdict(list)
    for item in records:
        by_patient[item["patient_id"]].append(item)
    chosen: List[Dict] = []
    reverse = scan_choice == "latest"
    for patient_id in sorted(by_patient):
        scans = sorted(
            by_patient[patient_id],
            key=lambda item: (item["scan_date"] is None, item["scan_date"]),
            reverse=reverse,
        )
        chosen.append(scans[0])
    return chosen


def record_matches_image_filter(item: Dict, image_filter: str) -> bool:
    if image_filter == "none":
        return True
    image_flags = item.get("image_flags", {})
    if image_filter == "ct-suv":
        return bool(image_flags.get("ct")) and bool(image_flags.get("suv"))
    if image_filter == "any":
        return any(bool(value) for value in image_flags.values())
    return True


def filter_records_by_images(records: Sequence[Dict], image_filter: str) -> List[Dict]:
    return [item for item in records if record_matches_image_filter(item, image_filter)]


def valid_numbers(values: Iterable[float]) -> List[float]:
    return [value for value in values if value is not None and not math.isnan(value)]


def mean_sd_text(values: Iterable[float], decimals: int = 2) -> str:
    data = valid_numbers(values)
    if not data:
        return "NA"
    mean_value = statistics.mean(data)
    sd_value = statistics.stdev(data) if len(data) > 1 else 0.0
    return f"{mean_value:.{decimals}f} ± {sd_value:.{decimals}f}"


def median_range_text(values: Iterable[float], decimals: int = 2) -> str:
    data = valid_numbers(values)
    if not data:
        return "NA"
    median_value = statistics.median(data)
    return f"{median_value:.{decimals}f} ({min(data):.{decimals}f}–{max(data):.{decimals}f})"


def maybe_months_between(record: Dict) -> float:
    baseline = parse_any_date(get_value(record, "psa time"))
    followup = parse_any_date(get_value(record, "relapsetime"))
    if baseline is None or followup is None:
        return math.nan
    days = max((followup - baseline).days, 0)
    return days / 30.4375


def summarize(
    records: Sequence[Dict],
    sex_mode: str,
    therapy_mode: str,
    unit: str,
    dicom_records: Optional[Dict[str, Dict]] = None,
    dicom_records_by_date: Optional[Dict[str, List[Dict]]] = None,
    assignment_crosswalk: Optional[Dict[str, str]] = None,
) -> Tuple[List[Tuple[str, str]], List[str]]:
    raw_records = [item["record"] for item in records]
    n_total = len(records)
    ages = [parse_float(get_value(record, "Age")) for record in raw_records]
    initial_psa = [parse_psa(get_value(record, "Initial PSA")) for record in raw_records]
    pre_psma_psa = [parse_psa(get_value(record, "PRE PSMA PSA")) for record in raw_records]
    post_psma_psa = [parse_psa(get_value(record, "Post PSMA PSA")) for record in raw_records]
    pseudo_followup_months = [maybe_months_between(record) for record in raw_records]

    sex_counts = Counter(parse_sex_bucket(record, sex_mode) for record in raw_records)
    race_counts = Counter(parse_race_bucket(get_value(record, "Race")) for record in raw_records)
    gleason_counts = Counter(parse_gleason_bucket(get_value(record, "Gleason score")) for record in raw_records)

    dicom_matches = [
        match_dicom_record(
            item["scan_name"],
            dicom_records or {},
            dicom_records_by_date or {},
            assignment_crosswalk or {},
        )
        for item in records
    ]
    dicom_stats = [extract_tracer_stats(match[0]) for match in dicom_matches]
    match_source_counts = Counter(match_source for _, match_source in dicom_matches)
    tracer_counts = Counter(stat["tracer"] for stat in dicom_stats)
    radionuclide_counts = Counter(stat["radionuclide"] for stat in dicom_stats)
    scanner_counts = Counter(stat["scanner_model"] for stat in dicom_stats)
    dose_mbq = [float(stat["dose_mbq"]) for stat in dicom_stats]

    therapy_func = dominant_therapy_bucket if therapy_mode == "dominant" else combination_therapy_bucket
    therapy_counts = Counter(therapy_func(record, "Post PSMA") for record in raw_records)

    rows: List[Tuple[str, str]] = []
    rows.append((f"Dataset ({'patients' if unit == 'patient' else 'scans'})", str(n_total)))
    rows.append(("Age (y) (mean ± SD)", mean_sd_text(ages)))
    rows.append(("Sex", ""))
    for label in ["Men", "Women", "Unknown"]:
        rows.append((f"  {label}", str(sex_counts.get(label, 0))))
    rows.append(("Race", ""))
    for label in ["White", "Black or African American", "Asian", "Other", "Unknown"]:
        rows.append((f"  {label}", str(race_counts.get(label, 0))))
    rows.append(("Gleason score", ""))
    for label in ["NA", "≤6", "7", "8", "9", "10"]:
        rows.append((f"  {label}", str(gleason_counts.get(label, 0))))
    rows.append(("Initial PSA level (ng/mL)", median_range_text(initial_psa)))
    rows.append(("Pre-PSMA PSA level (ng/mL)", median_range_text(pre_psma_psa)))
    rows.append(("Post-PSMA PSA level (ng/mL)", median_range_text(post_psma_psa)))
    rows.append(("PSA follow-up interval (mo)", median_range_text(pseudo_followup_months)))
    rows.append(("Tracer", ""))
    for label in ["DCFPyL / Pylarify", "68Ga-PSMA", "Fluciclovine", "NaF", "FDG", "Solution / unspecified", "Missing"]:
        if label in tracer_counts or label == "Missing":
            rows.append((f"  {label}", str(tracer_counts.get(label, 0))))
    extra_tracers = [label for label in tracer_counts if label not in {"DCFPyL / Pylarify", "68Ga-PSMA", "Fluciclovine", "NaF", "FDG", "Solution / unspecified", "Missing"}]
    for label in sorted(extra_tracers):
        rows.append((f"  {label}", str(tracer_counts.get(label, 0))))
    rows.append(("Radionuclide", ""))
    for label in ["18F", "68Ga", "Missing"]:
        rows.append((f"  {label}", str(radionuclide_counts.get(label, 0))))
    rows.append(("Injected dose (MBq)", median_range_text(dose_mbq, decimals=1)))
    rows.append(("Scanner model", ""))
    for label, count in scanner_counts.most_common():
        rows.append((f"  {label}", str(count)))
    rows.append(("Post-PSMA therapy", ""))

    if therapy_mode == "dominant":
        therapy_order = ["NA", "None", "Local", "Systemic androgen-targeted", "Systemic and cytotoxic"]
    else:
        therapy_order = ["NA", "None"] + sorted(
            [label for label in therapy_counts if label not in {"NA", "None"}],
            key=lambda label: (label.count("+"), label),
        )
    for label in therapy_order:
        rows.append((f"  {label}", str(therapy_counts.get(label, 0))))

    notes: List[str] = []
    if sex_mode == "infer-male" and all(get_value(record, key) is None for record in raw_records for key in ["Sex", "sex", "Gender", "gender"]):
        notes.append("Sex is not stored in these JSONs; this summary infers an all-male cohort because the dataset is prostate cancer.")
    notes.append("PSA doubling time is not available as a direct field in these JSONs.")
    notes.append("The displayed PSA follow-up interval is the time from 'psa time' to 'relapsetime' when both dates exist; it is not a doubling-time estimate.")
    notes.append("Tracer, radionuclide, dose, and scanner-model statistics come from matched scan-level DICOM metadata JSONs in JHU_dicom_info.")
    notes.append("When explicit radionuclide coding is absent but the tracer label is a known agent such as DCFPyL/Pylarify, the radionuclide is inferred from the tracer identity.")
    if match_source_counts.get("assignment-crosswalk", 0) > 0:
        notes.append(f"{match_source_counts.get('assignment-crosswalk', 0)} selected scans were matched through the follow-up assignment workbook that maps renamed XNAT session IDs back to old PSMA aliases.")
    if match_source_counts.get("date-fallback", 0) > 0:
        notes.append(f"{match_source_counts.get('date-fallback', 0)} selected scans used a unique same-date fallback match because no exact scan-name DICOM JSON was found.")
    if tracer_counts.get("Missing", 0) > 0:
        notes.append("Some selected scans do not have matched radiopharmaceutical metadata in JHU_dicom_info, so their tracer statistics are reported as Missing.")
    if therapy_mode == "dominant":
        notes.append("Dominant therapy mode collapses combination treatments into a single highest-intensity category.")
    else:
        notes.append("Combination therapy mode preserves exact combinations of local, androgen-targeted, and cytotoxic treatment flags.")
    return rows, notes


def format_text(rows: Sequence[Tuple[str, str]], notes: Sequence[str]) -> str:
    left_width = max(len(label) for label, _ in rows)
    lines: List[str] = []
    for label, value in rows:
        if value:
            lines.append(f"{label.ljust(left_width)}  {value}")
        else:
            lines.append(label)
    if notes:
        lines.append("")
        lines.append("Notes")
        for note in notes:
            lines.append(f"- {note}")
    return "\n".join(lines)


def format_markdown(rows: Sequence[Tuple[str, str]], notes: Sequence[str]) -> str:
    lines = ["| Variable | Value |", "| --- | --- |"]
    for label, value in rows:
        lines.append(f"| {label} | {value} |")
    if notes:
        lines.append("")
        lines.append("**Notes**")
        for note in notes:
            lines.append(f"- {note}")
    return "\n".join(lines)


def write_csv(rows: Sequence[Tuple[str, str]], output_path: str) -> None:
    with open(output_path, "w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Variable", "Value"])
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    records = load_records(args.clinical_dir, args.image_dir)
    imaging_records = filter_records_by_images(records, args.image_filter)
    selected = choose_records(imaging_records, args.unit, args.scan_choice)
    dicom_records, dicom_records_by_date = load_dicom_records(args.dicom_dir)
    assignment_crosswalk = load_assignment_crosswalk(args.assignment_xlsx)
    rows, notes = summarize(
        selected,
        args.sex_mode,
        args.therapy_mode,
        args.unit,
        dicom_records,
        dicom_records_by_date,
        assignment_crosswalk,
    )
    if args.image_filter != "none":
        excluded_count = len(records) - len(imaging_records)
        if args.unit == "patient":
            total_patients = len({item["patient_id"] for item in records})
            included_patients = len({item["patient_id"] for item in imaging_records})
            notes.insert(
                0,
                (
                    f"This summary is restricted to patients with available processed images in {args.image_dir}; "
                    f"{included_patients} of {total_patients} patients had scans meeting the '{args.image_filter}' image filter."
                ),
            )
        else:
            notes.insert(
                0,
                (
                    f"This summary is restricted to scans with available processed images in {args.image_dir}; "
                    f"{len(imaging_records)} of {len(records)} clinical scans met the '{args.image_filter}' image filter."
                ),
            )
        if excluded_count > 0:
            notes.append(
                f"{excluded_count} clinical scan records were excluded because the required processed images were not present."
            )

    if args.format == "csv":
        if not args.output:
            raise ValueError("--output is required when --format csv is used.")
        write_csv(rows, args.output)
        return

    rendered = format_markdown(rows, notes) if args.format == "markdown" else format_text(rows, notes)
    if args.output:
        with open(args.output, "w") as handle:
            handle.write(rendered + "\n")
    else:
        print(rendered)


if __name__ == "__main__":
    main()
