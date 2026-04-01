import argparse
import csv
import glob
import math
import os
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Sequence, Tuple

import matplotlib.pyplot as plt
import numpy as np
import SimpleITK as sitk
from radiomics import featureextractor
from scipy import ndimage


THIS_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_PREPROCESSED_DIR = '/scratch2/jchen/DATA/HECKTOR25/preprocessed'
DEFAULT_CLINICAL_XLSX = '/scratch2/jchen/DATA/HECKTOR25/HECKTOR_2025_Training_EHR_with_Data_Dictionary.xlsx'
DEFAULT_OUTPUT_DIR = os.path.join(THIS_DIR, 'population_stats_tumor_radiomics')


RADIOMICS_IMAGE_TYPES = ('Original',)
PET_RADIOMICS_CLASSES = ('shape', 'firstorder', 'glcm', 'glrlm', 'glszm', 'gldm', 'ngtdm')
CT_RADIOMICS_CLASSES = ('firstorder', 'glcm', 'glrlm', 'glszm', 'gldm', 'ngtdm')


def add_bool_flag(
    parser: argparse.ArgumentParser,
    name: str,
    dest: str,
    default: bool,
    help_enabled: str,
    help_disabled: str,
    positive_aliases: Optional[Sequence[str]] = None,
    negative_aliases: Optional[Sequence[str]] = None,
) -> None:
    positive_flags = [f'--{name}'] + list(positive_aliases or [])
    negative_flags = [f'--no-{name}'] + list(negative_aliases or [])
    group = parser.add_mutually_exclusive_group()
    group.add_argument(*positive_flags, dest=dest, action='store_true', help=help_enabled)
    group.add_argument(*negative_flags, dest=dest, action='store_false', help=help_disabled)
    parser.set_defaults(**{dest: default})


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Baseline HECKTOR25 survival analysis using tumor statistics and PET/CT radiomics.'
    )
    parser.add_argument('--preprocessed-dir', default=DEFAULT_PREPROCESSED_DIR)
    parser.add_argument('--clinical-xlsx', default=DEFAULT_CLINICAL_XLSX)
    parser.add_argument('--clinical-sheet', default='Data')
    parser.add_argument('--output-dir', default=DEFAULT_OUTPUT_DIR)
    parser.add_argument('--task', choices=['all', '1', '2', '3'], default='all')
    parser.add_argument('--max-cases', type=int, default=0, help='Limit processing to the first N matched cases for quick debugging; 0 uses all cases.')
    parser.add_argument('--lesion-min-voxels', type=int, default=10)
    parser.add_argument('--min-coverage', type=int, default=15)
    parser.add_argument('--min-events', type=int, default=5)
    parser.add_argument('--feature-selection', choices=['conventional', 'all', 'topk', 'fdr05', 'fdr10'], default='conventional')
    parser.add_argument('--topk', type=int, default=12, help='Used for explicit top-k selection or as the fallback size for conventional FDR-first selection.')
    parser.add_argument('--max-selected-features', type=int, default=10, help='Maximum number of imaging features retained after ranking/selection. Set to 0 to disable capping.')
    parser.add_argument('--correlation-threshold', type=float, default=0.90)
    parser.add_argument('--pet-bin-width', type=float, default=0.25)
    parser.add_argument('--ct-bin-width', type=float, default=25.0)
    parser.add_argument('--force-2d', action='store_true', help='Run radiomics in 2D mode instead of 3D.')
    add_bool_flag(
        parser,
        name='tumor-stats',
        dest='use_tumor_stats',
        default=True,
        help_enabled='Include handcrafted tumor statistics such as MTV, TLA, lesion count, SUV, and CT summaries.',
        help_disabled='Disable handcrafted tumor statistics.',
        positive_aliases=['--use-tumor-stats'],
    )
    add_bool_flag(
        parser,
        name='pet-radiomics',
        dest='use_pet_radiomics',
        default=True,
        help_enabled='Include PET radiomics from predicted tumor VOIs.',
        help_disabled='Disable PET radiomics.',
    )
    add_bool_flag(
        parser,
        name='ct-radiomics',
        dest='use_ct_radiomics',
        default=True,
        help_enabled='Include CT radiomics from predicted tumor VOIs.',
        help_disabled='Disable CT radiomics.',
    )
    add_bool_flag(
        parser,
        name='drop-correlated',
        dest='drop_correlated',
        default=True,
        help_enabled='Greedily remove highly correlated imaging features before multivariable score construction.',
        help_disabled='Keep correlated imaging features.',
    )
    return parser.parse_args()


def column_ref_to_index(ref: str) -> int:
    col = ''.join(ch for ch in ref if ch.isalpha())
    out = 0
    for ch in col:
        out = out * 26 + (ord(ch.upper()) - ord('A') + 1)
    return out - 1


def parse_xlsx_cell_value(cell: ET.Element, shared_strings: Sequence[str], ns: str):
    t = cell.attrib.get('t')
    if t == 'inlineStr':
        is_elem = cell.find(f'{ns}is')
        if is_elem is None:
            return ''
        return ''.join(txt.text or '' for txt in is_elem.iter(f'{ns}t'))
    v = cell.find(f'{ns}v')
    if v is None:
        return None
    text = v.text
    if text is None:
        return None
    if t == 's':
        return shared_strings[int(text)]
    if t == 'b':
        return int(text)
    try:
        num = float(text)
        return int(num) if num.is_integer() else num
    except Exception:
        return text


def load_rows_openpyxl(path: str, sheet_name: str) -> List[Dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(value).strip() if value is not None else '' for value in rows[0]]
    data_rows: List[Dict[str, object]] = []
    for values in rows[1:]:
        row = {header[index]: values[index] if index < len(values) else None for index in range(len(header))}
        data_rows.append(row)
    return data_rows


def load_rows_xlsx_xml(path: str, sheet_name: str) -> List[Dict[str, object]]:
    ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    rel_ns = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
    with zipfile.ZipFile(path) as zf:
        shared_strings: List[str] = []
        if 'xl/sharedStrings.xml' in zf.namelist():
            shared_root = ET.fromstring(zf.read('xl/sharedStrings.xml'))
            for si in shared_root.findall(f'{ns}si'):
                txt = ''.join(t.text or '' for t in si.iter(f'{ns}t'))
                shared_strings.append(txt)

        workbook_root = ET.fromstring(zf.read('xl/workbook.xml'))
        rel_root = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
        rel_map = {rel.attrib['Id']: rel.attrib['Target'] for rel in rel_root}

        target = None
        sheets = workbook_root.find(f'{ns}sheets')
        if sheets is None:
            raise RuntimeError('Workbook is missing sheets metadata.')
        for sheet in sheets:
            if sheet.attrib.get('name') == sheet_name:
                rid = sheet.attrib[f'{rel_ns}id']
                target = rel_map[rid]
                break
        if target is None:
            raise RuntimeError(f'Sheet {sheet_name!r} not found in workbook.')

        sheet_root = ET.fromstring(zf.read('xl/' + target))
        rows_xml = sheet_root.findall(f'.//{ns}sheetData/{ns}row')
        if not rows_xml:
            return []

        header_cells = rows_xml[0].findall(f'{ns}c')
        header_map: Dict[int, str] = {}
        for cell in header_cells:
            ref = cell.attrib.get('r', '')
            idx = column_ref_to_index(ref)
            value = parse_xlsx_cell_value(cell, shared_strings, ns)
            header_map[idx] = str(value).strip() if value is not None else ''

        max_idx = max(header_map) if header_map else -1
        headers = [header_map.get(i, '') for i in range(max_idx + 1)]
        data_rows: List[Dict[str, object]] = []
        for row_xml in rows_xml[1:]:
            values: List[Optional[object]] = [None] * len(headers)
            for cell in row_xml.findall(f'{ns}c'):
                ref = cell.attrib.get('r', '')
                idx = column_ref_to_index(ref)
                if idx >= len(values):
                    values.extend([None] * (idx + 1 - len(values)))
                    headers.extend([''] * (idx + 1 - len(headers)))
                values[idx] = parse_xlsx_cell_value(cell, shared_strings, ns)
            row = {headers[i]: values[i] for i in range(len(headers)) if headers[i]}
            data_rows.append(row)
    return data_rows


def load_clinical_rows(path: str, sheet_name: str) -> List[Dict[str, object]]:
    try:
        return load_rows_openpyxl(path, sheet_name)
    except Exception:
        return load_rows_xlsx_xml(path, sheet_name)


def to_float(value: object) -> float:
    if value is None or value == '':
        return math.nan
    try:
        return float(value)
    except Exception:
        return math.nan


def to_int(value: object) -> Optional[int]:
    if value is None or value == '':
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def clean_text(value: object) -> str:
    return '' if value is None else str(value).strip()


def filter_rows(rows: Sequence[Dict[str, object]], task: str) -> List[Dict[str, object]]:
    if task == 'all':
        return list(rows)
    key = f'Task {task}'
    return [row for row in rows if to_int(row.get(key)) == 1]


def stage_to_ord(stage: object, prefix: str) -> float:
    txt = clean_text(stage).upper()
    if not txt:
        return math.nan
    if txt.startswith(prefix.upper()):
        tail = txt[len(prefix):]
        digits = ''.join(ch for ch in tail if ch.isdigit())
        if digits:
            return float(int(digits))
        if tail.startswith('X'):
            return math.nan
    return math.nan


def map_binary(value: object) -> float:
    code = to_int(value)
    if code in (0, 1):
        return float(code)
    return math.nan


def build_clinical_matrix(row: Dict[str, object]) -> np.ndarray:
    return np.array([
        to_float(row.get('Age')),
        map_binary(row.get('Gender')),
        map_binary(row.get('Tobacco Consumption')),
        map_binary(row.get('Alcohol Consumption')),
        to_float(row.get('Performance Status')),
        map_binary(row.get('Treatment')),
        stage_to_ord(row.get('T-stage'), 'T'),
        stage_to_ord(row.get('N-stage'), 'N'),
        stage_to_ord(row.get('M-stage'), 'M'),
        map_binary(row.get('HPV Status')),
        to_float(row.get('CenterID')),
    ], dtype=float)


def collect_case_ids(data_dir: str, require_seg: bool = True) -> List[str]:
    out: List[str] = []
    for ct_path in sorted(glob.glob(os.path.join(data_dir, '*_CT.nii.gz'))):
        case_id = os.path.basename(ct_path)[:-len('_CT.nii.gz')]
        pet_path = os.path.join(data_dir, f'{case_id}_PET.nii.gz')
        seg_path = os.path.join(data_dir, f'{case_id}_SEG.nii.gz')
        if os.path.exists(pet_path) and ((not require_seg) or os.path.exists(seg_path)):
            out.append(case_id)
    return out


def cox_sort_by_time(time, event, X=None):
    idx = np.argsort(time)
    time_s = time[idx]
    event_s = event[idx]
    if X is None:
        return time_s, event_s, None, idx
    return time_s, event_s, X[idx, ...], idx


def cox_partial_grad_hess(beta, X, time, event, l2=0.0):
    if X.ndim == 1:
        X = X[:, None]
    n, p = X.shape
    xb = X @ beta
    r = np.exp(np.clip(xb, -50.0, 50.0))
    order = np.argsort(time)
    time = time[order]
    event = event[order]
    X = X[order, :]
    r = r[order]
    cr = np.cumsum(r[::-1])[::-1]
    cX = np.cumsum((r[:, None] * X)[::-1, :], axis=0)[::-1, :]
    g = np.zeros(p)
    fisher = np.zeros((p, p))
    i = 0
    while i < n:
        t = time[i]
        j = i
        while j < n and time[j] == t:
            j += 1
        d = int(event[i:j].sum())
        if d > 0:
            s0 = cr[i]
            s1 = cX[i, :]
            mu = s1 / max(s0, 1e-12)
            xe_sum = X[i:j, :][event[i:j] == 1].sum(axis=0)
            g += xe_sum - d * mu
            X_risk = X[i:, :]
            r_risk = r[i:]
            s2 = (X_risk.T * r_risk) @ X_risk / max(s0, 1e-12)
            fisher += d * (s2 - np.outer(mu, mu))
        i = j
    if l2 > 0:
        g -= l2 * beta
        fisher += l2 * np.eye(p)
    return g, fisher


def cox_fit(X, time, event, l2=1e-4, max_iter=60, tol=1e-6):
    X = np.asarray(X, dtype=float)
    time = np.asarray(time, dtype=float)
    event = np.asarray(event, dtype=int)
    if X.ndim == 1:
        X = X[:, None]
    mask = np.isfinite(time) & np.isfinite(event) & np.all(np.isfinite(X), axis=1)
    X, time, event = X[mask], time[mask], event[mask]
    if X.shape[0] < 10 or event.sum() < 5:
        p = X.shape[1]
        return np.zeros(p), np.full(p, np.inf), np.zeros(p), np.ones(p)
    time_s, event_s, X_s, _ = cox_sort_by_time(time, event, X)
    p = X_s.shape[1]
    beta = np.zeros(p)
    l2_used = float(l2)
    for _ in range(max_iter):
        g, fisher = cox_partial_grad_hess(beta, X_s, time_s, event_s, l2=l2_used)
        try:
            cond = np.linalg.cond(fisher)
        except np.linalg.LinAlgError:
            cond = np.inf
        if not np.isfinite(cond) or cond > 1e8:
            fisher = fisher + max(1e-6, l2_used) * np.eye(p)
            l2_used = min(max(l2_used * 10.0, 1e-6), 1.0)
        try:
            step = np.linalg.solve(fisher, g)
        except np.linalg.LinAlgError:
            step = np.linalg.pinv(fisher) @ g
        beta = beta + step
        if float(np.max(np.abs(step))) < tol:
            break
    _, fisher_final = cox_partial_grad_hess(beta, X_s, time_s, event_s, l2=l2_used)
    try:
        var = np.linalg.inv(fisher_final)
    except np.linalg.LinAlgError:
        var = np.linalg.pinv(fisher_final)
    diagv = np.diag(var)
    diagv = np.where(diagv > 1e-12, diagv, 1e-12)
    se = np.sqrt(diagv)
    z = beta / se
    try:
        from scipy.special import ndtr
        pvals = 2.0 * (1.0 - ndtr(np.abs(z)))
    except Exception:
        from math import erf, sqrt
        ndtr_vec = np.vectorize(lambda x: 0.5 * (1.0 + erf(x / sqrt(2.0))))
        pvals = 2.0 * (1.0 - ndtr_vec(np.abs(z)))
    return beta, se, z, pvals


def benjamini_hochberg(pvals, q=0.05):
    p = np.asarray(pvals, dtype=float)
    valid = np.isfinite(p)
    m = int(valid.sum())
    sig = np.zeros_like(p, dtype=bool)
    qvals = np.full_like(p, np.nan, dtype=float)
    if m == 0:
        return sig, qvals
    p_valid = p[valid]
    order = np.argsort(p_valid)
    ranked = p_valid[order]
    thresh = q * (np.arange(1, m + 1) / m)
    passed = ranked <= thresh
    sig_valid = np.zeros_like(p_valid, dtype=bool)
    if passed.any():
        cutoff = ranked[int(np.max(np.where(passed)[0]))]
        sig_valid = p_valid <= cutoff
    q_tmp = ranked * m / np.arange(1, m + 1)
    q_tmp = np.minimum.accumulate(q_tmp[::-1])[::-1]
    qvals_valid = np.empty_like(p_valid)
    qvals_valid[order] = np.clip(q_tmp, 0, 1)
    qvals[valid] = qvals_valid
    sig[valid] = sig_valid
    return sig, qvals


def concordance_index(time, event, risk):
    n = len(time)
    num = 0.0
    den = 0.0
    for i in range(n):
        for j in range(n):
            if time[i] < time[j] and event[i] == 1:
                den += 1.0
                if risk[i] > risk[j]:
                    num += 1.0
                elif risk[i] == risk[j]:
                    num += 0.5
    return num / den if den > 0 else np.nan


def zscore_matrix(A: np.ndarray):
    A = np.asarray(A, dtype=float)
    finite_mask = np.isfinite(A)
    counts = finite_mask.sum(axis=0)
    safe = np.where(finite_mask, A, 0.0)
    denom = np.maximum(counts, 1)
    means = safe.sum(axis=0) / denom
    centered = np.where(finite_mask, A - means, 0.0)
    stds = np.sqrt((centered ** 2).sum(axis=0) / denom)
    means = np.where(counts > 0, means, 0.0)
    stds = np.where(np.isfinite(stds) & (stds >= 1e-9), stds, 1.0)
    filled = np.where(finite_mask, A, means)
    Z = (filled - means) / stds
    if Z.ndim == 2:
        Z[:, counts == 0] = 0.0
    return Z, means, stds


def plot_km_with_ticks_two_groups(
    time_days: np.ndarray,
    event: np.ndarray,
    grp_high: np.ndarray,
    out_path: str,
    title: str,
    censor_merge_window_days: float = 45.0,
):
    def merge_censor_ticks(times, levels, merge_window_days):
        times = np.asarray(times, dtype=float)
        levels = np.asarray(levels, dtype=float)
        if times.size == 0 or merge_window_days is None or merge_window_days <= 0:
            return times, levels
        order = np.argsort(times)
        times = times[order]
        levels = levels[order]
        merged_times = []
        merged_levels = []
        cluster_times = [float(times[0])]
        cluster_levels = [float(levels[0])]
        for idx in range(1, times.size):
            if float(times[idx]) - cluster_times[-1] <= float(merge_window_days):
                cluster_times.append(float(times[idx]))
                cluster_levels.append(float(levels[idx]))
            else:
                merged_times.append(float(np.mean(cluster_times)))
                merged_levels.append(float(np.mean(cluster_levels)))
                cluster_times = [float(times[idx])]
                cluster_levels = [float(levels[idx])]
        merged_times.append(float(np.mean(cluster_times)))
        merged_levels.append(float(np.mean(cluster_levels)))
        return np.asarray(merged_times, dtype=float), np.asarray(merged_levels, dtype=float)

    def km_curve(t, e):
        order = np.argsort(t)
        t = np.asarray(t, dtype=float)[order]
        e = np.asarray(e, dtype=int)[order]
        uniq_times = np.unique(t)
        n_at_risk = float(len(t))
        surv = 1.0
        xs = [0.0]
        ys = [1.0]
        censor_times = []
        censor_levels = []
        ci_x = [0.0]
        ci_lo = [1.0]
        ci_hi = [1.0]
        var_acc = 0.0
        for ut in uniq_times:
            mask_ut = (t == ut)
            d = float((e[mask_ut] == 1).sum())
            c = float((e[mask_ut] == 0).sum())
            if n_at_risk > 0 and d > 0:
                surv *= (1.0 - d / n_at_risk)
                xs.extend([ut, ut])
                ys.extend([ys[-1], surv])
                var_acc += d / (n_at_risk * max(n_at_risk - d, 1.0))
                se = (surv ** 2) * var_acc
                delta = 1.96 * np.sqrt(max(se, 0.0))
                lo = max(0.0, surv - delta)
                hi = min(1.0, surv + delta)
                ci_x.extend([ut, ut])
                ci_lo.extend([ci_lo[-1], lo])
                ci_hi.extend([ci_hi[-1], hi])
            if c > 0:
                censor_times.extend([ut] * int(c))
                censor_levels.extend([surv] * int(c))
            n_at_risk -= (d + c)
        censor_times, censor_levels = merge_censor_ticks(censor_times, censor_levels, censor_merge_window_days)
        return np.asarray(xs), np.asarray(ys), np.asarray(censor_times), np.asarray(censor_levels), np.asarray(ci_x), np.asarray(ci_lo), np.asarray(ci_hi)

    mask_fin = np.isfinite(time_days) & np.isfinite(event)
    t = np.asarray(time_days, dtype=float)[mask_fin]
    e = np.asarray(event, dtype=int)[mask_fin]
    grp_high = np.asarray(grp_high, dtype=bool)[mask_fin]
    if grp_high.sum() == 0 or (~grp_high).sum() == 0:
        return
    xh, yh, cth, cyh, cixh, ciloh, cihih = km_curve(t[grp_high], e[grp_high])
    xl, yl, ctl, cyl, cixl, cilol, cihil = km_curve(t[~grp_high], e[~grp_high])
    plt.figure(figsize=(6.5, 4.5), dpi=140)
    plt.step(xh, yh, where='post', color='tab:red', label='High group')
    plt.step(xl, yl, where='post', color='tab:blue', label='Low group')
    if cixh.size:
        plt.fill_between(cixh, ciloh, cihih, color='tab:red', alpha=0.15, step='post')
    if cixl.size:
        plt.fill_between(cixl, cilol, cihil, color='tab:blue', alpha=0.15, step='post')
    if cth.size:
        plt.scatter(cth, cyh, marker='|', color='tab:red', s=80, linewidths=1.5)
    if ctl.size:
        plt.scatter(ctl, cyl, marker='|', color='tab:blue', s=80, linewidths=1.5)
    plt.ylim(0.0, 1.05)
    plt.xlabel('Time (days)')
    plt.ylabel('Relapse-free survival probability')
    plt.title(title)
    plt.grid(alpha=0.25)
    plt.legend(loc='best')
    try:
        plt.savefig(out_path, bbox_inches='tight')
    finally:
        plt.close()


def sanitize_mask(mask: np.ndarray, lesion_min_voxels: int) -> Tuple[np.ndarray, int]:
    mask = np.asarray(mask > 0.5, dtype=bool)
    if not mask.any():
        return mask, 0
    structure = np.ones((3, 3, 3), dtype=np.uint8)
    labeled, num = ndimage.label(mask.astype(np.uint8), structure=structure)
    if num <= 0:
        return np.zeros_like(mask, dtype=bool), 0
    sizes = ndimage.sum(np.ones_like(mask, dtype=np.float32), labels=labeled, index=np.arange(1, num + 1))
    keep_labels = np.nonzero(np.asarray(sizes) >= float(max(1, lesion_min_voxels)))[0] + 1
    if keep_labels.size == 0:
        return np.zeros_like(mask, dtype=bool), 0
    keep_mask = np.isin(labeled, keep_labels)
    return keep_mask.astype(bool), int(keep_labels.size)


def extract_tumor_statistics(
    pet_arr_zyx: np.ndarray,
    ct_arr_zyx: np.ndarray,
    mask_zyx: np.ndarray,
    spacing_xyz: Tuple[float, float, float],
) -> Tuple[np.ndarray, List[str]]:
    spacing_xyz = tuple(float(v) for v in spacing_xyz)
    voxel_volume_mm3 = float(np.prod(spacing_xyz))
    coords = np.argwhere(mask_zyx)
    pet_vals = pet_arr_zyx[mask_zyx]
    ct_vals = ct_arr_zyx[mask_zyx]
    extents_zyx = (coords.max(axis=0) - coords.min(axis=0) + 1).astype(float)
    centroid_zyx = coords.mean(axis=0).astype(float)
    spacing_zyx = np.array([spacing_xyz[2], spacing_xyz[1], spacing_xyz[0]], dtype=float)
    extent_mm_zyx = extents_zyx * spacing_zyx
    centroid_mm_zyx = centroid_zyx * spacing_zyx
    names = [
        'tumor_mtv_mm3',
        'tumor_tla_like',
        'tumor_lesion_count',
        'tumor_voxel_count',
        'tumor_suv_mean',
        'tumor_suv_max',
        'tumor_suv_p90',
        'tumor_suv_p95',
        'tumor_suv_std',
        'tumor_ct_mean',
        'tumor_ct_min',
        'tumor_ct_p10',
        'tumor_centroid_x_mm',
        'tumor_centroid_y_mm',
        'tumor_centroid_z_mm',
        'tumor_extent_x_mm',
        'tumor_extent_y_mm',
        'tumor_extent_z_mm',
    ]
    values = np.array([
        float(mask_zyx.sum()) * voxel_volume_mm3,
        float(np.nansum(pet_vals)) * voxel_volume_mm3,
        math.nan,
        float(mask_zyx.sum()),
        float(np.nanmean(pet_vals)),
        float(np.nanmax(pet_vals)),
        float(np.nanpercentile(pet_vals, 90.0)),
        float(np.nanpercentile(pet_vals, 95.0)),
        float(np.nanstd(pet_vals)),
        float(np.nanmean(ct_vals)),
        float(np.nanmin(ct_vals)),
        float(np.nanpercentile(ct_vals, 10.0)),
        float(centroid_mm_zyx[2]),
        float(centroid_mm_zyx[1]),
        float(centroid_mm_zyx[0]),
        float(extent_mm_zyx[2]),
        float(extent_mm_zyx[1]),
        float(extent_mm_zyx[0]),
    ], dtype=float)
    return values, names


def build_radiomics_extractor(bin_width: float, feature_classes: Sequence[str], force_2d: bool):
    settings = {
        'binWidth': float(bin_width),
        'normalize': False,
        'removeOutliers': None,
        'force2D': bool(force_2d),
        'label': 1,
        'additionalInfo': False,
    }
    extractor = featureextractor.RadiomicsFeatureExtractor(**settings)
    extractor.disableAllImageTypes()
    for image_type in RADIOMICS_IMAGE_TYPES:
        extractor.enableImageTypeByName(image_type)
    extractor.disableAllFeatures()
    for feature_class in feature_classes:
        extractor.enableFeatureClassByName(feature_class)
    return extractor


def sitk_mask_from_array(mask_zyx: np.ndarray, reference_image: sitk.Image) -> sitk.Image:
    mask_image = sitk.GetImageFromArray(mask_zyx.astype(np.uint8))
    mask_image.CopyInformation(reference_image)
    return mask_image


def extract_radiomics_features(
    image: sitk.Image,
    mask_zyx: np.ndarray,
    extractor,
    prefix: str,
) -> Tuple[np.ndarray, List[str]]:
    mask_image = sitk_mask_from_array(mask_zyx, image)
    raw_features = extractor.execute(image, mask_image)
    names: List[str] = []
    values: List[float] = []
    for key in sorted(raw_features):
        if key.startswith('diagnostics_'):
            continue
        value = raw_features[key]
        try:
            value_float = float(value)
        except Exception:
            continue
        names.append(f'{prefix}_{key}')
        values.append(value_float)
    return np.asarray(values, dtype=float), names


def compute_univariate_feature_stats(
    X: np.ndarray,
    feature_names: Sequence[str],
    time_np: np.ndarray,
    event_np: np.ndarray,
    min_coverage: int,
    min_events: int,
) -> Dict[str, np.ndarray]:
    feat_beta = []
    feat_z = []
    feat_p = []
    feat_cov = []
    feat_evt = []
    feat_std = []
    for j in range(X.shape[1]):
        col = X[:, j]
        mfin = np.isfinite(col)
        n_cov = int(mfin.sum())
        e_cov = int(event_np[mfin].sum())
        std_val = float(np.nanstd(col)) if n_cov > 0 else math.nan
        feat_cov.append(n_cov)
        feat_evt.append(e_cov)
        feat_std.append(std_val)
        if n_cov < min_coverage or e_cov < min_events or (not np.isfinite(std_val)) or std_val < 1e-9:
            feat_beta.append(0.0)
            feat_z.append(0.0)
            feat_p.append(1.0)
            continue
        mean_val = float(np.nanmean(col))
        xz = (np.where(np.isfinite(col), col, mean_val) - mean_val) / (std_val + 1e-9)
        xz = np.clip(xz, -8.0, 8.0)
        beta, _, z_val, p_val = cox_fit(xz, time_np, event_np, l2=1e-4, max_iter=60)
        feat_beta.append(float(np.atleast_1d(beta)[0]))
        feat_z.append(float(np.atleast_1d(z_val)[0]))
        feat_p.append(float(np.atleast_1d(p_val)[0]))
    feat_beta = np.asarray(feat_beta, dtype=float)
    feat_z = np.asarray(feat_z, dtype=float)
    feat_p = np.asarray(feat_p, dtype=float)
    feat_cov = np.asarray(feat_cov, dtype=int)
    feat_evt = np.asarray(feat_evt, dtype=int)
    feat_std = np.asarray(feat_std, dtype=float)
    sig05, q05 = benjamini_hochberg(feat_p, q=0.05)
    sig10, q10 = benjamini_hochberg(feat_p, q=0.10)
    return {
        'feature_names': np.asarray(list(feature_names), dtype=object),
        'beta': feat_beta,
        'z': feat_z,
        'p': feat_p,
        'q05': q05,
        'q10': q10,
        'sig05': sig05,
        'sig10': sig10,
        'coverage': feat_cov,
        'events': feat_evt,
        'std': feat_std,
    }


def greedy_correlation_prune(
    X: np.ndarray,
    ranked_indices: np.ndarray,
    threshold: float,
) -> np.ndarray:
    if ranked_indices.size == 0:
        return ranked_indices
    if threshold is None or threshold >= 1.0 or X.ndim != 2 or X.shape[0] < 2 or X.shape[1] < 2:
        return ranked_indices.copy()
    Xz, _, _ = zscore_matrix(X)
    corr = np.corrcoef(Xz, rowvar=False)
    if np.ndim(corr) != 2:
        return ranked_indices.copy()
    corr = np.nan_to_num(corr, nan=0.0, posinf=0.0, neginf=0.0)
    keep: List[int] = []
    dropped = np.zeros(X.shape[1], dtype=bool)
    for idx in ranked_indices:
        idx = int(idx)
        if dropped[idx]:
            continue
        keep.append(idx)
        correlated = np.abs(corr[idx]) >= float(threshold)
        correlated[idx] = False
        dropped |= correlated
    return np.asarray(keep, dtype=int)


def select_feature_indices(
    stats: Dict[str, np.ndarray],
    feature_selection: str,
    topk: int,
    X: np.ndarray,
    drop_correlated: bool,
    correlation_threshold: float,
    max_selected_features: int,
) -> Tuple[np.ndarray, np.ndarray, str]:
    def _cap_selected(selected_indices: np.ndarray, selection_label: str) -> Tuple[np.ndarray, str]:
        if int(max_selected_features) <= 0 or selected_indices.size <= int(max_selected_features):
            return selected_indices, selection_label
        capped = selected_indices[:int(max_selected_features)]
        return capped, f'{selection_label}+cap({int(max_selected_features)})'

    feat_p = stats['p']
    q05 = stats['q05']
    q10 = stats['q10']
    finite_idx = np.where(np.isfinite(feat_p))[0]
    ranked = finite_idx[np.argsort(feat_p[finite_idx])]
    pruned_ranked = greedy_correlation_prune(X, ranked, correlation_threshold) if drop_correlated else ranked
    if feature_selection == 'conventional':
        selected_q05 = np.asarray([idx for idx in pruned_ranked if np.isfinite(q05[idx]) and q05[idx] <= 0.05], dtype=int)
        selected_q10 = np.asarray([idx for idx in pruned_ranked if np.isfinite(q10[idx]) and q10[idx] <= 0.10], dtype=int)
        if selected_q05.size > 0:
            selected = selected_q05
            selection_used = 'conventional:fdr05'
        elif selected_q10.size > 0:
            selected = selected_q10
            selection_used = 'conventional:fdr10'
        else:
            selected = pruned_ranked[:min(int(topk), pruned_ranked.size)] if int(topk) > 0 else pruned_ranked
            selection_used = f'conventional:topk_fallback(k={int(topk)})'
    elif feature_selection == 'all':
        selected = pruned_ranked
        selection_used = 'all'
    elif feature_selection == 'topk':
        selected = pruned_ranked[:min(int(topk), pruned_ranked.size)] if int(topk) > 0 else pruned_ranked
        selection_used = f'topk(k={int(topk)})'
    elif feature_selection == 'fdr05':
        selected = np.asarray([idx for idx in pruned_ranked if np.isfinite(q05[idx]) and q05[idx] <= 0.05], dtype=int)
        selection_used = 'fdr05'
    elif feature_selection == 'fdr10':
        selected = np.asarray([idx for idx in pruned_ranked if np.isfinite(q10[idx]) and q10[idx] <= 0.10], dtype=int)
        selection_used = 'fdr10'
    else:
        raise ValueError(f'Unsupported feature selection mode: {feature_selection}')
    selected, selection_used = _cap_selected(selected, selection_used)
    return selected, pruned_ranked, selection_used


def write_case_feature_table(
    output_csv: str,
    patient_ids: Sequence[str],
    time_np: np.ndarray,
    event_np: np.ndarray,
    Xclin: np.ndarray,
    Ximg: np.ndarray,
    imaging_names: Sequence[str],
) -> None:
    clinical_names = [
        'Age', 'Gender', 'Tobacco Consumption', 'Alcohol Consumption', 'Performance Status',
        'Treatment', 'T-stage', 'N-stage', 'M-stage', 'HPV Status', 'CenterID'
    ]
    with open(output_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['PatientID', 'RFS', 'Relapse'] + clinical_names + list(imaging_names))
        for i, patient_id in enumerate(patient_ids):
            writer.writerow([
                patient_id,
                float(time_np[i]),
                int(event_np[i]),
                *[float(x) if np.isfinite(x) else '' for x in Xclin[i]],
                *[float(x) if np.isfinite(x) else '' for x in Ximg[i]],
            ])


def main() -> None:
    args = parse_args()
    if not (args.use_tumor_stats or args.use_pet_radiomics or args.use_ct_radiomics):
        raise RuntimeError('Enable at least one of tumor stats, PET radiomics, or CT radiomics.')

    os.makedirs(args.output_dir, exist_ok=True)

    rows = filter_rows(load_clinical_rows(args.clinical_xlsx, args.clinical_sheet), args.task)
    clinical_by_id = {clean_text(row.get('PatientID')): row for row in rows if clean_text(row.get('PatientID'))}

    case_ids_all = collect_case_ids(args.preprocessed_dir, require_seg=True)
    case_ids = [case_id for case_id in case_ids_all if case_id in clinical_by_id]
    if int(args.max_cases) > 0:
        case_ids = case_ids[:int(args.max_cases)]
    if not case_ids:
        raise RuntimeError('No HECKTOR cases matched between preprocessed images, segmentations, and workbook.')

    print(f'Clinical rows after task filter: {len(rows)}')
    print(f'Cases with CT/PET/SEG and clinical data: {len(case_ids)}')

    pet_extractor = None
    ct_extractor = None
    if args.use_pet_radiomics:
        pet_extractor = build_radiomics_extractor(args.pet_bin_width, PET_RADIOMICS_CLASSES, args.force_2d)
    if args.use_ct_radiomics:
        ct_extractor = build_radiomics_extractor(args.ct_bin_width, CT_RADIOMICS_CLASSES, args.force_2d)

    patient_ids: List[str] = []
    clinical_rows_list: List[np.ndarray] = []
    imaging_feature_rows: List[np.ndarray] = []
    imaging_feature_names: Optional[List[str]] = None
    time_to_event: List[float] = []
    event_indicator: List[int] = []
    lesion_counts: List[int] = []
    skipped_cases: List[Tuple[str, str]] = []

    for case_index, case_id in enumerate(case_ids, start=1):
        row = clinical_by_id.get(case_id)
        if row is None:
            continue
        t_relapse = to_float(row.get('RFS'))
        e_relapse = to_int(row.get('Relapse'))
        if not np.isfinite(t_relapse) or e_relapse not in (0, 1):
            print(f'Skip {case_id}: invalid RFS/Relapse outcome.')
            skipped_cases.append((case_id, 'invalid outcome'))
            continue

        ct_path = os.path.join(args.preprocessed_dir, f'{case_id}_CT.nii.gz')
        pet_path = os.path.join(args.preprocessed_dir, f'{case_id}_PET.nii.gz')
        seg_path = os.path.join(args.preprocessed_dir, f'{case_id}_SEG.nii.gz')
        try:
            ct_img = sitk.ReadImage(ct_path)
            pet_img = sitk.ReadImage(pet_path)
            seg_img = sitk.ReadImage(seg_path)
        except Exception as exc:
            print(f'Skip {case_id}: read failure: {exc}')
            skipped_cases.append((case_id, f'read failure: {exc}'))
            continue

        pet_arr = sitk.GetArrayFromImage(pet_img).astype(np.float32)
        ct_arr = sitk.GetArrayFromImage(ct_img).astype(np.float32)
        seg_arr = sitk.GetArrayFromImage(seg_img).astype(np.float32)
        if pet_arr.shape != ct_arr.shape or seg_arr.shape != ct_arr.shape:
            print(f'Skip {case_id}: shape mismatch among CT/PET/SEG')
            skipped_cases.append((case_id, 'shape mismatch among CT/PET/SEG'))
            continue

        mask_arr, lesion_count = sanitize_mask(seg_arr, args.lesion_min_voxels)
        if not mask_arr.any():
            print(f'Skip {case_id}: empty tumor mask after lesion filtering')
            skipped_cases.append((case_id, 'empty tumor mask after lesion filtering'))
            continue

        case_parts: List[np.ndarray] = []
        case_names: List[str] = []

        if args.use_tumor_stats:
            stat_values, stat_names = extract_tumor_statistics(pet_arr, ct_arr, mask_arr, pet_img.GetSpacing())
            if stat_values.size >= 3:
                stat_values[2] = float(lesion_count)
            case_parts.append(stat_values)
            case_names.extend(stat_names)

        if args.use_pet_radiomics and pet_extractor is not None:
            try:
                pet_values, pet_names = extract_radiomics_features(pet_img, mask_arr, pet_extractor, 'pet')
            except Exception as exc:
                print(f'Skip {case_id}: PET radiomics failure: {exc}')
                skipped_cases.append((case_id, f'PET radiomics failure: {exc}'))
                continue
            case_parts.append(pet_values)
            case_names.extend(pet_names)

        if args.use_ct_radiomics and ct_extractor is not None:
            try:
                ct_values, ct_names = extract_radiomics_features(ct_img, mask_arr, ct_extractor, 'ct')
            except Exception as exc:
                print(f'Skip {case_id}: CT radiomics failure: {exc}')
                skipped_cases.append((case_id, f'CT radiomics failure: {exc}'))
                continue
            case_parts.append(ct_values)
            case_names.extend(ct_names)

        if imaging_feature_names is None:
            imaging_feature_names = list(case_names)
        elif list(case_names) != imaging_feature_names:
            print(f'Skip {case_id}: inconsistent feature layout across cases')
            skipped_cases.append((case_id, 'inconsistent feature layout across cases'))
            continue

        patient_ids.append(case_id)
        time_to_event.append(float(t_relapse))
        event_indicator.append(int(e_relapse))
        clinical_rows_list.append(build_clinical_matrix(row))
        imaging_feature_rows.append(np.concatenate(case_parts, axis=0) if case_parts else np.zeros((0,), dtype=float))
        lesion_counts.append(int(lesion_count))
        print(
            f'Processed {case_index}/{len(case_ids)}: {case_id} | '
            f'RFS={t_relapse} days | event={e_relapse} | lesions={lesion_count} | features={len(case_names)}'
        )

    if not imaging_feature_rows:
        raise RuntimeError('No analyzable cases remained after segmentation/radiomics filtering.')

    Xclin = np.stack(clinical_rows_list, axis=0)
    Ximg = np.vstack(imaging_feature_rows)
    time_np = np.asarray(time_to_event, dtype=float)
    event_np = np.asarray(event_indicator, dtype=int)
    feature_names = imaging_feature_names or []

    print(f'Dataset for analysis: N={Ximg.shape[0]}, Features={Ximg.shape[1]}, Events={int(event_np.sum())}')

    case_feature_csv = os.path.join(args.output_dir, 'tumor_radiomics_case_features.csv')
    write_case_feature_table(case_feature_csv, patient_ids, time_np, event_np, Xclin, Ximg, feature_names)

    if skipped_cases:
        skip_csv = os.path.join(args.output_dir, 'skipped_cases.csv')
        with open(skip_csv, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['PatientID', 'reason'])
            writer.writerows(skipped_cases)

    Xclin_z, _, _ = zscore_matrix(Xclin)
    b_clin, _, _, _ = cox_fit(Xclin_z, time_np, event_np, l2=1e-4, max_iter=60)
    risk_clin = Xclin_z @ np.nan_to_num(b_clin, nan=0.0, posinf=0.0, neginf=0.0)
    c_clin = concordance_index(time_np, event_np, risk_clin)
    plot_km_with_ticks_two_groups(
        time_np,
        event_np,
        risk_clin >= np.nanmedian(risk_clin),
        os.path.join(args.output_dir, 'km_clinical_median.png'),
        'KM: clinical model (median split)',
    )

    imaging_stats = compute_univariate_feature_stats(
        Ximg,
        feature_names,
        time_np,
        event_np,
        args.min_coverage,
        args.min_events,
    )

    feat_csv = os.path.join(args.output_dir, 'tumor_radiomics_univariate_cox.csv')
    with open(feat_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['feature', 'beta', 'HR', 'z', 'p', 'q05', 'q10', 'coverage', 'events', 'std'])
        for idx, name in enumerate(feature_names):
            beta = float(imaging_stats['beta'][idx])
            hr = np.exp(np.clip(beta, -50.0, 50.0)) if np.isfinite(beta) else np.nan
            writer.writerow([
                name,
                beta,
                hr,
                float(imaging_stats['z'][idx]),
                float(imaging_stats['p'][idx]),
                float(imaging_stats['q05'][idx]),
                float(imaging_stats['q10'][idx]),
                int(imaging_stats['coverage'][idx]),
                int(imaging_stats['events'][idx]),
                float(imaging_stats['std'][idx]) if np.isfinite(imaging_stats['std'][idx]) else '',
            ])

    selected_indices, pruned_ranked, selection_used = select_feature_indices(
        imaging_stats,
        args.feature_selection,
        args.topk,
        Ximg,
        args.drop_correlated,
        args.correlation_threshold,
        args.max_selected_features,
    )

    Xfeat_z, _, _ = zscore_matrix(Ximg[:, selected_indices]) if selected_indices.size > 0 else (np.zeros((Ximg.shape[0], 0)), None, None)
    X_clin_only = np.nan_to_num(Xclin_z, nan=0.0, posinf=0.0, neginf=0.0)
    X_feat_only = np.nan_to_num(Xfeat_z, nan=0.0, posinf=0.0, neginf=0.0)
    X_with_feats = np.hstack([X_clin_only, X_feat_only]) if X_feat_only.shape[1] > 0 else X_clin_only

    b_cf, _, _, _ = cox_fit(X_with_feats, time_np, event_np, l2=1e-3, max_iter=80)
    risk_all = X_with_feats @ np.nan_to_num(b_cf, nan=0.0, posinf=0.0, neginf=0.0)
    c_all = concordance_index(time_np, event_np, risk_all)

    if X_feat_only.shape[1] > 0:
        b_feat, _, _, _ = cox_fit(X_feat_only, time_np, event_np, l2=1e-3, max_iter=80)
        risk_feat = X_feat_only @ np.nan_to_num(b_feat, nan=0.0, posinf=0.0, neginf=0.0)
        c_feat = concordance_index(time_np, event_np, risk_feat)
        plot_km_with_ticks_two_groups(
            time_np,
            event_np,
            risk_feat >= np.nanmedian(risk_feat),
            os.path.join(args.output_dir, 'km_imaging_median.png'),
            'KM: tumor stats + radiomics (median split)',
        )
    else:
        risk_feat = np.zeros_like(time_np)
        c_feat = np.nan

    plot_km_with_ticks_two_groups(
        time_np,
        event_np,
        risk_all >= np.nanmedian(risk_all),
        os.path.join(args.output_dir, 'km_clinical_plus_imaging_median.png'),
        'KM: clinical + tumor radiomics (median split)',
    )

    selected_csv = os.path.join(args.output_dir, 'selected_imaging_features.csv')
    with open(selected_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['rank', 'feature_index', 'feature_name', 'beta', 'HR', 'p', 'q05', 'q10'])
        for rank, idx in enumerate(selected_indices, start=1):
            beta = float(imaging_stats['beta'][idx])
            writer.writerow([
                rank,
                int(idx),
                feature_names[int(idx)],
                beta,
                np.exp(np.clip(beta, -50.0, 50.0)) if np.isfinite(beta) else np.nan,
                float(imaging_stats['p'][idx]),
                float(imaging_stats['q05'][idx]),
                float(imaging_stats['q10'][idx]),
            ])

    with open(os.path.join(args.output_dir, 'cox_summary.txt'), 'w') as f:
        f.write(f'Task filter={args.task}\n')
        f.write(f'Patients N={Ximg.shape[0]}, Imaging features={Ximg.shape[1]}\n')
        f.write(f'Events={int(event_np.sum())}, Median RFS={np.nanmedian(time_np):.1f} days\n')
        f.write(f'Tumor stats enabled={args.use_tumor_stats}\n')
        f.write(f'PET radiomics enabled={args.use_pet_radiomics}\n')
        f.write(f'CT radiomics enabled={args.use_ct_radiomics}\n')
        f.write(f'PET bin width={args.pet_bin_width}\n')
        f.write(f'CT bin width={args.ct_bin_width}\n')
        f.write(f'Force 2D radiomics={args.force_2d}\n')
        f.write(f'Greedy correlation pruning enabled={args.drop_correlated}\n')
        f.write(f'Correlation threshold={args.correlation_threshold:.3f}\n')
        f.write(f'Max selected imaging features={args.max_selected_features}\n')
        f.write(f'Min lesion voxels={args.lesion_min_voxels}\n')
        if lesion_counts:
            f.write(f'Median retained lesion count={np.median(np.asarray(lesion_counts, dtype=float)):.1f}\n')
        f.write(f'Clinical-only C-index={c_clin:.3f}\n')
        if args.feature_selection == 'topk':
            selection_desc = f'topk (k={args.topk}, selected={selected_indices.size})'
        elif args.feature_selection == 'conventional':
            selection_desc = f'conventional [{selection_used}] (selected={selected_indices.size})'
        else:
            selection_desc = f'{args.feature_selection} (selected={selected_indices.size})'
        f.write(f'Feature selection mode={selection_desc}\n')
        f.write(f'Correlation-pruned ranked candidates={pruned_ranked.size}\n')
        f.write(f'Clinical+imaging C-index={c_all:.3f}\n')
        f.write(f'Imaging-only C-index={c_feat:.3f}\n')
        f.write('Clinical covariates included: age, gender, tobacco, alcohol, performance status, treatment, T/N/M stage, HPV, CenterID.\n')
        if selected_indices.size > 0:
            selected_names = [feature_names[int(idx)] for idx in selected_indices]
            f.write('Selected imaging features=' + ', '.join(selected_names) + '\n')
        if skipped_cases:
            f.write(f'Skipped cases={len(skipped_cases)}\n')

    print('Saved outputs to:', args.output_dir)
    print(f'Clinical-only C-index={c_clin:.3f}')
    print(f'Clinical + imaging C-index={c_all:.3f}')
    print(f'Imaging-only C-index={c_feat:.3f}')
    print(f'Saved case feature table: {case_feature_csv}')
    print(f'Saved feature table: {feat_csv}')


if __name__ == '__main__':
    main()
