import argparse
import csv
import glob
import math
import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Sequence, Tuple

import matplotlib.pyplot as plt
import nibabel as nib
import numpy as np
import SimpleITK as sitk
import torch
import torch.nn.functional as F
from scipy import ndimage
from torch.utils.data import DataLoader, Dataset

try:
    from radiomics import featureextractor
except Exception:
    featureextractor = None

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
TUTORIAL_DIR = os.path.abspath(os.path.join(THIS_DIR, '..'))
REPO_ROOT = os.path.abspath(os.path.join(THIS_DIR, '..', '..', '..'))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from MIR.models import TemplateCreation, SpatialTransformer, EncoderFeatureExtractor, SITReg
from MIR.models.SITReg import ReLUFactory, GroupNormalizerFactory
from MIR.models.SITReg.composable_mapping import DataFormat
from MIR.models.SITReg.deformation_inversion_layer.fixed_point_iteration import (
    AndersonSolver,
    AndersonSolverArguments,
    MaxElementWiseAbsStopCriterion,
    RelativeL2ErrorStopCriterion,
)


DEFAULT_PREPROCESSED_DIR = '/scratch2/jchen/DATA/HECKTOR25/preprocessed'
DEFAULT_CLINICAL_XLSX = '/scratch2/jchen/DATA/HECKTOR25/HECKTOR_2025_Training_EHR_with_Data_Dictionary.xlsx'
DEFAULT_ATLAS_CT_DIR = '/scratch/jchen/python_projects/custom_packages/MIR/tutorials/HECKTOR25/atlas/ct/SITRegAtlas_JHU_SSIM_1_MS_1_diffusion_1'
DEFAULT_ATLAS_PET_DIR = '/scratch/jchen/python_projects/custom_packages/MIR/tutorials/HECKTOR25/atlas/suv/SITRegAtlas_JHU_SSIM_1_MS_1_diffusion_1'
DEFAULT_ATLAS_CTSEG_PATH = '/scratch/jchen/python_projects/custom_packages/MIR/tutorials/HECKTOR25/atlas/seg/SITRegAtlas_JHU_SSIM_1_MS_1_diffusion_1/ctseg_atlas_118lbls.nii.gz'
DEFAULT_CHECKPOINT_DIR = '/scratch/jchen/python_projects/custom_packages/MIR/tutorials/HECKTOR25/experiments/SITRegAtlas_JHU_SSIM_1_MS_1_diffusion_1'
DEFAULT_OUTPUT_DIR = os.path.join(THIS_DIR, 'population_stats')
INPUT_SHAPE = (192, 192, 144)
VOXEL_SPACING_MM = (2.8, 2.8, 3.8)
BLOOD_POOL_LABELS = (52, 54, 55, 56, 57, 58, 59, 60, 62)
BRAIN_LABELS = (90,)
RADIOMICS_IMAGE_TYPES = ('Original',)
PET_RADIOMICS_CLASSES = ('shape', 'firstorder', 'glcm', 'glrlm', 'glszm', 'gldm', 'ngtdm')
CT_RADIOMICS_CLASSES = ('firstorder', 'glcm', 'glrlm', 'glszm', 'gldm', 'ngtdm')


def add_bool_flag(
    parser: argparse.ArgumentParser,
    name: str,
    dest: str,
    default: bool,
    help_enabled: str,
    help_disabled: str,
    positive_aliases: Optional[Sequence[str]] = None,
    negative_aliases: Optional[Sequence[str]] = None,
) -> None:
    positive_flags = [f'--{name}'] + list(positive_aliases or [])
    negative_flags = [f'--no-{name}'] + list(negative_aliases or [])
    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        *positive_flags,
        dest=dest,
        action='store_true',
        help=help_enabled,
    )
    group.add_argument(
        *negative_flags,
        dest=dest,
        action='store_false',
        help=help_disabled,
    )
    parser.set_defaults(**{dest: default})


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Population survival analysis for HECKTOR25 atlas-space PET data.')
    parser.add_argument('--preprocessed-dir', default=DEFAULT_PREPROCESSED_DIR)
    parser.add_argument('--clinical-xlsx', default=DEFAULT_CLINICAL_XLSX)
    parser.add_argument('--clinical-sheet', default='Data')
    parser.add_argument('--atlas-ct-dir', default=DEFAULT_ATLAS_CT_DIR)
    parser.add_argument('--atlas-pet-dir', default=DEFAULT_ATLAS_PET_DIR)
    parser.add_argument('--atlas-ctseg-path', default=DEFAULT_ATLAS_CTSEG_PATH)
    parser.add_argument('--checkpoint-dir', default=DEFAULT_CHECKPOINT_DIR)
    parser.add_argument('--output-dir', default=DEFAULT_OUTPUT_DIR)
    parser.add_argument('--task', choices=['all', '1', '2', '3'], default='all')
    parser.add_argument('--max-cases', type=int, default=0, help='Limit processing to the first N matched cases for quick debugging; 0 uses all cases.')
    parser.add_argument('--batch-size', type=int, default=1)
    parser.add_argument('--num-workers', type=int, default=2)
    parser.add_argument('--device', default='cuda:0')
    parser.add_argument('--voxelwise', action='store_true', help='Run voxelwise univariate Cox maps.')
    parser.add_argument('--voxel-chunk', type=int, default=75000)
    parser.add_argument('--min-coverage', type=int, default=15)
    parser.add_argument('--min-events', type=int, default=5)
    parser.add_argument('--feature-selection', choices=['conventional', 'all', 'topk', 'fdr05', 'fdr10'], default='conventional', help='Rule used to select imaging features for the multivariable imaging score. The conventional mode uses correlation pruning plus FDR q<=0.05, then q<=0.10, then top-k fallback if needed.')
    parser.add_argument('--topk', type=int, default=12, help='Number of univariate-ranked imaging features to include when --feature-selection=topk, or as the fallback size for --feature-selection=conventional.')
    parser.add_argument('--max-selected-features', type=int, default=10, help='Maximum number of imaging features retained after ranking/selection. Set to 0 to disable capping.')
    parser.add_argument('--drop-correlation-threshold', type=float, default=0.90, help='Greedy absolute-correlation threshold used to prune redundant imaging features before multivariable score construction.')
    parser.add_argument('--smooth-sigma', type=float, default=1.0)
    parser.add_argument('--lesion-min-voxels', type=int, default=10)
    parser.add_argument('--pet-radiomics-bin-width', type=float, default=0.25)
    parser.add_argument('--ct-radiomics-bin-width', type=float, default=25.0)
    parser.add_argument('--radiomics-force-2d', action='store_true', help='Run radiomics in 2D mode instead of 3D.')
    add_bool_flag(
        parser,
        name='tumor-features',
        dest='use_tumor_features',
        default=True,
        help_enabled='Enable tumor/lesion-derived features, maps, and Z-score tumor summaries.',
        help_disabled='Disable tumor/lesion-derived features, maps, and Z-score tumor summaries.',
        positive_aliases=['--use-tumor-features', '--use-lesion-seg'],
    )
    add_bool_flag(
        parser,
        name='atlas-ctseg-features',
        dest='use_atlas_ctseg_features',
        default=True,
        help_enabled='Enable atlas-region PET/CT biomarker features.',
        help_disabled='Disable atlas-region PET/CT biomarker features.',
    )
    add_bool_flag(
        parser,
        name='atlas-tumor-overlap-features',
        dest='use_atlas_tumor_overlap_features',
        default=True,
        help_enabled='Enable compact tumor-localized atlas overlap features.',
        help_disabled='Disable compact tumor-localized atlas overlap features.',
    )
    add_bool_flag(
        parser,
        name='atlas-burden-features',
        dest='use_atlas_burden_features',
        default=True,
        help_enabled='Enable compact atlas PET burden features across regions.',
        help_disabled='Disable compact atlas PET burden features across regions.',
    )
    add_bool_flag(
        parser,
        name='atlas-z-features',
        dest='use_atlas_z_features',
        default=True,
        help_enabled='Enable atlas z-score summary features from the population atlas map.',
        help_disabled='Disable atlas z-score summary features from the population atlas map.',
    )
    add_bool_flag(
        parser,
        name='pet-radiomics',
        dest='use_pet_radiomics',
        default=False,
        help_enabled='Enable PET radiomics from warped tumor VOIs.',
        help_disabled='Disable PET radiomics from warped tumor VOIs.',
    )
    add_bool_flag(
        parser,
        name='ct-radiomics',
        dest='use_ct_radiomics',
        default=False,
        help_enabled='Enable CT radiomics from warped tumor VOIs.',
        help_disabled='Disable CT radiomics from warped tumor VOIs.',
    )
    parser.add_argument('--atlas-region-min-voxels', type=int, default=1, help='Minimum atlas-region voxel count required to include a CTSeg atlas label as a biomarker feature block.')
    parser.add_argument('--z-threshold', type=float, default=2.0)
    parser.add_argument('--pet-normalization', choices=['none', 'blood_pool', 'brain', 'blood_pool_then_brain'], default='blood_pool_then_brain')
    parser.add_argument('--reference-min-voxels', type=int, default=40)
    parser.add_argument('--reference-erosion-iters', type=int, default=1)
    return parser.parse_args()


def column_ref_to_index(ref: str) -> int:
    col = ''.join(ch for ch in ref if ch.isalpha())
    out = 0
    for ch in col:
        out = out * 26 + (ord(ch.upper()) - ord('A') + 1)
    return out - 1


def load_rows_openpyxl(path: str, sheet_name: str) -> List[Dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(value).strip() if value is not None else '' for value in rows[0]]
    data_rows: List[Dict[str, object]] = []
    for values in rows[1:]:
        row = {header[index]: values[index] if index < len(values) else None for index in range(len(header))}
        data_rows.append(row)
    return data_rows


def load_rows_xlsx_xml(path: str, sheet_name: str) -> List[Dict[str, object]]:
    ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
    rel_ns = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
    with zipfile.ZipFile(path) as zf:
        shared_strings: List[str] = []
        if 'xl/sharedStrings.xml' in zf.namelist():
            shared_root = ET.fromstring(zf.read('xl/sharedStrings.xml'))
            for si in shared_root.findall(f'{ns}si'):
                txt = ''.join(t.text or '' for t in si.iter(f'{ns}t'))
                shared_strings.append(txt)

        workbook_root = ET.fromstring(zf.read('xl/workbook.xml'))
        rel_root = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
        rel_map = {rel.attrib['Id']: rel.attrib['Target'] for rel in rel_root}

        target = None
        sheets = workbook_root.find(f'{ns}sheets')
        if sheets is None:
            raise RuntimeError('Workbook is missing sheets metadata.')
        for sheet in sheets:
            if sheet.attrib.get('name') == sheet_name:
                rid = sheet.attrib[f'{rel_ns}id']
                target = rel_map[rid]
                break
        if target is None:
            raise RuntimeError(f'Sheet {sheet_name!r} not found in workbook.')

        sheet_root = ET.fromstring(zf.read('xl/' + target))
        rows_xml = sheet_root.findall(f'.//{ns}sheetData/{ns}row')
        if not rows_xml:
            return []

        header_cells = rows_xml[0].findall(f'{ns}c')
        header_map: Dict[int, str] = {}
        for cell in header_cells:
            ref = cell.attrib.get('r', '')
            idx = column_ref_to_index(ref)
            value = parse_xlsx_cell_value(cell, shared_strings, ns)
            header_map[idx] = str(value).strip() if value is not None else ''

        max_idx = max(header_map) if header_map else -1
        headers = [header_map.get(i, '') for i in range(max_idx + 1)]
        data_rows: List[Dict[str, object]] = []
        for row_xml in rows_xml[1:]:
            values: List[Optional[object]] = [None] * len(headers)
            for cell in row_xml.findall(f'{ns}c'):
                ref = cell.attrib.get('r', '')
                idx = column_ref_to_index(ref)
                if idx >= len(values):
                    values.extend([None] * (idx + 1 - len(values)))
                    headers.extend([''] * (idx + 1 - len(headers)))
                values[idx] = parse_xlsx_cell_value(cell, shared_strings, ns)
            row = {headers[i]: values[i] for i in range(len(headers)) if headers[i]}
            data_rows.append(row)
    return data_rows


def parse_xlsx_cell_value(cell: ET.Element, shared_strings: Sequence[str], ns: str):
    t = cell.attrib.get('t')
    if t == 'inlineStr':
        is_elem = cell.find(f'{ns}is')
        if is_elem is None:
            return ''
        return ''.join(txt.text or '' for txt in is_elem.iter(f'{ns}t'))
    v = cell.find(f'{ns}v')
    if v is None:
        return None
    text = v.text
    if text is None:
        return None
    if t == 's':
        return shared_strings[int(text)]
    if t == 'b':
        return int(text)
    try:
        num = float(text)
        return int(num) if num.is_integer() else num
    except Exception:
        return text


def load_clinical_rows(path: str, sheet_name: str) -> List[Dict[str, object]]:
    try:
        return load_rows_openpyxl(path, sheet_name)
    except Exception:
        return load_rows_xlsx_xml(path, sheet_name)


def to_float(value: object) -> float:
    if value is None or value == '':
        return math.nan
    try:
        return float(value)
    except Exception:
        return math.nan


def to_int(value: object) -> Optional[int]:
    if value is None or value == '':
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def clean_text(value: object) -> str:
    return '' if value is None else str(value).strip()


def filter_rows(rows: Sequence[Dict[str, object]], task: str) -> List[Dict[str, object]]:
    if task == 'all':
        return list(rows)
    key = f'Task {task}'
    return [row for row in rows if to_int(row.get(key)) == 1]


def stage_to_ord(stage: object, prefix: str) -> float:
    txt = clean_text(stage).upper()
    if not txt:
        return math.nan
    if txt.startswith(prefix.upper()):
        tail = txt[len(prefix):]
        digits = ''.join(ch for ch in tail if ch.isdigit())
        if digits:
            return float(int(digits))
        if tail.startswith('X'):
            return math.nan
    return math.nan


def map_binary(value: object) -> float:
    code = to_int(value)
    if code in (0, 1):
        return float(code)
    return math.nan


def build_clinical_matrix(row: Dict[str, object]) -> np.ndarray:
    return np.array([
        to_float(row.get('Age')),
        map_binary(row.get('Gender')),
        map_binary(row.get('Tobacco Consumption')),
        map_binary(row.get('Alcohol Consumption')),
        to_float(row.get('Performance Status')),
        map_binary(row.get('Treatment')),
        stage_to_ord(row.get('T-stage'), 'T'),
        stage_to_ord(row.get('N-stage'), 'N'),
        stage_to_ord(row.get('M-stage'), 'M'),
        map_binary(row.get('HPV Status')),
        to_float(row.get('CenterID')),
    ], dtype=float)


def norm_ct(img: np.ndarray) -> np.ndarray:
    x = img.copy()
    x[x < -300] = -300
    x[x > 300] = 300
    return (x - x.min()) / (x.max() - x.min() + 1e-8)


def create_model(device: torch.device, input_shape=INPUT_SHAPE) -> SITReg:
    feature_extractor = EncoderFeatureExtractor(
        n_input_channels=1,
        activation_factory=ReLUFactory(),
        n_features_per_resolution=[12, 16, 32, 64, 128, 128],
        n_convolutions_per_resolution=[2, 2, 2, 2, 2, 2],
        input_shape=input_shape,
        normalizer_factory=GroupNormalizerFactory(2),
    ).to(device)
    solver_forward = AndersonSolver(
        MaxElementWiseAbsStopCriterion(min_iterations=2, max_iterations=50, threshold=1e-2),
        AndersonSolverArguments(memory_length=4),
    )
    solver_backward = AndersonSolver(
        RelativeL2ErrorStopCriterion(min_iterations=2, max_iterations=50, threshold=1e-2),
        AndersonSolverArguments(memory_length=4),
    )
    return SITReg(
        feature_extractor=feature_extractor,
        n_transformation_convolutions_per_resolution=[2, 2, 2, 2, 2, 2],
        n_transformation_features_per_resolution=[12, 64, 128, 256, 256, 256],
        max_control_point_multiplier=0.99,
        affine_transformation_type=None,
        input_voxel_size=(1.0, 1.0, 1.0),
        input_shape=input_shape,
        transformation_downsampling_factor=(1.0, 1.0, 1.0),
        forward_fixed_point_solver=solver_forward,
        backward_fixed_point_solver=solver_backward,
        activation_factory=ReLUFactory(),
        normalizer_factory=GroupNormalizerFactory(4),
    ).to(device)


def make_affine_from_pixdim(pixdim):
    affine = np.eye(4, dtype=np.float32)
    affine[0, 0] = pixdim[0]
    affine[1, 1] = pixdim[1]
    affine[2, 2] = pixdim[2]
    return affine


def gaussian_kernel3d(sigma_vox: float, size_vox: Tuple[int, int, int]) -> torch.Tensor:
    ax = [torch.arange(-(s // 2), s // 2 + 1, dtype=torch.float32) for s in size_vox]
    zz, yy, xx = torch.meshgrid(ax[0], ax[1], ax[2], indexing='ij')
    g = torch.exp(-(xx ** 2 + yy ** 2 + zz ** 2) / (2 * sigma_vox ** 2))
    return g / g.sum()


def smooth_inside_mask(vol: torch.Tensor, m_float: torch.Tensor, sigma: float = 1.0) -> torch.Tensor:
    rad = max(2, int(math.ceil(3.0 * float(sigma))))
    ksz = 2 * rad + 1
    k = gaussian_kernel3d(float(sigma), (ksz, ksz, ksz)).to(vol.device)[None, None, ...]
    finite = torch.isfinite(vol)
    w = (finite & (m_float > 0)).float()
    v = torch.where(finite, vol, torch.zeros_like(vol))
    num = F.conv3d((v * w)[None, None, ...], k, padding='same')
    den = F.conv3d(w[None, None, ...], k, padding='same').clamp_min(1e-6)
    return (num / den)[0, 0]


def robust_region_reference(
    pet_np: np.ndarray,
    mask_np: np.ndarray,
    min_voxels: int = 40,
    erosion_iters: int = 1,
) -> Tuple[float, int]:
    mask = np.asarray(mask_np, dtype=bool) & np.isfinite(pet_np)
    if not mask.any():
        return np.nan, 0
    if erosion_iters and erosion_iters > 0:
        eroded = ndimage.binary_erosion(mask, structure=np.ones((3, 3, 3), dtype=np.uint8), iterations=int(erosion_iters))
        if eroded.any():
            mask = eroded
    vals = pet_np[mask]
    vals = vals[np.isfinite(vals)]
    if vals.size < int(min_voxels):
        return np.nan, int(vals.size)
    lo, hi = np.percentile(vals, [10.0, 90.0])
    trimmed = vals[(vals >= lo) & (vals <= hi)]
    use_vals = trimmed if trimmed.size >= int(min_voxels // 2) else vals
    return float(np.median(use_vals)), int(vals.size)


def compute_pet_reference(
    pet_np: np.ndarray,
    ctseg_np: np.ndarray,
    mode: str,
    min_voxels: int,
    erosion_iters: int,
) -> Tuple[float, str, int]:
    ctseg_int = np.rint(ctseg_np).astype(np.int32)
    candidates: List[Tuple[str, Tuple[int, ...]]] = []
    if mode == 'blood_pool':
        candidates = [('blood_pool', BLOOD_POOL_LABELS)]
    elif mode == 'brain':
        candidates = [('brain', BRAIN_LABELS)]
    elif mode == 'blood_pool_then_brain':
        candidates = [('blood_pool', BLOOD_POOL_LABELS), ('brain', BRAIN_LABELS)]
    else:
        return np.nan, 'none', 0

    for name, labels in candidates:
        ref_mask = np.isin(ctseg_int, labels)
        ref_value, n_vox = robust_region_reference(pet_np, ref_mask, min_voxels=min_voxels, erosion_iters=erosion_iters)
        if np.isfinite(ref_value) and ref_value > 0:
            return ref_value, name, n_vox
    return np.nan, 'none', 0


def save_vector_as_vol(vec: np.ndarray, name: str, affine: np.ndarray, shape: Tuple[int, int, int], mv_bool: np.ndarray):
    out = np.zeros(int(np.prod(shape)), dtype=np.float32)
    out[mv_bool.reshape(-1)] = vec.astype(np.float32)
    nib.save(nib.Nifti1Image(out.reshape(shape), affine), name)


def save_volume(vol: np.ndarray, name: str, affine: np.ndarray):
    nib.save(nib.Nifti1Image(vol.astype(np.float32), affine), name)


def _cox_sort_by_time(time, event, X=None):
    idx = np.argsort(time)
    time_s = time[idx]
    event_s = event[idx]
    if X is None:
        return time_s, event_s, None, idx
    return time_s, event_s, X[idx, ...], idx


def _cox_partial_grad_hess(beta, X, time, event, l2=0.0):
    if X.ndim == 1:
        X = X[:, None]
    n, p = X.shape
    xb = X @ beta
    r = np.exp(np.clip(xb, -50.0, 50.0))
    order = np.argsort(time)
    time = time[order]
    event = event[order]
    X = X[order, :]
    r = r[order]
    cr = np.cumsum(r[::-1])[::-1]
    cX = np.cumsum((r[:, None] * X)[::-1, :], axis=0)[::-1, :]
    g = np.zeros(p)
    fisher = np.zeros((p, p))
    i = 0
    while i < n:
        t = time[i]
        j = i
        while j < n and time[j] == t:
            j += 1
        d = int(event[i:j].sum())
        if d > 0:
            s0 = cr[i]
            s1 = cX[i, :]
            mu = s1 / max(s0, 1e-12)
            xe_sum = X[i:j, :][event[i:j] == 1].sum(axis=0)
            g += xe_sum - d * mu
            X_risk = X[i:, :]
            r_risk = r[i:]
            s2 = (X_risk.T * r_risk) @ X_risk / max(s0, 1e-12)
            fisher += d * (s2 - np.outer(mu, mu))
        i = j
    if l2 > 0:
        g -= l2 * beta
        fisher += l2 * np.eye(p)
    return g, fisher


def cox_fit(X, time, event, l2=1e-4, max_iter=60, tol=1e-6):
    X = np.asarray(X, dtype=float)
    time = np.asarray(time, dtype=float)
    event = np.asarray(event, dtype=int)
    if X.ndim == 1:
        X = X[:, None]
    mask = np.isfinite(time) & np.isfinite(event) & np.all(np.isfinite(X), axis=1)
    X, time, event = X[mask], time[mask], event[mask]
    if X.shape[0] < 10 or event.sum() < 5:
        p = X.shape[1]
        return np.zeros(p), np.full(p, np.inf), np.zeros(p), np.ones(p)
    time_s, event_s, X_s, _ = _cox_sort_by_time(time, event, X)
    p = X_s.shape[1]
    beta = np.zeros(p)
    l2_used = float(l2)
    for _ in range(max_iter):
        g, fisher = _cox_partial_grad_hess(beta, X_s, time_s, event_s, l2=l2_used)
        try:
            cond = np.linalg.cond(fisher)
        except np.linalg.LinAlgError:
            cond = np.inf
        if not np.isfinite(cond) or cond > 1e8:
            fisher = fisher + max(1e-6, l2_used) * np.eye(p)
            l2_used = min(max(l2_used * 10.0, 1e-6), 1.0)
        try:
            step = np.linalg.solve(fisher, g)
        except np.linalg.LinAlgError:
            step = np.linalg.pinv(fisher) @ g
        beta_new = beta + step
        beta = beta_new
        if float(np.max(np.abs(step))) < tol:
            break
    _, fisher_final = _cox_partial_grad_hess(beta, X_s, time_s, event_s, l2=l2_used)
    try:
        var = np.linalg.inv(fisher_final)
    except np.linalg.LinAlgError:
        var = np.linalg.pinv(fisher_final)
    diagv = np.diag(var)
    diagv = np.where(diagv > 1e-12, diagv, 1e-12)
    se = np.sqrt(diagv)
    z = beta / se
    try:
        from scipy.special import ndtr
        pvals = 2.0 * (1.0 - ndtr(np.abs(z)))
    except Exception:
        from math import erf, sqrt
        ndtr_vec = np.vectorize(lambda x: 0.5 * (1.0 + erf(x / sqrt(2.0))))
        pvals = 2.0 * (1.0 - ndtr_vec(np.abs(z)))
    return beta, se, z, pvals


def benjamini_hochberg(pvals, q=0.05):
    p = np.asarray(pvals, dtype=float)
    valid = np.isfinite(p)
    m = int(valid.sum())
    sig = np.zeros_like(p, dtype=bool)
    qvals = np.full_like(p, np.nan, dtype=float)
    if m == 0:
        return sig, qvals
    p_valid = p[valid]
    order = np.argsort(p_valid)
    ranked = p_valid[order]
    thresh = q * (np.arange(1, m + 1) / m)
    passed = ranked <= thresh
    sig_valid = np.zeros_like(p_valid, dtype=bool)
    if passed.any():
        cutoff = ranked[int(np.max(np.where(passed)[0]))]
        sig_valid = p_valid <= cutoff
    q_tmp = ranked * m / np.arange(1, m + 1)
    q_tmp = np.minimum.accumulate(q_tmp[::-1])[::-1]
    qvals_valid = np.empty_like(p_valid)
    qvals_valid[order] = np.clip(q_tmp, 0, 1)
    qvals[valid] = qvals_valid
    sig[valid] = sig_valid
    return sig, qvals


def concordance_index(time, event, risk):
    n = len(time)
    num = 0.0
    den = 0.0
    for i in range(n):
        for j in range(n):
            if time[i] < time[j] and event[i] == 1:
                den += 1.0
                if risk[i] > risk[j]:
                    num += 1.0
                elif risk[i] == risk[j]:
                    num += 0.5
    return num / den if den > 0 else np.nan


def zscore_matrix(A: np.ndarray):
    A = np.asarray(A, dtype=float)
    finite_mask = np.isfinite(A)
    counts = finite_mask.sum(axis=0)
    safe = np.where(finite_mask, A, 0.0)
    denom = np.maximum(counts, 1)
    means = safe.sum(axis=0) / denom
    centered = np.where(finite_mask, A - means, 0.0)
    stds = np.sqrt((centered ** 2).sum(axis=0) / denom)
    means = np.where(counts > 0, means, 0.0)
    stds = np.where(np.isfinite(stds) & (stds >= 1e-9), stds, 1.0)
    filled = np.where(finite_mask, A, means)
    Z = (filled - means) / stds
    if Z.ndim == 2:
        Z[:, counts == 0] = 0.0
    return Z, means, stds


def plot_km_with_ticks_two_groups(
    time_days: np.ndarray,
    event: np.ndarray,
    grp_high: np.ndarray,
    out_path: str,
    title: str,
    censor_merge_window_days: float = 45.0,
):
    def _merge_censor_ticks(times, levels, merge_window_days):
        times = np.asarray(times, dtype=float)
        levels = np.asarray(levels, dtype=float)
        if times.size == 0 or merge_window_days is None or merge_window_days <= 0:
            return times, levels
        order = np.argsort(times)
        times = times[order]
        levels = levels[order]
        merged_times = []
        merged_levels = []
        cluster_times = [float(times[0])]
        cluster_levels = [float(levels[0])]
        for idx in range(1, times.size):
            if float(times[idx]) - cluster_times[-1] <= float(merge_window_days):
                cluster_times.append(float(times[idx]))
                cluster_levels.append(float(levels[idx]))
            else:
                merged_times.append(float(np.mean(cluster_times)))
                merged_levels.append(float(np.mean(cluster_levels)))
                cluster_times = [float(times[idx])]
                cluster_levels = [float(levels[idx])]
        merged_times.append(float(np.mean(cluster_times)))
        merged_levels.append(float(np.mean(cluster_levels)))
        return np.asarray(merged_times, dtype=float), np.asarray(merged_levels, dtype=float)

    def _km_curve(t, e):
        order = np.argsort(t)
        t = np.asarray(t, dtype=float)[order]
        e = np.asarray(e, dtype=int)[order]
        uniq_times = np.unique(t)
        n_at_risk = float(len(t))
        surv = 1.0
        xs = [0.0]
        ys = [1.0]
        censor_times = []
        censor_levels = []
        ci_x = [0.0]
        ci_lo = [1.0]
        ci_hi = [1.0]
        var_acc = 0.0
        for ut in uniq_times:
            mask_ut = (t == ut)
            d = float((e[mask_ut] == 1).sum())
            c = float((e[mask_ut] == 0).sum())
            if n_at_risk > 0 and d > 0:
                surv *= (1.0 - d / n_at_risk)
                xs.extend([ut, ut])
                ys.extend([ys[-1], surv])
                var_acc += d / (n_at_risk * max(n_at_risk - d, 1.0))
                se = (surv ** 2) * var_acc
                delta = 1.96 * np.sqrt(max(se, 0.0))
                lo = max(0.0, surv - delta)
                hi = min(1.0, surv + delta)
                ci_x.extend([ut, ut])
                ci_lo.extend([ci_lo[-1], lo])
                ci_hi.extend([ci_hi[-1], hi])
            if c > 0:
                censor_times.extend([ut] * int(c))
                censor_levels.extend([surv] * int(c))
            n_at_risk -= (d + c)
        censor_times, censor_levels = _merge_censor_ticks(censor_times, censor_levels, censor_merge_window_days)
        return np.asarray(xs), np.asarray(ys), np.asarray(censor_times), np.asarray(censor_levels), np.asarray(ci_x), np.asarray(ci_lo), np.asarray(ci_hi)

    mask_fin = np.isfinite(time_days) & np.isfinite(event)
    t = np.asarray(time_days, dtype=float)[mask_fin]
    e = np.asarray(event, dtype=int)[mask_fin]
    grp_high = np.asarray(grp_high, dtype=bool)[mask_fin]
    if grp_high.sum() == 0 or (~grp_high).sum() == 0:
        return
    xh, yh, cth, cyh, cixh, ciloh, cihih = _km_curve(t[grp_high], e[grp_high])
    xl, yl, ctl, cyl, cixl, cilol, cihil = _km_curve(t[~grp_high], e[~grp_high])
    plt.figure(figsize=(6.5, 4.5), dpi=140)
    plt.step(xh, yh, where='post', color='tab:red', label='High group')
    plt.step(xl, yl, where='post', color='tab:blue', label='Low group')
    if cixh.size:
        plt.fill_between(cixh, ciloh, cihih, color='tab:red', alpha=0.15, step='post')
    if cixl.size:
        plt.fill_between(cixl, cilol, cihil, color='tab:blue', alpha=0.15, step='post')
    if cth.size:
        plt.scatter(cth, cyh, marker='|', color='tab:red', s=80, linewidths=1.5)
    if ctl.size:
        plt.scatter(ctl, cyl, marker='|', color='tab:blue', s=80, linewidths=1.5)
    plt.ylim(0.0, 1.05)
    plt.xlabel('Time (days)')
    plt.ylabel('Relapse-free survival probability')
    plt.title(title)
    plt.grid(alpha=0.25)
    plt.legend(loc='best')
    try:
        plt.savefig(out_path, bbox_inches='tight')
    finally:
        plt.close()


class Hecktor25Dataset(Dataset):
    def __init__(self, data_dir: str, case_ids: Sequence[str]):
        self.data_dir = data_dir
        self.case_ids = list(case_ids)

    def __len__(self) -> int:
        return len(self.case_ids)

    def __getitem__(self, index: int):
        case_id = self.case_ids[index]
        ct_nib = nib.load(os.path.join(self.data_dir, f'{case_id}_CT.nii.gz'))
        pet_nib = nib.load(os.path.join(self.data_dir, f'{case_id}_PET.nii.gz'))
        seg_path = os.path.join(self.data_dir, f'{case_id}_SEG.nii.gz')
        ctseg_path = os.path.join(self.data_dir, f'{case_id}_CTSeg.nii.gz')
        seg_nib = nib.load(seg_path) if os.path.exists(seg_path) else None
        ctseg_nib = nib.load(ctseg_path) if os.path.exists(ctseg_path) else None
        ct_raw = ct_nib.get_fdata().astype(np.float32)
        ct_norm = norm_ct(ct_raw).astype(np.float32)
        pet = pet_nib.get_fdata().astype(np.float32)
        seg = seg_nib.get_fdata().astype(np.float32) if seg_nib is not None else np.zeros_like(ct_raw, dtype=np.float32)
        ctseg = ctseg_nib.get_fdata().astype(np.float32) if ctseg_nib is not None else np.zeros_like(ct_raw, dtype=np.float32)
        return {
            'case_id': case_id,
            'ct_norm': torch.from_numpy(np.ascontiguousarray(ct_norm[None, ...])),
            'ct_raw': torch.from_numpy(np.ascontiguousarray(ct_raw[None, ...])),
            'pet': torch.from_numpy(np.ascontiguousarray(pet[None, ...])),
            'seg': torch.from_numpy(np.ascontiguousarray(seg[None, ...])),
            'ctseg': torch.from_numpy(np.ascontiguousarray(ctseg[None, ...])),
        }


def sanitize_tumor_mask(seg: np.ndarray, lesion_min_voxels: int) -> Tuple[np.ndarray, int]:
    mask = np.asarray(seg > 0.5, dtype=bool)
    if not mask.any():
        return mask, 0
    structure = np.ones((3, 3, 3), dtype=np.uint8)
    labeled, num = ndimage.label(mask.astype(np.uint8), structure=structure)
    if num <= 0:
        return np.zeros_like(mask, dtype=bool), 0
    sizes = ndimage.sum(np.ones_like(mask, dtype=np.float32), labels=labeled, index=np.arange(1, num + 1))
    keep_labels = np.nonzero(np.asarray(sizes) >= float(max(1, lesion_min_voxels)))[0] + 1
    if keep_labels.size == 0:
        return np.zeros_like(mask, dtype=bool), 0
    keep_mask = np.isin(labeled, keep_labels)
    return keep_mask.astype(bool), int(keep_labels.size)


def extract_tumor_features(pet: np.ndarray, ct: np.ndarray, seg: np.ndarray, affine: np.ndarray, lesion_min_voxels: int) -> Tuple[np.ndarray, List[str], np.ndarray]:
    names = [
        'tumor_volume_mm3', 'tumor_lesion_count', 'tumor_voxel_count', 'tumor_suv_mean', 'tumor_suv_std',
        'tumor_suv_max', 'tumor_suv_p90', 'tumor_suv_p95', 'tumor_tlg_like', 'tumor_ct_mean',
        'tumor_ct_min', 'tumor_ct_p10', 'tumor_centroid_x', 'tumor_centroid_y', 'tumor_centroid_z',
        'tumor_extent_x_mm', 'tumor_extent_y_mm', 'tumor_extent_z_mm'
    ]
    mask, lesion_count = sanitize_tumor_mask(seg, lesion_min_voxels)
    if not mask.any():
        feats = np.array([0.0, 0.0, 0.0] + [np.nan] * (len(names) - 3), dtype=float)
        return feats, names, mask
    try:
        voxel_vol = float(abs(np.linalg.det(affine[:3, :3])))
    except Exception:
        voxel_vol = float(np.prod(VOXEL_SPACING_MM))
    pet_vals = pet[mask]
    ct_vals = ct[mask]
    coords = np.argwhere(mask)
    centroid = coords.mean(axis=0).astype(float)
    extents = (coords.max(axis=0) - coords.min(axis=0) + 1).astype(float)
    spacing = np.sqrt((np.asarray(affine[:3, :3]) ** 2).sum(axis=0)) if affine is not None else np.asarray(VOXEL_SPACING_MM, dtype=float)
    feats = np.array([
        float(mask.sum()) * voxel_vol,
        float(lesion_count),
        float(mask.sum()),
        float(np.nanmean(pet_vals)),
        float(np.nanstd(pet_vals)),
        float(np.nanmax(pet_vals)),
        float(np.nanpercentile(pet_vals, 90.0)),
        float(np.nanpercentile(pet_vals, 95.0)),
        float(np.nansum(pet_vals) * voxel_vol),
        float(np.nanmean(ct_vals)),
        float(np.nanmin(ct_vals)),
        float(np.nanpercentile(ct_vals, 10.0)),
        float(centroid[0]),
        float(centroid[1]),
        float(centroid[2]),
        float(extents[0] * spacing[0]),
        float(extents[1] * spacing[1]),
        float(extents[2] * spacing[2]),
    ], dtype=float)
    return feats, names, mask


def affine_spacing_xyz(affine: np.ndarray) -> Tuple[float, float, float]:
    if affine is None:
        return tuple(float(v) for v in VOXEL_SPACING_MM)
    spacing = np.sqrt((np.asarray(affine[:3, :3], dtype=float) ** 2).sum(axis=0))
    return tuple(float(v) for v in spacing)


def array_xyz_to_sitk_image(arr_xyz: np.ndarray, affine: np.ndarray) -> sitk.Image:
    image = sitk.GetImageFromArray(np.transpose(arr_xyz.astype(np.float32), (2, 1, 0)))
    image.SetSpacing(affine_spacing_xyz(affine))
    return image


def mask_xyz_to_sitk(mask_xyz: np.ndarray, reference_image: sitk.Image) -> sitk.Image:
    mask_image = sitk.GetImageFromArray(np.transpose(mask_xyz.astype(np.uint8), (2, 1, 0)))
    mask_image.CopyInformation(reference_image)
    return mask_image


def build_radiomics_extractor(bin_width: float, feature_classes: Sequence[str], force_2d: bool):
    if featureextractor is None:
        raise RuntimeError('PyRadiomics is not available. Install the radiomics package or disable radiomics flags.')
    settings = {
        'binWidth': float(bin_width),
        'normalize': False,
        'removeOutliers': None,
        'force2D': bool(force_2d),
        'label': 1,
        'additionalInfo': False,
    }
    extractor = featureextractor.RadiomicsFeatureExtractor(**settings)
    extractor.disableAllImageTypes()
    for image_type in RADIOMICS_IMAGE_TYPES:
        extractor.enableImageTypeByName(image_type)
    extractor.disableAllFeatures()
    for feature_class in feature_classes:
        extractor.enableFeatureClassByName(feature_class)
    return extractor


def extract_radiomics_features_xyz(
    image_xyz: np.ndarray,
    mask_xyz: np.ndarray,
    affine: np.ndarray,
    extractor,
    prefix: str,
) -> Tuple[np.ndarray, List[str]]:
    image = array_xyz_to_sitk_image(image_xyz, affine)
    mask_image = mask_xyz_to_sitk(mask_xyz, image)
    raw_features = extractor.execute(image, mask_image)
    names: List[str] = []
    values: List[float] = []
    for key in sorted(raw_features):
        if key.startswith('diagnostics_'):
            continue
        value = raw_features[key]
        try:
            value_float = float(value)
        except Exception:
            continue
        names.append(f'{prefix}_{key}')
        values.append(value_float)
    return np.asarray(values, dtype=float), names


def compute_univariate_feature_stats(
    X: np.ndarray,
    time_np: np.ndarray,
    event_np: np.ndarray,
    min_coverage: int,
    min_events: int,
) -> Dict[str, np.ndarray]:
    feat_beta = []
    feat_z = []
    feat_p = []
    feat_cov = []
    feat_evt = []
    feat_std = []
    for j in range(X.shape[1]):
        col = X[:, j]
        mfin = np.isfinite(col)
        n_cov = int(mfin.sum())
        e_cov = int(event_np[mfin].sum())
        std_val = float(np.nanstd(col)) if n_cov > 0 else math.nan
        feat_cov.append(n_cov)
        feat_evt.append(e_cov)
        feat_std.append(std_val)
        if n_cov < min_coverage or e_cov < min_events or (not np.isfinite(std_val)) or std_val < 1e-9:
            feat_beta.append(0.0)
            feat_z.append(0.0)
            feat_p.append(1.0)
            continue
        mean_val = float(np.nanmean(col))
        xz = (np.where(np.isfinite(col), col, mean_val) - mean_val) / (std_val + 1e-9)
        xz = np.clip(xz, -8.0, 8.0)
        beta, _, z_val, p_val = cox_fit(xz, time_np, event_np, l2=1e-4, max_iter=60)
        feat_beta.append(float(np.atleast_1d(beta)[0]))
        feat_z.append(float(np.atleast_1d(z_val)[0]))
        feat_p.append(float(np.atleast_1d(p_val)[0]))
    feat_beta = np.asarray(feat_beta, dtype=float)
    feat_z = np.asarray(feat_z, dtype=float)
    feat_p = np.asarray(feat_p, dtype=float)
    feat_cov = np.asarray(feat_cov, dtype=int)
    feat_evt = np.asarray(feat_evt, dtype=int)
    feat_std = np.asarray(feat_std, dtype=float)
    sig05, q05 = benjamini_hochberg(feat_p, q=0.05)
    sig10, q10 = benjamini_hochberg(feat_p, q=0.10)
    return {
        'beta': feat_beta,
        'z': feat_z,
        'p': feat_p,
        'q05': q05,
        'q10': q10,
        'sig05': sig05,
        'sig10': sig10,
        'coverage': feat_cov,
        'events': feat_evt,
        'std': feat_std,
    }


def greedy_correlation_prune(X: np.ndarray, ranked_indices: np.ndarray, threshold: float) -> np.ndarray:
    if ranked_indices.size == 0:
        return ranked_indices
    if threshold is None or threshold >= 1.0 or X.ndim != 2 or X.shape[0] < 2 or X.shape[1] < 2:
        return ranked_indices.copy()
    Xz, _, _ = zscore_matrix(X)
    corr = np.corrcoef(Xz, rowvar=False)
    if np.ndim(corr) != 2:
        return ranked_indices.copy()
    corr = np.nan_to_num(corr, nan=0.0, posinf=0.0, neginf=0.0)
    keep: List[int] = []
    dropped = np.zeros(X.shape[1], dtype=bool)
    for idx in ranked_indices:
        idx = int(idx)
        if dropped[idx]:
            continue
        keep.append(idx)
        correlated = np.abs(corr[idx]) >= float(threshold)
        correlated[idx] = False
        dropped |= correlated
    return np.asarray(keep, dtype=int)


def select_feature_indices(
    feat_p: np.ndarray,
    q05: np.ndarray,
    q10: np.ndarray,
    X: np.ndarray,
    feature_selection: str,
    topk: int,
    correlation_threshold: float,
    max_selected_features: int,
) -> Tuple[np.ndarray, np.ndarray, str]:
    def _cap_selected(selected_indices: np.ndarray, selection_label: str) -> Tuple[np.ndarray, str]:
        if int(max_selected_features) <= 0 or selected_indices.size <= int(max_selected_features):
            return selected_indices, selection_label
        capped = selected_indices[:int(max_selected_features)]
        return capped, f'{selection_label}+cap({int(max_selected_features)})'

    finite_idx = np.where(np.isfinite(feat_p))[0]
    ranked = finite_idx[np.argsort(feat_p[finite_idx])]
    pruned_ranked = greedy_correlation_prune(X, ranked, correlation_threshold)
    if feature_selection == 'conventional':
        selected_q05 = np.asarray([idx for idx in pruned_ranked if np.isfinite(q05[idx]) and q05[idx] <= 0.05], dtype=int)
        selected_q10 = np.asarray([idx for idx in pruned_ranked if np.isfinite(q10[idx]) and q10[idx] <= 0.10], dtype=int)
        if selected_q05.size > 0:
            selected = selected_q05
            selection_used = 'conventional:fdr05'
        elif selected_q10.size > 0:
            selected = selected_q10
            selection_used = 'conventional:fdr10'
        else:
            selected = pruned_ranked[:min(int(topk), pruned_ranked.size)] if int(topk) > 0 else pruned_ranked
            selection_used = f'conventional:topk_fallback(k={int(topk)})'
    elif feature_selection == 'all':
        selected = pruned_ranked
        selection_used = 'all'
    elif feature_selection == 'topk':
        selected = pruned_ranked[:min(int(topk), pruned_ranked.size)] if int(topk) > 0 else pruned_ranked
        selection_used = f'topk(k={int(topk)})'
    elif feature_selection == 'fdr05':
        selected = np.asarray([idx for idx in pruned_ranked if np.isfinite(q05[idx]) and q05[idx] <= 0.05], dtype=int)
        selection_used = 'fdr05'
    elif feature_selection == 'fdr10':
        selected = np.asarray([idx for idx in pruned_ranked if np.isfinite(q10[idx]) and q10[idx] <= 0.10], dtype=int)
        selection_used = 'fdr10'
    else:
        raise ValueError(f'Unsupported feature selection mode: {feature_selection}')
    selected, selection_used = _cap_selected(selected, selection_used)
    return selected, pruned_ranked, selection_used


def extract_atlas_region_features(
    pet: np.ndarray,
    ct: np.ndarray,
    region_masks: Sequence[Tuple[str, np.ndarray]],
) -> Tuple[np.ndarray, List[str]]:
    feats: List[float] = []
    names: List[str] = []
    for region_name, region_mask_np in region_masks:
        region_mask = np.asarray(region_mask_np, dtype=bool)
        pet_vals = pet[region_mask]
        ct_vals = ct[region_mask]
        pet_vals = pet_vals[np.isfinite(pet_vals)]
        ct_vals = ct_vals[np.isfinite(ct_vals)]

        if pet_vals.size == 0:
            pet_mean = np.nan
            pet_p90 = np.nan
            pet_p95 = np.nan
            pet_max = np.nan
        else:
            pet_mean = float(np.nanmean(pet_vals))
            pet_p90 = float(np.nanpercentile(pet_vals, 90.0))
            pet_p95 = float(np.nanpercentile(pet_vals, 95.0))
            pet_max = float(np.nanmax(pet_vals))

        if ct_vals.size == 0:
            ct_mean = np.nan
            ct_p10 = np.nan
        else:
            ct_mean = float(np.nanmean(ct_vals))
            ct_p10 = float(np.nanpercentile(ct_vals, 10.0))

        prefix = region_name
        names.extend([
            f'{prefix}_pet_mean',
            f'{prefix}_pet_p90',
            f'{prefix}_pet_p95',
            f'{prefix}_pet_max',
            f'{prefix}_ct_mean',
            f'{prefix}_ct_p10',
        ])
        feats.extend([pet_mean, pet_p90, pet_p95, pet_max, ct_mean, ct_p10])
    return np.asarray(feats, dtype=float), names


def extract_atlas_tumor_overlap_features(
    pet: np.ndarray,
    tumor_mask: np.ndarray,
    region_masks: Sequence[Tuple[str, np.ndarray]],
) -> Tuple[np.ndarray, List[str]]:
    names = [
        'atlas_overlap_num_regions',
        'atlas_overlap_entropy',
        'atlas_overlap_top1_fraction',
        'atlas_overlap_top2_fraction',
        'atlas_overlap_top3_fraction',
        'atlas_overlap_top3_sum_fraction',
        'atlas_overlap_fraction_std',
        'atlas_overlap_regions_gt10pct',
        'atlas_overlap_regions_gt20pct',
        'atlas_overlap_region_pet_mean_max',
        'atlas_overlap_region_pet_mean_top3_mean',
        'atlas_overlap_region_pet_mean_std',
    ]
    tumor_mask = np.asarray(tumor_mask, dtype=bool)
    if not tumor_mask.any():
        return np.asarray([0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, np.nan, np.nan, np.nan], dtype=float), names

    total_tumor_vox = float(tumor_mask.sum())
    overlap_fracs: List[float] = []
    overlap_pet_means: List[float] = []
    for _, region_mask_np in region_masks:
        overlap_mask = tumor_mask & np.asarray(region_mask_np, dtype=bool)
        overlap_vox = int(overlap_mask.sum())
        if overlap_vox <= 0:
            continue
        frac = float(overlap_vox) / total_tumor_vox
        overlap_fracs.append(frac)
        pet_vals = pet[overlap_mask]
        pet_vals = pet_vals[np.isfinite(pet_vals)]
        overlap_pet_means.append(float(np.nanmean(pet_vals)) if pet_vals.size > 0 else np.nan)

    if not overlap_fracs:
        return np.asarray([0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, np.nan, np.nan, np.nan], dtype=float), names

    frac_arr = np.asarray(overlap_fracs, dtype=float)
    pet_mean_arr = np.asarray(overlap_pet_means, dtype=float)
    frac_sorted = np.sort(frac_arr)[::-1]
    top1 = float(frac_sorted[0]) if frac_sorted.size >= 1 else 0.0
    top2 = float(frac_sorted[1]) if frac_sorted.size >= 2 else 0.0
    top3 = float(frac_sorted[2]) if frac_sorted.size >= 3 else 0.0
    top3_sum = float(frac_sorted[:3].sum()) if frac_sorted.size >= 1 else 0.0
    entropy = float(-np.sum(frac_arr * np.log(np.clip(frac_arr, 1e-8, None))))
    valid_pet_mean = pet_mean_arr[np.isfinite(pet_mean_arr)]
    if valid_pet_mean.size == 0:
        pet_mean_max = np.nan
        pet_mean_top3 = np.nan
        pet_mean_std = np.nan
    else:
        pet_mean_sorted = np.sort(valid_pet_mean)[::-1]
        pet_mean_max = float(pet_mean_sorted[0])
        pet_mean_top3 = float(np.mean(pet_mean_sorted[:min(3, pet_mean_sorted.size)]))
        pet_mean_std = float(np.std(valid_pet_mean))

    feats = np.asarray([
        float(frac_arr.size),
        entropy,
        top1,
        top2,
        top3,
        top3_sum,
        float(np.std(frac_arr)),
        float((frac_arr >= 0.10).sum()),
        float((frac_arr >= 0.20).sum()),
        pet_mean_max,
        pet_mean_top3,
        pet_mean_std,
    ], dtype=float)
    return feats, names


def extract_atlas_burden_features(
    pet: np.ndarray,
    region_masks: Sequence[Tuple[str, np.ndarray]],
) -> Tuple[np.ndarray, List[str]]:
    names = [
        'atlas_burden_num_regions',
        'atlas_burden_pet_mean_top1',
        'atlas_burden_pet_mean_top3_mean',
        'atlas_burden_pet_mean_std',
        'atlas_burden_pet_mean_gt1_count',
        'atlas_burden_pet_mean_gt1p2_count',
        'atlas_burden_pet_mean_gt1p5_count',
        'atlas_burden_pet_mean_excess_gt1_sum',
        'atlas_burden_pet_mean_excess_gt1p2_sum',
        'atlas_burden_pet_max_top1',
        'atlas_burden_pet_max_top3_mean',
        'atlas_burden_pet_max_gt2_count',
        'atlas_burden_pet_max_gt3_count',
        'atlas_burden_pet_max_excess_gt2_sum',
    ]
    region_pet_mean: List[float] = []
    region_pet_max: List[float] = []
    for _, region_mask_np in region_masks:
        region_mask = np.asarray(region_mask_np, dtype=bool)
        pet_vals = pet[region_mask]
        pet_vals = pet_vals[np.isfinite(pet_vals)]
        if pet_vals.size == 0:
            continue
        region_pet_mean.append(float(np.nanmean(pet_vals)))
        region_pet_max.append(float(np.nanmax(pet_vals)))

    if not region_pet_mean:
        return np.asarray([0.0, np.nan, np.nan, np.nan, 0.0, 0.0, 0.0, 0.0, 0.0, np.nan, np.nan, 0.0, 0.0, 0.0], dtype=float), names

    mean_arr = np.asarray(region_pet_mean, dtype=float)
    max_arr = np.asarray(region_pet_max, dtype=float)
    mean_sorted = np.sort(mean_arr)[::-1]
    max_sorted = np.sort(max_arr)[::-1]
    feats = np.asarray([
        float(mean_arr.size),
        float(mean_sorted[0]),
        float(np.mean(mean_sorted[:min(3, mean_sorted.size)])),
        float(np.std(mean_arr)),
        float((mean_arr > 1.0).sum()),
        float((mean_arr > 1.2).sum()),
        float((mean_arr > 1.5).sum()),
        float(np.sum(np.clip(mean_arr - 1.0, a_min=0.0, a_max=None))),
        float(np.sum(np.clip(mean_arr - 1.2, a_min=0.0, a_max=None))),
        float(max_sorted[0]),
        float(np.mean(max_sorted[:min(3, max_sorted.size)])),
        float((max_arr > 2.0).sum()),
        float((max_arr > 3.0).sum()),
        float(np.sum(np.clip(max_arr - 2.0, a_min=0.0, a_max=None))),
    ], dtype=float)
    return feats, names


def build_atlas_region_masks(
    atlas_ctseg: np.ndarray,
    atlas_ct: np.ndarray,
    min_voxels: int,
) -> List[Tuple[str, np.ndarray]]:
    region_masks: List[Tuple[str, np.ndarray]] = []
    labeled_union = np.zeros_like(atlas_ctseg, dtype=bool)
    min_voxels = max(1, int(min_voxels))
    for label_id in np.unique(atlas_ctseg):
        label_id = int(label_id)
        if label_id <= 0:
            continue
        label_mask = np.asarray(atlas_ctseg == label_id, dtype=bool)
        if int(label_mask.sum()) < int(min_voxels):
            continue
        region_masks.append((f'atlas_lbl_{label_id:03d}', label_mask))
        labeled_union |= label_mask

    body_mask = np.asarray(atlas_ct > 0.01, dtype=bool)
    unlabeled_body_mask = body_mask & (~labeled_union)
    if int(unlabeled_body_mask.sum()) >= int(min_voxels):
        region_masks.append(('atlas_body_unlabeled', unlabeled_body_mask))
    return region_masks


def compute_atlas_z_stats(Ximg: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    mean_vec = np.nanmean(Ximg, axis=0).astype(np.float32)
    std_vec = np.nanstd(Ximg, axis=0).astype(np.float32)
    std_vec = np.where(np.isfinite(std_vec) & (std_vec >= 1e-6), std_vec, 1.0).astype(np.float32)
    return mean_vec, std_vec


def filter_feature_columns(
    feature_matrix: np.ndarray,
    feature_names: Sequence[str],
    excluded_prefixes: Sequence[str],
) -> Tuple[np.ndarray, List[str]]:
    keep_indices = [
        index for index, name in enumerate(feature_names)
        if not any(name.startswith(prefix) for prefix in excluded_prefixes)
    ]
    filtered_names = [feature_names[index] for index in keep_indices]
    if feature_matrix.ndim != 2:
        raise ValueError('feature_matrix must be a 2D array.')
    return feature_matrix[:, keep_indices], filtered_names


def write_case_feature_table(
    output_csv: str,
    patient_ids: Sequence[str],
    time_np: np.ndarray,
    event_np: np.ndarray,
    Xclin: np.ndarray,
    Ximg: np.ndarray,
    imaging_names: Sequence[str],
) -> None:
    clinical_names = [
        'Age', 'Gender', 'Tobacco Consumption', 'Alcohol Consumption', 'Performance Status',
        'Treatment', 'T-stage', 'N-stage', 'M-stage', 'HPV Status', 'CenterID'
    ]
    with open(output_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['PatientID', 'RFS', 'Relapse'] + clinical_names + list(imaging_names))
        for i, patient_id in enumerate(patient_ids):
            writer.writerow([
                patient_id,
                float(time_np[i]),
                int(event_np[i]),
                *[float(x) if np.isfinite(x) else '' for x in Xclin[i]],
                *[float(x) if np.isfinite(x) else '' for x in Ximg[i]],
            ])


def add_z_features(feature_rows: List[np.ndarray], feature_names: List[str], Ximg: np.ndarray, seg_masks: List[np.ndarray], keep_mask: np.ndarray, body_shape: Tuple[int, int, int], affine: np.ndarray, z_threshold: float):
    mean_vec, std_vec = compute_atlas_z_stats(Ximg)
    z_names = ['tumor_z_mean', 'tumor_z_max', 'tumor_z_p90', 'tumor_z_p95', f'tumor_zvol_gt_{z_threshold:g}_mm3']
    try:
        voxel_vol = float(abs(np.linalg.det(affine[:3, :3])))
    except Exception:
        voxel_vol = float(np.prod(VOXEL_SPACING_MM))
    for index in range(len(feature_rows)):
        x = Ximg[index]
        z = (x - mean_vec) / std_vec
        z3d = np.full(body_shape, np.nan, dtype=np.float32)
        z3d.reshape(-1)[keep_mask.reshape(-1)] = z.astype(np.float32)
        tumor_mask = seg_masks[index]
        if tumor_mask is None or not tumor_mask.any():
            z_feats = np.array([np.nan, np.nan, np.nan, np.nan, 0.0], dtype=float)
        else:
            z_vals = z3d[tumor_mask]
            z_vals = z_vals[np.isfinite(z_vals)]
            if z_vals.size == 0:
                z_feats = np.array([np.nan, np.nan, np.nan, np.nan, 0.0], dtype=float)
            else:
                z_feats = np.array([
                    float(np.nanmean(z_vals)),
                    float(np.nanmax(z_vals)),
                    float(np.nanpercentile(z_vals, 90.0)),
                    float(np.nanpercentile(z_vals, 95.0)),
                    float((z_vals > z_threshold).sum()) * voxel_vol,
                ], dtype=float)
        feature_rows[index] = np.concatenate([feature_rows[index], z_feats], axis=0)
    feature_names = feature_names + z_names
    return feature_rows, feature_names, mean_vec, std_vec


def add_atlas_z_summary_features(
    feature_rows: List[np.ndarray],
    feature_names: List[str],
    Ximg: np.ndarray,
    seg_masks: List[np.ndarray],
    region_masks: Sequence[Tuple[str, np.ndarray]],
    keep_mask: np.ndarray,
    body_shape: Tuple[int, int, int],
    affine: np.ndarray,
    z_threshold: float,
):
    mean_vec, std_vec = compute_atlas_z_stats(Ximg)
    try:
        voxel_vol = float(abs(np.linalg.det(affine[:3, :3])))
    except Exception:
        voxel_vol = float(np.prod(VOXEL_SPACING_MM))
    z_names = [
        'atlas_z_region_mean_top1',
        'atlas_z_region_mean_top3_mean',
        'atlas_z_region_mean_std',
        'atlas_z_region_mean_gt1_count',
        'atlas_z_region_mean_gt2_count',
        'atlas_z_region_mean_excess_gt1_sum',
        'atlas_z_region_mean_top1_minus_top3_mean',
        'atlas_z_region_max_top1',
        'atlas_z_region_max_top3_mean',
        'atlas_z_region_max_gt2_count',
        'atlas_z_region_max_gt3_count',
        'atlas_z_overlap_mean',
        'atlas_z_overlap_pos_mean',
        'atlas_z_overlap_max',
        'atlas_z_overlap_p90',
        'atlas_z_overlap_top10_mean',
        'atlas_z_overlap_frac_gt1',
        'atlas_z_overlap_frac_gt2',
        'atlas_z_overlap_excess_gt1_mm3',
        'atlas_z_overlap_excess_gt2_mm3',
        f'atlas_z_overlap_vol_gt_{z_threshold:g}_mm3',
        'atlas_z_overlap_region_mean_top1',
        'atlas_z_overlap_region_mean_top3_mean',
        'atlas_z_overlap_region_mean_std',
        'atlas_z_overlap_region_mean_weighted',
    ]
    for index in range(len(feature_rows)):
        x = Ximg[index]
        z = (x - mean_vec) / std_vec
        z3d = np.full(body_shape, np.nan, dtype=np.float32)
        z3d.reshape(-1)[keep_mask.reshape(-1)] = z.astype(np.float32)

        region_mean_vals: List[float] = []
        region_max_vals: List[float] = []
        for _, region_mask_np in region_masks:
            vals = z3d[np.asarray(region_mask_np, dtype=bool)]
            vals = vals[np.isfinite(vals)]
            if vals.size == 0:
                continue
            region_mean_vals.append(float(np.nanmean(vals)))
            region_max_vals.append(float(np.nanmax(vals)))

        if region_mean_vals:
            region_mean_arr = np.asarray(region_mean_vals, dtype=float)
            region_mean_sorted = np.sort(region_mean_arr)[::-1]
            z_region_mean_top1 = float(region_mean_sorted[0])
            z_region_mean_top3 = float(np.mean(region_mean_sorted[:min(3, region_mean_sorted.size)]))
            z_region_mean_std = float(np.std(region_mean_arr))
            z_region_mean_gt1 = float((region_mean_arr > 1.0).sum())
            z_region_mean_gt2 = float((region_mean_arr > 2.0).sum())
            z_region_mean_excess = float(np.sum(np.clip(region_mean_arr - 1.0, a_min=0.0, a_max=None)))
            z_region_mean_top1_minus_top3 = float(z_region_mean_top1 - z_region_mean_top3)
        else:
            z_region_mean_top1 = np.nan
            z_region_mean_top3 = np.nan
            z_region_mean_std = np.nan
            z_region_mean_gt1 = 0.0
            z_region_mean_gt2 = 0.0
            z_region_mean_excess = 0.0
            z_region_mean_top1_minus_top3 = np.nan

        if region_max_vals:
            region_max_arr = np.asarray(region_max_vals, dtype=float)
            region_max_sorted = np.sort(region_max_arr)[::-1]
            z_region_max_top1 = float(region_max_sorted[0])
            z_region_max_top3 = float(np.mean(region_max_sorted[:min(3, region_max_sorted.size)]))
            z_region_max_gt2 = float((region_max_arr > 2.0).sum())
            z_region_max_gt3 = float((region_max_arr > 3.0).sum())
        else:
            z_region_max_top1 = np.nan
            z_region_max_top3 = np.nan
            z_region_max_gt2 = 0.0
            z_region_max_gt3 = 0.0

        tumor_mask = np.asarray(seg_masks[index], dtype=bool) if seg_masks[index] is not None else np.zeros(body_shape, dtype=bool)
        if tumor_mask.any():
            overlap_vals = z3d[tumor_mask]
            overlap_vals = overlap_vals[np.isfinite(overlap_vals)]
            if overlap_vals.size > 0:
                z_overlap_mean = float(np.nanmean(overlap_vals))
                overlap_pos_vals = overlap_vals[overlap_vals > 0.0]
                z_overlap_pos_mean = float(np.nanmean(overlap_pos_vals)) if overlap_pos_vals.size > 0 else 0.0
                z_overlap_max = float(np.nanmax(overlap_vals))
                z_overlap_p90 = float(np.nanpercentile(overlap_vals, 90.0))
                top_count = max(1, int(math.ceil(0.10 * overlap_vals.size)))
                overlap_sorted = np.sort(overlap_vals)[::-1]
                z_overlap_top10_mean = float(np.mean(overlap_sorted[:top_count]))
                z_overlap_frac_gt1 = float(np.mean(overlap_vals > 1.0))
                z_overlap_frac_gt2 = float(np.mean(overlap_vals > 2.0))
                z_overlap_excess_gt1 = float(np.sum(np.clip(overlap_vals - 1.0, a_min=0.0, a_max=None)) * voxel_vol)
                z_overlap_excess_gt2 = float(np.sum(np.clip(overlap_vals - 2.0, a_min=0.0, a_max=None)) * voxel_vol)
                z_overlap_vol = float((overlap_vals > z_threshold).sum()) * voxel_vol
            else:
                z_overlap_mean = np.nan
                z_overlap_pos_mean = np.nan
                z_overlap_max = np.nan
                z_overlap_p90 = np.nan
                z_overlap_top10_mean = np.nan
                z_overlap_frac_gt1 = np.nan
                z_overlap_frac_gt2 = np.nan
                z_overlap_excess_gt1 = np.nan
                z_overlap_excess_gt2 = np.nan
                z_overlap_vol = 0.0

            overlap_region_means: List[float] = []
            overlap_region_weights: List[float] = []
            for _, region_mask_np in region_masks:
                overlap_mask = tumor_mask & np.asarray(region_mask_np, dtype=bool)
                vals = z3d[overlap_mask]
                vals = vals[np.isfinite(vals)]
                if vals.size == 0:
                    continue
                overlap_region_means.append(float(np.nanmean(vals)))
                overlap_region_weights.append(float(vals.size))
            if overlap_region_means:
                overlap_region_arr = np.asarray(overlap_region_means, dtype=float)
                overlap_region_sorted = np.sort(overlap_region_arr)[::-1]
                z_overlap_region_top1 = float(overlap_region_sorted[0])
                z_overlap_region_top3 = float(np.mean(overlap_region_sorted[:min(3, overlap_region_sorted.size)]))
                z_overlap_region_std = float(np.std(overlap_region_arr))
                z_overlap_region_weighted = float(np.average(overlap_region_arr, weights=np.asarray(overlap_region_weights, dtype=float)))
            else:
                z_overlap_region_top1 = np.nan
                z_overlap_region_top3 = np.nan
                z_overlap_region_std = np.nan
                z_overlap_region_weighted = np.nan
        else:
            z_overlap_mean = np.nan
            z_overlap_pos_mean = np.nan
            z_overlap_max = np.nan
            z_overlap_p90 = np.nan
            z_overlap_top10_mean = np.nan
            z_overlap_frac_gt1 = np.nan
            z_overlap_frac_gt2 = np.nan
            z_overlap_excess_gt1 = np.nan
            z_overlap_excess_gt2 = np.nan
            z_overlap_vol = 0.0
            z_overlap_region_top1 = np.nan
            z_overlap_region_top3 = np.nan
            z_overlap_region_std = np.nan
            z_overlap_region_weighted = np.nan

        z_feats = np.asarray([
            z_region_mean_top1,
            z_region_mean_top3,
            z_region_mean_std,
            z_region_mean_gt1,
            z_region_mean_gt2,
            z_region_mean_excess,
            z_region_mean_top1_minus_top3,
            z_region_max_top1,
            z_region_max_top3,
            z_region_max_gt2,
            z_region_max_gt3,
            z_overlap_mean,
            z_overlap_pos_mean,
            z_overlap_max,
            z_overlap_p90,
            z_overlap_top10_mean,
            z_overlap_frac_gt1,
            z_overlap_frac_gt2,
            z_overlap_excess_gt1,
            z_overlap_excess_gt2,
            z_overlap_vol,
            z_overlap_region_top1,
            z_overlap_region_top3,
            z_overlap_region_std,
            z_overlap_region_weighted,
        ], dtype=float)
        feature_rows[index] = np.concatenate([feature_rows[index], z_feats], axis=0)

    feature_names = feature_names + z_names
    return feature_rows, feature_names, mean_vec, std_vec


def collect_case_ids(data_dir: str, require_seg: bool = True) -> List[str]:
    out = []
    for ct_path in sorted(glob.glob(os.path.join(data_dir, '*_CT.nii.gz'))):
        case_id = os.path.basename(ct_path)[:-len('_CT.nii.gz')]
        pet_path = os.path.join(data_dir, f'{case_id}_PET.nii.gz')
        seg_path = os.path.join(data_dir, f'{case_id}_SEG.nii.gz')
        if os.path.exists(pet_path) and ((not require_seg) or os.path.exists(seg_path)):
            out.append(case_id)
    return out


def latest_file_in_dir(folder: str, suffix: str) -> str:
    files = sorted(glob.glob(os.path.join(folder, f'*{suffix}')))
    if not files:
        raise FileNotFoundError(f'No files with suffix {suffix!r} found in {folder}')
    return files[-1]


def run_voxelwise_cox(Ximg: np.ndarray, time_np: np.ndarray, event_np: np.ndarray, keep_mask: np.ndarray, affine: np.ndarray,
                     shape: Tuple[int, int, int], output_dir: str, voxel_chunk: int, min_coverage: int, min_events: int):
    V = Ximg.shape[1]
    betas = np.full(V, np.nan, dtype=float)
    ses = np.full(V, np.nan, dtype=float)
    zs = np.full(V, np.nan, dtype=float)
    ps = np.full(V, np.nan, dtype=float)
    cover = np.zeros(V, dtype=np.int32)
    stdev = np.zeros(V, dtype=float)
    events_cover = np.zeros(V, dtype=np.int32)
    for start in range(0, V, voxel_chunk):
        end = min(start + voxel_chunk, V)
        Xi = Ximg[:, start:end]
        for j in range(Xi.shape[1]):
            xcol = Xi[:, j]
            mask_fin = np.isfinite(xcol)
            n_cov = int(mask_fin.sum())
            e_cov = int(event_np[mask_fin].sum())
            s = np.nanstd(xcol)
            cover[start + j] = n_cov
            events_cover[start + j] = e_cov
            stdev[start + j] = s if np.isfinite(s) else np.nan
            if n_cov < min_coverage or e_cov < min_events or (not np.isfinite(s)) or s < 1e-6:
                continue
            m = np.nanmean(xcol)
            xz = (np.where(np.isfinite(xcol), xcol, m) - m) / (s + 1e-9)
            xz = np.clip(xz, -8.0, 8.0)
            b, se, z, p = cox_fit(xz, time_np, event_np, l2=1e-4, max_iter=40)
            betas[start + j] = float(np.atleast_1d(b)[0])
            ses[start + j] = float(np.atleast_1d(se)[0])
            zs[start + j] = float(np.atleast_1d(z)[0])
            ps[start + j] = float(np.atleast_1d(p)[0])
        print(f'Voxelwise Cox progress: {end}/{V}')
    sig05, q05 = benjamini_hochberg(ps, q=0.05)
    sig10, q10 = benjamini_hochberg(ps, q=0.10)
    hr = np.exp(np.clip(betas, -50.0, 50.0))
    beta_map = betas.copy()
    save_vector_as_vol(hr, os.path.join(output_dir, 'hr_map.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(beta_map.astype(np.float32), os.path.join(output_dir, 'beta_map.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(zs.astype(np.float32), os.path.join(output_dir, 'z_map.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(np.nan_to_num(ps, nan=1.0).astype(np.float32), os.path.join(output_dir, 'p_map.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(np.nan_to_num(q05, nan=1.0).astype(np.float32), os.path.join(output_dir, 'q_map_q05.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(np.nan_to_num(q10, nan=1.0).astype(np.float32), os.path.join(output_dir, 'q_map_q10.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(sig05.astype(np.float32), os.path.join(output_dir, 'sig_mask_fdr_q05.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(sig10.astype(np.float32), os.path.join(output_dir, 'sig_mask_fdr_q10.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(cover.astype(np.float32), os.path.join(output_dir, 'coverage_map.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(np.nan_to_num(stdev, nan=0.0).astype(np.float32), os.path.join(output_dir, 'std_map.nii.gz'), affine, shape, keep_mask)
    save_vector_as_vol(events_cover.astype(np.float32), os.path.join(output_dir, 'events_map.nii.gz'), affine, shape, keep_mask)
    valid_fit = np.isfinite(betas) & np.isfinite(zs) & np.isfinite(ps)
    return {
        'sig05': sig05,
        'sig10': sig10,
        'q05': q05,
        'q10': q10,
        'cover': cover,
        'events_cover': events_cover,
        'valid_fit': valid_fit,
    }


def main():
    args = parse_args()
    if (args.use_pet_radiomics or args.use_ct_radiomics) and featureextractor is None:
        raise RuntimeError('Radiomics requested, but PyRadiomics is not available in this environment.')
    os.makedirs(args.output_dir, exist_ok=True)
    device = torch.device(args.device if ('cuda' in args.device and torch.cuda.is_available()) else 'cpu')
    print(f'Using device: {device}')

    atlas_ct_path = latest_file_in_dir(args.atlas_ct_dir, '.nii.gz')
    atlas_pet_path = latest_file_in_dir(args.atlas_pet_dir, '.nii.gz')
    checkpoint_path = latest_file_in_dir(args.checkpoint_dir, '.pth.tar')
    print(f'Atlas CT: {atlas_ct_path}')
    print(f'Atlas PET: {atlas_pet_path}')
    print(f'Checkpoint: {checkpoint_path}')

    atlas_ct_nib = nib.load(atlas_ct_path)
    atlas_pet_nib = nib.load(atlas_pet_path)
    atlas_ct = atlas_ct_nib.get_fdata().astype(np.float32)
    atlas_pet = atlas_pet_nib.get_fdata().astype(np.float32)
    atlas_affine = atlas_ct_nib.affine.astype(np.float32)
    H, W, D = atlas_ct.shape

    atlas_ctseg = None
    atlas_region_masks: List[Tuple[str, np.ndarray]] = []
    if args.use_atlas_ctseg_features:
        atlas_ctseg_nib = nib.load(args.atlas_ctseg_path)
        atlas_ctseg = np.rint(atlas_ctseg_nib.get_fdata()).astype(np.int32)
        if atlas_ctseg.shape != atlas_ct.shape:
            raise RuntimeError(
                f'Atlas CTSeg shape mismatch: got {atlas_ctseg.shape}, expected {atlas_ct.shape} from atlas CT.'
            )
        atlas_region_masks = build_atlas_region_masks(
            atlas_ctseg,
            atlas_ct,
            args.atlas_region_min_voxels,
        )
        print(f'Atlas CTSeg: {args.atlas_ctseg_path}')
        print(f'Atlas CTSeg biomarker regions retained: {len(atlas_region_masks)}')

    rows = filter_rows(load_clinical_rows(args.clinical_xlsx, args.clinical_sheet), args.task)
    clinical_by_id = {clean_text(row.get('PatientID')): row for row in rows if clean_text(row.get('PatientID'))}

    case_ids_all = collect_case_ids(args.preprocessed_dir, require_seg=True)
    case_ids = [case_id for case_id in case_ids_all if case_id in clinical_by_id]
    if int(args.max_cases) > 0:
        case_ids = case_ids[:int(args.max_cases)]
    if not case_ids:
        raise RuntimeError('No HECKTOR cases matched between preprocessed images and workbook.')
    print(f'Clinical rows after task filter: {len(rows)}')
    print(f'Cases with CT/PET/SEG and clinical data: {len(case_ids)}')

    pet_radiomics_extractor = None
    ct_radiomics_extractor = None
    if args.use_pet_radiomics:
        pet_radiomics_extractor = build_radiomics_extractor(args.pet_radiomics_bin_width, PET_RADIOMICS_CLASSES, args.radiomics_force_2d)
    if args.use_ct_radiomics:
        ct_radiomics_extractor = build_radiomics_extractor(args.ct_radiomics_bin_width, CT_RADIOMICS_CLASSES, args.radiomics_force_2d)

    dataset = Hecktor25Dataset(args.preprocessed_dir, case_ids)
    loader = DataLoader(dataset, batch_size=args.batch_size, shuffle=False, num_workers=args.num_workers, pin_memory=(device.type == 'cuda'))

    sitreg_model = create_model(device, INPUT_SHAPE)
    model = TemplateCreation(sitreg_model, INPUT_SHAPE, use_sitreg=True).to(device)
    ckpt = torch.load(checkpoint_path, map_location='cpu')
    state_dict = ckpt.get('state_dict', ckpt)
    model.load_state_dict(state_dict)
    model.eval()

    spatial_trans = SpatialTransformer((H, W, D), mode='bilinear').to(device)
    spatial_trans_nn = SpatialTransformer((H, W, D), mode='nearest').to(device)
    atlas_ct_t = torch.from_numpy(atlas_ct[None, None, ...]).to(device).float()
    mask_3d = torch.from_numpy((atlas_ct > 0.01) | (atlas_pet > 0.01)).to(device)
    keep_mask = mask_3d.detach().cpu().numpy().astype(bool)

    baseline_X_list: List[np.ndarray] = []
    feature_rows: List[np.ndarray] = []
    feature_names: Optional[List[str]] = None
    clinical_rows: List[np.ndarray] = []
    patient_ids: List[str] = []
    time_to_event: List[float] = []
    event_indicator: List[int] = []
    pet_volumes: List[np.ndarray] = []
    tumor_masks: List[np.ndarray] = []
    tumor_count_vol = np.zeros((H, W, D), dtype=np.float32)
    reference_methods: List[str] = []
    reference_values: List[float] = []
    reference_voxels: List[int] = []
    lesion_counts: List[int] = []

    with torch.no_grad():
        for batch_idx, batch in enumerate(loader, start=1):
            case_id = batch['case_id'][0] if isinstance(batch['case_id'], (list, tuple)) else batch['case_id']
            row = clinical_by_id.get(str(case_id))
            if row is None:
                continue
            t_relapse = to_float(row.get('RFS'))
            e_relapse = to_int(row.get('Relapse'))
            if not np.isfinite(t_relapse) or e_relapse not in (0, 1):
                print(f'Skip {case_id}: invalid RFS/Relapse outcome.')
                continue

            ct_norm = batch['ct_norm'].to(device).float()
            ct_raw = batch['ct_raw'].to(device).float()
            pet_raw = batch['pet'].to(device).float()
            seg_raw = batch['seg'].to(device).float()
            ctseg_raw = batch['ctseg'].to(device).float()
            atlas_ct_rep = atlas_ct_t.repeat(ct_norm.shape[0], 1, 1, 1, 1)
            outputs = model((atlas_ct_rep, ct_norm))
            if len(outputs) == 5:
                _, _, pos_flow, neg_flow, _ = outputs
            else:
                _, _, pos_flow, neg_flow = outputs[:4]

            pet_warp = spatial_trans(pet_raw, neg_flow)[0, 0]
            ct_warp = spatial_trans(ct_raw, neg_flow)[0, 0]
            seg_warp = spatial_trans_nn(seg_raw, neg_flow)[0, 0]
            ctseg_warp = spatial_trans_nn(ctseg_raw, neg_flow)[0, 0]
            pet_smooth = smooth_inside_mask(pet_warp, mask_3d.float(), sigma=args.smooth_sigma)
            pet_smooth_np = pet_smooth.detach().cpu().numpy()
            ctseg_np = ctseg_warp.detach().cpu().numpy()
            ref_value = np.nan
            ref_method = 'none'
            ref_nvox = 0
            if args.pet_normalization != 'none':
                ref_value, ref_method, ref_nvox = compute_pet_reference(
                    pet_smooth_np,
                    ctseg_np,
                    mode=args.pet_normalization,
                    min_voxels=args.reference_min_voxels,
                    erosion_iters=args.reference_erosion_iters,
                )
            if np.isfinite(ref_value) and ref_value > 0:
                pet_metric = pet_smooth / float(ref_value)
            else:
                pet_metric = pet_smooth
            body_vals = pet_smooth[mask_3d.bool()]
            if body_vals.numel() == 0:
                print(f'Skip {case_id}: empty atlas body mask.')
                continue

            xvec = pet_metric.reshape(-1)[mask_3d.reshape(-1).bool()].detach().cpu().numpy()
            pet_metric_np = pet_metric.detach().cpu().numpy()
            ct_warp_np = ct_warp.detach().cpu().numpy()
            case_feature_parts: List[np.ndarray] = []
            case_feature_names: List[str] = []
            tumor_mask_np = np.zeros((H, W, D), dtype=bool)
            lesion_count = 0

            if args.use_tumor_features or args.use_pet_radiomics or args.use_ct_radiomics or args.use_atlas_tumor_overlap_features:
                tumor_mask_np, lesion_count = sanitize_tumor_mask(
                    seg_warp.detach().cpu().numpy(),
                    args.lesion_min_voxels,
                )
            if (args.use_pet_radiomics or args.use_ct_radiomics) and not tumor_mask_np.any():
                print(f'Skip {case_id}: empty tumor mask after lesion filtering; radiomics requested.')
                continue

            if args.use_tumor_features:
                lesion_feats, lesion_names, tumor_mask_np = extract_tumor_features(
                    pet_metric_np,
                    ct_warp_np,
                    tumor_mask_np.astype(np.float32),
                    atlas_affine,
                    args.lesion_min_voxels,
                )
                case_feature_parts.append(lesion_feats)
                case_feature_names.extend(lesion_names)

            if args.use_pet_radiomics:
                if tumor_mask_np.any():
                    pet_rad_feats, pet_rad_names = extract_radiomics_features_xyz(
                        pet_metric_np,
                        tumor_mask_np,
                        atlas_affine,
                        pet_radiomics_extractor,
                        'pet',
                    )
                else:
                    pet_rad_feats, pet_rad_names = np.zeros((0,), dtype=float), []
                case_feature_parts.append(pet_rad_feats)
                case_feature_names.extend(pet_rad_names)

            if args.use_ct_radiomics:
                if tumor_mask_np.any():
                    ct_rad_feats, ct_rad_names = extract_radiomics_features_xyz(
                        ct_warp_np,
                        tumor_mask_np,
                        atlas_affine,
                        ct_radiomics_extractor,
                        'ct',
                    )
                else:
                    ct_rad_feats, ct_rad_names = np.zeros((0,), dtype=float), []
                case_feature_parts.append(ct_rad_feats)
                case_feature_names.extend(ct_rad_names)

            if args.use_atlas_ctseg_features and atlas_region_masks:
                atlas_region_feats, atlas_region_names = extract_atlas_region_features(
                    pet_metric_np,
                    ct_warp_np,
                    atlas_region_masks,
                )
                case_feature_parts.append(atlas_region_feats)
                case_feature_names.extend(atlas_region_names)

            if args.use_atlas_burden_features and atlas_region_masks:
                atlas_burden_feats, atlas_burden_names = extract_atlas_burden_features(
                    pet_metric_np,
                    atlas_region_masks,
                )
                case_feature_parts.append(atlas_burden_feats)
                case_feature_names.extend(atlas_burden_names)

            if args.use_atlas_tumor_overlap_features and atlas_region_masks:
                atlas_overlap_feats, atlas_overlap_names = extract_atlas_tumor_overlap_features(
                    pet_metric_np,
                    tumor_mask_np,
                    atlas_region_masks,
                )
                case_feature_parts.append(atlas_overlap_feats)
                case_feature_names.extend(atlas_overlap_names)

            if feature_names is None:
                feature_names = case_feature_names
            elif feature_names != case_feature_names:
                raise RuntimeError(f'Feature layout mismatch encountered for case {case_id}.')
            feats_full = np.concatenate(case_feature_parts, axis=0) if case_feature_parts else np.zeros((0,), dtype=float)

            baseline_X_list.append(xvec)
            feature_rows.append(feats_full)
            clinical_rows.append(build_clinical_matrix(row))
            patient_ids.append(str(case_id))
            time_to_event.append(float(t_relapse))
            event_indicator.append(int(e_relapse))
            pet_np = pet_metric.detach().cpu().numpy().astype(np.float32)
            pet_volumes.append(pet_np)
            tumor_masks.append(tumor_mask_np.astype(bool))
            if args.use_tumor_features:
                tumor_count_vol += tumor_mask_np.astype(np.float32)
            reference_methods.append(ref_method)
            reference_values.append(float(ref_value) if np.isfinite(ref_value) else np.nan)
            reference_voxels.append(int(ref_nvox))
            lesion_counts.append(int(lesion_count))
            print(
                f'Processed {batch_idx}/{len(loader)}: {case_id} | RFS={t_relapse} days | '
                f'event={e_relapse} | lesions={lesion_count} | '
                f'ref={ref_method} ({ref_value if np.isfinite(ref_value) else float("nan"):.3f}, n={ref_nvox})'
            )

    if not baseline_X_list:
        raise RuntimeError('No analyzable cases remained after filtering.')

    Ximg = np.stack(baseline_X_list, axis=0)
    Xclin = np.stack(clinical_rows, axis=0)
    atlas_X = np.vstack(feature_rows)
    time_np = np.asarray(time_to_event, dtype=float)
    event_np = np.asarray(event_indicator, dtype=int)
    mean_vec, std_vec = compute_atlas_z_stats(Ximg)

    if args.use_tumor_features:
        feature_rows, feature_names, mean_vec, std_vec = add_z_features(
            list(atlas_X),
            feature_names or [],
            Ximg,
            tumor_masks,
            keep_mask,
            (H, W, D),
            atlas_affine,
            args.z_threshold,
        )
        atlas_X = np.vstack(feature_rows)

    if args.use_atlas_z_features and atlas_region_masks:
        feature_rows, feature_names, mean_vec, std_vec = add_atlas_z_summary_features(
            list(atlas_X),
            feature_names or [],
            Ximg,
            tumor_masks,
            atlas_region_masks,
            keep_mask,
            (H, W, D),
            atlas_affine,
            args.z_threshold,
        )
        atlas_X = np.vstack(feature_rows)

    if not args.use_tumor_features and feature_names is not None:
        atlas_X, feature_names = filter_feature_columns(
            atlas_X,
            feature_names,
            excluded_prefixes=('tumor_',),
        )
        stale_tumor_map = os.path.join(args.output_dir, 'tumor_prevalence_map.nii.gz')
        if os.path.exists(stale_tumor_map):
            os.remove(stale_tumor_map)

    save_vector_as_vol(mean_vec, os.path.join(args.output_dir, 'pet_atlas_mean.nii.gz'), atlas_affine, (H, W, D), keep_mask)
    save_vector_as_vol(std_vec, os.path.join(args.output_dir, 'pet_atlas_std.nii.gz'), atlas_affine, (H, W, D), keep_mask)
    if args.use_tumor_features and len(tumor_masks) > 0:
        save_volume((tumor_count_vol / float(len(tumor_masks))).astype(np.float32), os.path.join(args.output_dir, 'tumor_prevalence_map.nii.gz'), atlas_affine)

    ref_csv = os.path.join(args.output_dir, 'pet_reference_summary.csv')
    with open(ref_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['PatientID', 'reference_method', 'reference_value', 'reference_voxels'])
        for patient_id, method, value, nvox in zip(patient_ids, reference_methods, reference_values, reference_voxels):
            writer.writerow([patient_id, method, value, nvox])

    case_feature_csv = os.path.join(args.output_dir, 'imaging_case_features.csv')
    write_case_feature_table(case_feature_csv, patient_ids, time_np, event_np, Xclin, atlas_X, feature_names or [])

    print(f'Dataset for analysis: N={Ximg.shape[0]}, Vmask={Ximg.shape[1]}, F={atlas_X.shape[1]}')

    Xclin_z, _, _ = zscore_matrix(Xclin)
    b_clin, _, _, _ = cox_fit(Xclin_z, time_np, event_np, l2=1e-4, max_iter=60)
    risk_clin = Xclin_z @ np.nan_to_num(b_clin, nan=0.0, posinf=0.0, neginf=0.0)
    c_clin = concordance_index(time_np, event_np, risk_clin)
    plot_km_with_ticks_two_groups(time_np, event_np, risk_clin >= np.nanmedian(risk_clin), os.path.join(args.output_dir, 'km_clinical_median.png'), 'KM: clinical model (median split)')

    feat_stats = compute_univariate_feature_stats(
        atlas_X,
        time_np,
        event_np,
        args.min_coverage,
        args.min_events,
    )
    feat_beta = feat_stats['beta']
    feat_z = feat_stats['z']
    feat_p = feat_stats['p']
    feat_cov = feat_stats['coverage']
    feat_evt = feat_stats['events']
    q_feat05 = feat_stats['q05']
    q_feat10 = feat_stats['q10']
    feat_std = feat_stats['std']

    feat_csv = os.path.join(args.output_dir, 'imaging_features_univariate_cox.csv')
    with open(feat_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['feature', 'beta', 'HR', 'z', 'p', 'q05', 'q10', 'coverage', 'events', 'std'])
        for k, nm in enumerate(feature_names or [f'feat_{k}' for k in range(len(feat_beta))]):
            hr = np.exp(np.clip(feat_beta[k], -50.0, 50.0)) if np.isfinite(feat_beta[k]) else np.nan
            writer.writerow([nm, feat_beta[k], hr, feat_z[k], feat_p[k], q_feat05[k], q_feat10[k], int(feat_cov[k]), int(feat_evt[k]), feat_std[k]])

    sel, ranked_pruned, selection_used = select_feature_indices(
        feat_p,
        q_feat05,
        q_feat10,
        atlas_X,
        args.feature_selection,
        args.topk,
        args.drop_correlation_threshold,
        args.max_selected_features,
    )
    Xfeat_z, _, _ = zscore_matrix(atlas_X[:, sel]) if sel.size > 0 else (np.zeros((atlas_X.shape[0], 0)), None, None)
    X_clin_only = np.nan_to_num(Xclin_z, nan=0.0, posinf=0.0, neginf=0.0)
    X_with_feats = np.hstack([X_clin_only, np.nan_to_num(Xfeat_z, nan=0.0, posinf=0.0, neginf=0.0)]) if Xfeat_z.shape[1] > 0 else X_clin_only
    b_cf, _, _, _ = cox_fit(X_with_feats, time_np, event_np, l2=1e-3, max_iter=80)
    risk_all = X_with_feats @ np.nan_to_num(b_cf, nan=0.0, posinf=0.0, neginf=0.0)
    c_all = concordance_index(time_np, event_np, risk_all)
    if Xfeat_z.shape[1] > 0:
        X_feat_only = np.nan_to_num(Xfeat_z, nan=0.0, posinf=0.0, neginf=0.0)
        b_feat, _, _, _ = cox_fit(X_feat_only, time_np, event_np, l2=1e-3, max_iter=80)
        risk_feat = X_feat_only @ np.nan_to_num(b_feat, nan=0.0, posinf=0.0, neginf=0.0)
        c_feat = concordance_index(time_np, event_np, risk_feat)
        plot_km_with_ticks_two_groups(time_np, event_np, risk_feat >= np.nanmedian(risk_feat), os.path.join(args.output_dir, 'km_imaging_median.png'), 'KM: imaging score (median split)')
    else:
        c_feat = np.nan
    plot_km_with_ticks_two_groups(time_np, event_np, risk_all >= np.nanmedian(risk_all), os.path.join(args.output_dir, 'km_clinical_plus_imaging_median.png'), 'KM: clinical + imaging (median split)')

    voxelwise_report = None
    if args.voxelwise:
        voxelwise_report = run_voxelwise_cox(
            Ximg,
            time_np,
            event_np,
            keep_mask,
            atlas_affine,
            (H, W, D),
            args.output_dir,
            args.voxel_chunk,
            args.min_coverage,
            args.min_events,
        )

    try:
        idx_hi = int(np.nanargmax(risk_all))
        idx_lo = int(np.nanargmin(risk_all))
        mean_vol = np.full((H, W, D), np.nan, dtype=np.float32)
        std_vol = np.ones((H, W, D), dtype=np.float32)
        mean_vol.reshape(-1)[keep_mask.reshape(-1)] = mean_vec
        std_vol.reshape(-1)[keep_mask.reshape(-1)] = std_vec
        z_hi = (pet_volumes[idx_hi] - mean_vol) / std_vol
        z_lo = (pet_volumes[idx_lo] - mean_vol) / std_vol
        z_hi[~keep_mask] = 0.0
        z_lo[~keep_mask] = 0.0
        special_dir = os.path.join(args.output_dir, 'special_cases')
        os.makedirs(special_dir, exist_ok=True)
        save_volume(pet_volumes[idx_hi], os.path.join(special_dir, f'{patient_ids[idx_hi]}_PET_high_risk.nii.gz'), atlas_affine)
        save_volume(z_hi, os.path.join(special_dir, f'{patient_ids[idx_hi]}_Z_high_risk.nii.gz'), atlas_affine)
        save_volume(pet_volumes[idx_lo], os.path.join(special_dir, f'{patient_ids[idx_lo]}_PET_low_risk.nii.gz'), atlas_affine)
        save_volume(z_lo, os.path.join(special_dir, f'{patient_ids[idx_lo]}_Z_low_risk.nii.gz'), atlas_affine)
    except Exception as exc:
        print(f'Skipped special-case exports: {exc}')

    with open(os.path.join(args.output_dir, 'cox_summary.txt'), 'w') as f:
        f.write(f'Task filter={args.task}\n')
        f.write(f'PET normalization={args.pet_normalization}\n')
        f.write(f'Tumor feature extraction enabled={args.use_tumor_features}\n')
        f.write(f'Atlas CTSeg biomarker features enabled={args.use_atlas_ctseg_features}\n')
        f.write(f'PET radiomics enabled={args.use_pet_radiomics}\n')
        f.write(f'CT radiomics enabled={args.use_ct_radiomics}\n')
        f.write(f'PET radiomics bin width={args.pet_radiomics_bin_width}\n')
        f.write(f'CT radiomics bin width={args.ct_radiomics_bin_width}\n')
        f.write(f'Radiomics force 2D={args.radiomics_force_2d}\n')
        f.write(f'Redundancy-pruning correlation threshold={args.drop_correlation_threshold:.3f}\n')
        f.write(f'Max selected imaging features={args.max_selected_features}\n')
        if args.use_atlas_ctseg_features:
            f.write(f'Atlas CTSeg path={args.atlas_ctseg_path}\n')
            f.write(f'Atlas CTSeg regions used={len(atlas_region_masks)}\n')
        f.write(f'Patients N={Ximg.shape[0]}, Vmask={Ximg.shape[1]}, Features={atlas_X.shape[1]}\n')
        f.write(f'Events={int(event_np.sum())}, Median RFS={np.nanmedian(time_np):.1f} days\n')
        if lesion_counts:
            f.write(f'Median retained lesion count={np.median(np.asarray(lesion_counts, dtype=float)):.1f}\n')
        if reference_methods:
            method_counts: Dict[str, int] = {}
            for method in reference_methods:
                method_counts[method] = method_counts.get(method, 0) + 1
            f.write('Reference method counts=' + ', '.join(f'{key}:{value}' for key, value in sorted(method_counts.items())) + '\n')
            finite_refs = np.asarray(reference_values, dtype=float)
            if np.isfinite(finite_refs).any():
                f.write(f'Median reference value={np.nanmedian(finite_refs):.3f}\n')
        f.write(f'Clinical-only C-index={c_clin:.3f}\n')
        if args.feature_selection == 'topk':
            selection_desc = f'topk (k={args.topk}, selected={sel.size})'
        elif args.feature_selection == 'conventional':
            selection_desc = f'conventional [{selection_used}] (selected={sel.size})'
        else:
            selection_desc = f'{args.feature_selection} (selected={sel.size})'
        f.write(f'Feature selection mode={selection_desc}\n')
        f.write(f'Correlation-pruned ranked candidates={ranked_pruned.size}\n')
        f.write(f'Clinical+imaging C-index={c_all:.3f}\n')
        f.write(f'Imaging-only C-index={c_feat:.3f}\n')
        f.write('Clinical covariates included: age, gender, tobacco, alcohol, performance status, treatment, T/N/M stage, HPV, CenterID.\n')
        if sel.size > 0 and feature_names is not None:
            selected_names = [feature_names[k] for k in sel]
            f.write('Selected imaging features=' + ', '.join(selected_names) + '\n')
        if voxelwise_report is None:
            f.write('Voxelwise Cox: skipped\n')
        else:
            sig_count = int(np.isfinite(voxelwise_report['q05']).sum() and voxelwise_report['sig05'].sum())
            valid_fit = voxelwise_report['valid_fit']
            cover = voxelwise_report['cover']
            events_cover = voxelwise_report['events_cover']
            f.write(f'Voxelwise Cox significant voxels (FDR q<=0.05)={sig_count}\n')
            f.write(f'Voxelwise valid fits={int(valid_fit.sum())}/{int(valid_fit.size)}\n')
            if valid_fit.any():
                f.write(f'Median voxel coverage={np.nanmedian(cover[valid_fit]):.1f}\n')
                f.write(f'Median voxel events={np.nanmedian(events_cover[valid_fit]):.1f}\n')

    print('Saved outputs to:', args.output_dir)
    print(f'Clinical-only C-index={c_clin:.3f}')
    print(f'Clinical + imaging C-index={c_all:.3f}')
    print(f'Imaging-only C-index={c_feat:.3f}')
    print(f'Saved case feature table: {case_feature_csv}')
    print(f'Saved feature table: {feat_csv}')
    print(f'Saved PET reference summary: {ref_csv}')


if __name__ == '__main__':
    if torch.cuda.is_available():
        print('Number of GPU:', torch.cuda.device_count())
        for gpu_idx in range(torch.cuda.device_count()):
            print(f'     GPU #{gpu_idx}: {torch.cuda.get_device_name(gpu_idx)}')
    print('CUDA available?', torch.cuda.is_available())
    torch.manual_seed(42)
    np.random.seed(42)
    main()
