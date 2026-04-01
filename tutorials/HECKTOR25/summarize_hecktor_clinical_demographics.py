#!/usr/bin/env python3
import argparse
import csv
import math
import statistics
from collections import Counter
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


DEFAULT_XLSX = "/scratch2/jchen/DATA/HECKTOR25/HECKTOR_2025_Training_EHR_with_Data_Dictionary.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Summarize the HECKTOR 2025 training clinical workbook into a paper-style demographics table."
        )
    )
    parser.add_argument(
        "--input-xlsx",
        default=DEFAULT_XLSX,
        help="Path to the HECKTOR Excel workbook.",
    )
    parser.add_argument(
        "--sheet",
        default="Data",
        help="Workbook sheet containing patient-level clinical data.",
    )
    parser.add_argument(
        "--task",
        choices=["all", "1", "2", "3"],
        default="all",
        help="Optional HECKTOR task subset filter.",
    )
    parser.add_argument(
        "--output-format",
        choices=["text", "markdown", "csv"],
        default="text",
        help="Output format.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Optional output file path.",
    )
    return parser.parse_args()


def load_rows(path: str, sheet_name: str) -> List[Dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows: List[Dict[str, object]] = []
    for values in rows[1:]:
        row = {header[index]: values[index] for index in range(len(header))}
        data_rows.append(row)
    return data_rows


def filter_rows(rows: Sequence[Dict[str, object]], task: str) -> List[Dict[str, object]]:
    if task == "all":
        return list(rows)
    key = f"Task {task}"
    filtered: List[Dict[str, object]] = []
    for row in rows:
        if to_int(row.get(key)) == 1:
            filtered.append(row)
    return filtered


def to_float(value: object) -> float:
    if value is None or value == "":
        return math.nan
    try:
        return float(value)
    except Exception:
        return math.nan


def to_int(value: object) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def valid_numbers(values: Iterable[float]) -> List[float]:
    return [value for value in values if not math.isnan(value)]


def mean_sd_text(values: Iterable[float], decimals: int = 2) -> str:
    data = valid_numbers(values)
    if not data:
        return "NA"
    mean_value = statistics.mean(data)
    sd_value = statistics.stdev(data) if len(data) > 1 else 0.0
    return f"{mean_value:.{decimals}f} ± {sd_value:.{decimals}f}"


def median_range_text(values: Iterable[float], decimals: int = 2) -> str:
    data = valid_numbers(values)
    if not data:
        return "NA"
    return f"{statistics.median(data):.{decimals}f} ({min(data):.{decimals}f}–{max(data):.{decimals}f})"


def map_gender(value: object) -> str:
    code = to_int(value)
    if code == 0:
        return "Female"
    if code == 1:
        return "Male"
    return "Missing"


def map_binary_yes_no(value: object) -> str:
    code = to_int(value)
    if code == 0:
        return "No"
    if code == 1:
        return "Yes"
    return "Missing"


def map_treatment(value: object) -> str:
    code = to_int(value)
    if code == 0:
        return "Radiotherapy only"
    if code == 1:
        return "Chemoradiotherapy"
    return "Missing"


def map_hpv(value: object) -> str:
    code = to_int(value)
    if code == 0:
        return "Code 0"
    if code == 1:
        return "Code 1"
    return "Missing"


def map_relapse(value: object) -> str:
    code = to_int(value)
    if code == 0:
        return "No relapse / censored"
    if code == 1:
        return "Relapse"
    return "Missing"


def summarize(rows: Sequence[Dict[str, object]], task: str) -> Tuple[List[Tuple[str, str]], List[str]]:
    ages = [to_float(row.get("Age")) for row in rows]
    rfs_days = [to_float(row.get("RFS")) for row in rows]

    gender_counts = Counter(map_gender(row.get("Gender")) for row in rows)
    tobacco_counts = Counter(map_binary_yes_no(row.get("Tobacco Consumption")) for row in rows)
    alcohol_counts = Counter(map_binary_yes_no(row.get("Alcohol Consumption")) for row in rows)
    performance_counts = Counter(
        f"{int(value)}" if (value := to_int(row.get("Performance Status"))) is not None else "Missing"
        for row in rows
    )
    treatment_counts = Counter(map_treatment(row.get("Treatment")) for row in rows)
    t_stage_counts = Counter(clean_text(row.get("T-stage")) or "Missing" for row in rows)
    n_stage_counts = Counter(clean_text(row.get("N-stage")) or "Missing" for row in rows)
    m_stage_counts = Counter(clean_text(row.get("M-stage")) or "Missing" for row in rows)
    hpv_counts = Counter(map_hpv(row.get("HPV Status")) for row in rows)
    relapse_counts = Counter(map_relapse(row.get("Relapse")) for row in rows)
    center_counts = Counter(str(to_int(row.get("CenterID"))) if to_int(row.get("CenterID")) is not None else "Missing" for row in rows)

    table: List[Tuple[str, str]] = []
    cohort_label = "HECKTOR 2025 cohort" if task == "all" else f"HECKTOR 2025 Task {task} subset"
    table.append((cohort_label, str(len(rows))))
    table.append(("Age (y) (mean ± SD)", mean_sd_text(ages)))
    table.append(("Sex", ""))
    for label in ["Male", "Female", "Missing"]:
        table.append((f"  {label}", str(gender_counts.get(label, 0))))

    table.append(("Tobacco consumption history", ""))
    for label in ["Yes", "No", "Missing"]:
        table.append((f"  {label}", str(tobacco_counts.get(label, 0))))

    table.append(("Alcohol consumption history", ""))
    for label in ["Yes", "No", "Missing"]:
        table.append((f"  {label}", str(alcohol_counts.get(label, 0))))

    table.append(("Performance status", ""))
    for label in ["0", "1", "2", "3", "4", "Missing"]:
        table.append((f"  {label}", str(performance_counts.get(label, 0))))

    table.append(("Treatment", ""))
    for label in ["Radiotherapy only", "Chemoradiotherapy", "Missing"]:
        table.append((f"  {label}", str(treatment_counts.get(label, 0))))

    table.append(("T-stage", ""))
    for label in ["T0", "T1", "T2", "T3", "T4", "Missing"]:
        table.append((f"  {label}", str(t_stage_counts.get(label, 0))))

    table.append(("N-stage", ""))
    for label in ["N0", "N1", "N2", "N3", "Missing"]:
        table.append((f"  {label}", str(n_stage_counts.get(label, 0))))

    table.append(("M-stage", ""))
    for label in ["M0", "M1", "Missing"]:
        table.append((f"  {label}", str(m_stage_counts.get(label, 0))))

    table.append(("HPV status", ""))
    for label in ["Code 0", "Code 1", "Missing"]:
        table.append((f"  {label}", str(hpv_counts.get(label, 0))))

    table.append(("Relapse status", ""))
    for label in ["Relapse", "No relapse / censored", "Missing"]:
        table.append((f"  {label}", str(relapse_counts.get(label, 0))))

    table.append(("RFS (days) median (range)", median_range_text(rfs_days, decimals=0)))

    table.append(("CenterID", ""))
    for label in sorted(center_counts.keys(), key=lambda item: (item == "Missing", item)):
        table.append((f"  {label}", str(center_counts.get(label, 0))))

    notes = [
        "The workbook already appears patient-level; each row is treated as one unique HECKTOR case.",
        "Gender coding follows the workbook dictionary: 0=female, 1=male.",
        "Treatment is labeled using the workbook dictionary order: 0=radiotherapy only, 1=chemoradiotherapy.",
        "HPV status is reported as code 0/code 1 because the data dictionary lists binary encoding but does not explicitly define polarity.",
        "RFS is summarized in days using the paired `Relapse` event indicator.",
    ]
    return table, notes


def render_text(table: Sequence[Tuple[str, str]], notes: Sequence[str]) -> str:
    width = max(len(label) for label, _ in table)
    lines: List[str] = []
    for label, value in table:
        if value:
            lines.append(f"{label.ljust(width)}  {value}")
        else:
            lines.append(label)
    if notes:
        lines.append("")
        lines.append("Notes")
        for note in notes:
            lines.append(f"- {note}")
    return "\n".join(lines)


def render_markdown(table: Sequence[Tuple[str, str]], notes: Sequence[str]) -> str:
    lines = ["| Variable | Value |", "| --- | --- |"]
    for label, value in table:
        lines.append(f"| {label} | {value} |")
    if notes:
        lines.append("")
        lines.append("**Notes**")
        for note in notes:
            lines.append(f"- {note}")
    return "\n".join(lines)


def write_csv(table: Sequence[Tuple[str, str]], output_path: str) -> None:
    with open(output_path, "w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Variable", "Value"])
        writer.writerows(table)


def main() -> None:
    args = parse_args()
    rows = load_rows(args.input_xlsx, args.sheet)
    rows = filter_rows(rows, args.task)
    table, notes = summarize(rows, args.task)

    if args.output_format == "csv":
        if not args.output:
            raise ValueError("--output is required when --output-format csv is used.")
        write_csv(table, args.output)
        return

    rendered = render_markdown(table, notes) if args.output_format == "markdown" else render_text(table, notes)
    if args.output:
        with open(args.output, "w") as handle:
            handle.write(rendered + "\n")
    else:
        print(rendered)


if __name__ == "__main__":
    main()
