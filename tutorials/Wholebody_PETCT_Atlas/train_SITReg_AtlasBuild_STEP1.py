from __future__ import annotations

import argparse
import csv
import glob
import math
import os
import random
import re
import sys
import time
from datetime import datetime
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

import nibabel as nib
import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as F
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from natsort import natsorted
from torch import optim
from torch.utils.data import BatchSampler, Sampler
from torch.utils.data import DataLoader, Dataset
from torch.utils.tensorboard import SummaryWriter

from MIR.deformation_regularizer import Grad3d
from MIR.image_similarity import SSIM3D
from MIR.models import EncoderFeatureExtractor, SITReg, SpatialTransformer, TemplateCreation
from MIR.models.SITReg import GroupNormalizerFactory, ReLUFactory
from MIR.models.SITReg.composable_mapping import DataFormat
from MIR.models.SITReg.deformation_inversion_layer.fixed_point_iteration import (
    AndersonSolver,
    AndersonSolverArguments,
    MaxElementWiseAbsStopCriterion,
    RelativeL2ErrorStopCriterion,
)
from MIR.utils import AverageMeter


AUTOPET_FDG_ROOT = Path("/scratch2/jchen/DATA/AutoPET/affine_aligned/network")
AUTOPET_FDG_SPLITS = ("train", "val", "test")
AUTOPET_FDG_METADATA = Path("/scratch2/jchen/DATA/AutoPET/Metadata.csv")
AUTOPET_PSMA_DIR = Path("/scratch2/jchen/DATA/PSMA_autoPET/Preprocessed/autoPET")
AUTOPET_PSMA_METADATA = Path("/scratch2/jchen/DATA/PSMA_autoPET/psma_metadata.csv")
JHU_FDG_BATCH7_DIR = Path("/scratch2/jchen/DATA/JHU_FDG/batch7/preprocessed")
JHU_FDG_BATCH8_DIR = Path("/scratch2/jchen/DATA/JHU_FDG/batch8/preprocessed")
JHU_FDG_CLINICAL = Path("/scratch/jchen/python_projects/clinical_reports/batch78_img_clinical_os_merged_annotated.csv")
JHU_FDG_WEIGHT_XLSX = Path("/scratch2/jchen/DATA/JHU_FDG/CCDA10838.xlsx")
JHU_PSMA_DIR = Path("/scratch2/jchen/DATA/PSMA_JHU/Preprocessed/JHU")

DEFAULT_DATASET_FLIP_AXES_TO_JHU_FDG: Dict[str, Tuple[int, ...]] = {
    "autopet_fdg": (1,),
    "autopet_psma": (),
    "jhu_fdg": (),
    "jhu_psma": (1,),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Train whole-body SITReg atlases across AutoPET FDG, AutoPET PSMA, JHU FDG, "
            "and JHU PSMA with shared sex-specific CT atlases and dataset-specific PET atlases."
        )
    )
    parser.add_argument("--output-root", default="wholebody_petct_sitreg_step1", help="Output folder name under the current working directory.")
    parser.add_argument("--batch-size", type=int, default=1, help="Training batch size. Batches larger than 1 are supported when all items in a batch share the same atlas keys.")
    parser.add_argument("--num-workers", type=int, default=4)
    parser.add_argument("--max-epochs", type=int, default=500)
    parser.add_argument("--lr-model", type=float, default=1e-4)
    parser.add_argument("--lr-atlas", type=float, default=1e-3)
    parser.add_argument("--val-fraction", type=float, default=0.001)
    parser.add_argument("--save-every", type=int, default=1)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--target-shape", nargs=3, type=int, default=[192, 192, 256])
    parser.add_argument("--weights", nargs=3, type=float, default=[1.0, 1.0, 1.0], help="Loss weights: image similarity, mean-stream penalty, deformation regularization.")
    parser.add_argument("--interleave-by", choices=["dataset", "pet_key"], default="pet_key", help="How to interleave samples during training.")
    parser.add_argument("--init-samples-per-dataset", type=int, default=200, help="Maximum number of training scans to sample from each dataset when estimating initial atlases.")
    parser.add_argument("--profile-timing", action="store_true", help="Print detailed timing diagnostics for data wait, loading, transfer, forward, loss, backward, optimizer, validation, and checkpointing.")
    parser.add_argument("--timing-log-every", type=int, default=20, help="When timing is enabled, print per-iteration timing every N training iterations.")
    parser.add_argument("--timing-warmup-steps", type=int, default=3, help="Ignore the first N training iterations for timing averages to reduce startup noise.")
    parser.add_argument("--amp", action="store_true", help="Enable automatic mixed precision for training and validation. Leave off to run full FP32.")
    parser.add_argument("--preview-save-every-steps", type=int, default=250, help="Save overwriteable preview atlas NIfTI files every N training steps. Set to 0 to disable step-wise preview saves.")
    parser.add_argument(
        "--dataset-flips",
        default="autopet_fdg:1,autopet_psma:none,jhu_fdg:none,jhu_psma:1",
        help="Dataset-specific voxel flips to align all volumes to the JHU FDG convention. Format: dataset:axis|axis,dataset:none",
    )
    parser.add_argument(
        "--disable-orientation-standardization",
        action="store_true",
        help="Disable canonical reorientation plus dataset-specific flips. Use only for debugging if the standardized orientation looks wrong.",
    )
    return parser.parse_args()


@dataclass(frozen=True)
class SampleRecord:
    stem: str
    dataset: str
    pet_group: str
    sex: str
    ct_path: str
    pet_path: str
    seg_path: Optional[str] = None
    pet_suv_multiplier: Optional[float] = None

    @property
    def ct_key(self) -> str:
        return self.sex

    @property
    def pet_key(self) -> str:
        return f"{self.pet_group}_{self.sex}"


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_sex(value: object) -> Optional[str]:
    text = clean_text(value).lower()
    if text in {"m", "male", "man", "men"}:
        return "men"
    if text in {"f", "female", "woman", "women"}:
        return "women"
    return None


def normalize_bool(value: object) -> Optional[bool]:
    text = clean_text(value).lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return None


def choose_first(record: Dict[str, object], keys: Sequence[str]) -> str:
    for key in keys:
        if key in record:
            text = clean_text(record.get(key))
            if text:
                return text
    return ""


def parse_flip_axes(text: str) -> Tuple[int, ...]:
    cleaned = clean_text(text).lower()
    if not cleaned or cleaned in {"none", "no", "off"}:
        return ()
    axes = []
    for token in cleaned.split("|"):
        axis = int(token)
        if axis not in {0, 1, 2}:
            raise ValueError(f"Flip axis must be 0, 1, or 2; got {axis}.")
        axes.append(axis)
    return tuple(axes)


def parse_dataset_flip_config(config_text: str) -> Dict[str, Tuple[int, ...]]:
    dataset_flips = dict(DEFAULT_DATASET_FLIP_AXES_TO_JHU_FDG)
    if not clean_text(config_text):
        return dataset_flips
    for item in config_text.split(","):
        entry = clean_text(item)
        if not entry:
            continue
        if ":" not in entry:
            raise ValueError(f"Invalid dataset flip entry '{entry}'. Expected dataset:axes.")
        dataset_name, axes_text = entry.split(":", 1)
        dataset_name = clean_text(dataset_name)
        if not dataset_name:
            raise ValueError(f"Invalid dataset flip entry '{entry}'. Dataset name is empty.")
        dataset_flips[dataset_name] = parse_flip_axes(axes_text)
    return dataset_flips


def format_dataset_flip_config(dataset_flips: Dict[str, Tuple[int, ...]]) -> str:
    parts = []
    for dataset_name in sorted(dataset_flips):
        axes = dataset_flips[dataset_name]
        axes_text = "none" if not axes else "|".join(str(axis) for axis in axes)
        parts.append(f"{dataset_name}:{axes_text}")
    return ", ".join(parts)


def summarize_orientation_policy(dataset_flips: Dict[str, Tuple[int, ...]], enabled: bool) -> None:
    if not enabled:
        print("Orientation standardization disabled; volumes will be used as stored on disk.")
        return
    print("Orientation standardization enabled (canonical header reorientation + dataset-specific voxel flips to JHU FDG convention):")
    for dataset_name in sorted(dataset_flips):
        axes = dataset_flips[dataset_name]
        axes_text = "none" if not axes else ", ".join(str(axis) for axis in axes)
        print(f"  {dataset_name}: flip axes {axes_text}")


def infer_fieldnames(path: Path) -> List[str]:
    csv.field_size_limit(sys.maxsize)
    with path.open("r", newline="") as handle:
        reader = csv.reader(handle)
        return next(reader)


def load_autopet_fdg_sex_map() -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    with AUTOPET_FDG_METADATA.open("r", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            subject_id = clean_text(row.get("Subject ID"))
            if not subject_id:
                continue
            stem = subject_id.replace("PETCT_", "")
            sex = normalize_sex(row.get("sex"))
            if sex is not None and stem not in mapping:
                mapping[stem] = sex
    return mapping


def build_autopet_fdg_records() -> List[SampleRecord]:
    sex_map = load_autopet_fdg_sex_map()
    records: List[SampleRecord] = []
    seen_stems: set[str] = set()
    for split in AUTOPET_FDG_SPLITS:
        split_dir = AUTOPET_FDG_ROOT / split
        for ct_path in sorted(split_dir.glob("*_CTRes.nii.gz")):
            stem = ct_path.name[: -len("_CTRes.nii.gz")]
            if stem in seen_stems:
                continue
            pet_path = split_dir / f"{stem}_SUV.nii.gz"
            sex = sex_map.get(stem)
            if sex is None or not pet_path.exists():
                continue
            seen_stems.add(stem)
            records.append(
                SampleRecord(
                    stem=stem,
                    dataset="autopet_fdg",
                    pet_group="fdg",
                    sex=sex,
                    ct_path=str(ct_path),
                    pet_path=str(pet_path),
                    seg_path=None,
                )
            )
    return records


def build_autopet_psma_records() -> List[SampleRecord]:
    records: List[SampleRecord] = []
    for ct_path in sorted(AUTOPET_PSMA_DIR.glob("*_CT.nii.gz")):
        stem = ct_path.name[: -len("_CT.nii.gz")]
        if stem.endswith("_CT_seg"):
            continue
        pet_path = AUTOPET_PSMA_DIR / f"{stem}_SUV.nii.gz"
        if not pet_path.exists():
            continue
        records.append(
            SampleRecord(
                stem=stem,
                dataset="autopet_psma",
                    pet_group="psma",
                sex="men",
                ct_path=str(ct_path),
                pet_path=str(pet_path),
                seg_path=None,
            )
        )
    return records


def parse_datetime_flexible(value: object) -> Optional[datetime]:
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y%m%d", "%m/%d/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_float_or_none(value: object) -> Optional[float]:
    text = clean_text(value)
    if not text or text.lower() in {"null", "none", "nan"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_injected_dose_to_bq(value: object) -> Optional[float]:
    text = clean_text(value).lower()
    if not text:
        return None
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(mci|uci|ci|mbq|kbq|bq)", text)
    if match is None:
        return None
    amount = float(match.group(1))
    unit = match.group(2)
    scale_by_unit = {
        "ci": 3.7e10,
        "mci": 3.7e7,
        "uci": 3.7e4,
        "mbq": 1.0e6,
        "kbq": 1.0e3,
        "bq": 1.0,
    }
    return amount * scale_by_unit[unit]


def load_jhu_fdg_weight_rows() -> Dict[str, List[Tuple[datetime, float]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(JHU_FDG_WEIGHT_XLSX, read_only=True, data_only=True)
    sheet = workbook["Weight"]
    rows_by_mrn: Dict[str, List[Tuple[datetime, float]]] = defaultdict(list)
    header = [clean_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {name: index for index, name in enumerate(header)}
    mrn_index = header_index.get("patient_mrn")
    exam_index = header_index.get("Exam_Started_Date")
    weight_index = header_index.get("Weight_in_kg")
    if mrn_index is None or exam_index is None or weight_index is None:
        raise RuntimeError("JHU FDG weight workbook is missing one of patient_mrn, Exam_Started_Date, or Weight_in_kg columns.")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        mrn = clean_text(row[mrn_index])
        exam_date = row[exam_index]
        weight_kg = parse_float_or_none(row[weight_index])
        if not mrn or exam_date is None or weight_kg is None or weight_kg <= 0:
            continue
        if isinstance(exam_date, datetime):
            exam_dt = exam_date
        else:
            exam_dt = parse_datetime_flexible(exam_date)
        if exam_dt is None:
            continue
        rows_by_mrn[mrn].append((exam_dt, weight_kg))

    for mrn in rows_by_mrn:
        rows_by_mrn[mrn].sort(key=lambda item: item[0])
    return rows_by_mrn


def find_nearest_weight_kg(
    rows_by_mrn: Dict[str, List[Tuple[datetime, float]]],
    mrn: str,
    study_date: Optional[datetime],
) -> Optional[float]:
    candidates = rows_by_mrn.get(mrn, [])
    if not candidates:
        return None
    if study_date is None:
        return candidates[-1][1]
    nearest_date, nearest_weight = min(candidates, key=lambda item: abs((item[0] - study_date).total_seconds()))
    return nearest_weight


def load_jhu_fdg_clinical_info() -> Tuple[Dict[str, Tuple[str, float]], Dict[str, Dict[str, object]]]:
    mapping: Dict[str, Tuple[str, float]] = {}
    status_by_session: Dict[str, Dict[str, object]] = {}
    weight_rows = load_jhu_fdg_weight_rows()
    csv.field_size_limit(sys.maxsize)
    with JHU_FDG_CLINICAL.open("r", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            session_id = choose_first(row, ["XNATSessionID", "patient_id", "XnatSessionID"])
            tracer = choose_first(row, ["tracer", "Tracer"])
            sex = normalize_sex(choose_first(row, ["demo_Gender", "Gender", "gender"]))
            mrn = choose_first(row, ["MRN", "patient_mrn"])
            study_date = parse_datetime_flexible(choose_first(row, ["StudyDate", "study_date"]))
            injected_dose_bq = parse_injected_dose_to_bq(row.get("injected_dose_with_units_llm"))
            weight_kg = find_nearest_weight_kg(weight_rows, mrn, study_date) if mrn else None
            if not session_id:
                continue
            if "fdg" not in tracer.lower():
                continue
            reasons: List[str] = []
            if sex is None:
                reasons.append("missing_sex")
            if injected_dose_bq is None or injected_dose_bq <= 0:
                reasons.append("missing_injected_dose")
            if weight_kg is None or weight_kg <= 0:
                reasons.append("missing_weight")
            status_by_session.setdefault(
                session_id,
                {
                    "mrn": mrn,
                    "study_date": study_date.isoformat() if study_date is not None else "",
                    "sex": sex or "",
                    "injected_dose_bq": injected_dose_bq,
                    "weight_kg": weight_kg,
                    "reasons": reasons,
                },
            )
            if reasons:
                continue
            pet_suv_multiplier = (weight_kg * 1000.0) / injected_dose_bq
            mapping.setdefault(session_id, (sex, pet_suv_multiplier))
    return mapping, status_by_session


def load_jhu_fdg_clinical_map() -> Dict[str, Tuple[str, float]]:
    mapping, _ = load_jhu_fdg_clinical_info()
    return mapping


def save_jhu_fdg_exclusion_report(output_path: Path) -> None:
    clinical_map, status_by_session = load_jhu_fdg_clinical_info()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    excluded_rows: List[List[object]] = []
    for root in (JHU_FDG_BATCH7_DIR, JHU_FDG_BATCH8_DIR):
        for ct_path in sorted(root.glob("*_CT.nii.gz")):
            stem = ct_path.name[: -len("_CT.nii.gz")]
            pet_path = root / f"{stem}_PET.nii.gz"
            status = status_by_session.get(stem, {})
            reasons: List[str] = []
            if not pet_path.exists():
                reasons.append("missing_pet_file")
            if stem not in status_by_session:
                reasons.append("missing_fdg_clinical_row")
            else:
                reasons.extend(status.get("reasons", []))
            if not reasons:
                continue
            unique_reasons = sorted(set(reasons))
            excluded_rows.append(
                [
                    stem,
                    str(root),
                    str(ct_path),
                    str(pet_path),
                    int(pet_path.exists()),
                    status.get("mrn", ""),
                    status.get("study_date", ""),
                    status.get("sex", ""),
                    status.get("injected_dose_bq", ""),
                    status.get("weight_kg", ""),
                    "|".join(unique_reasons),
                ]
            )
    with output_path.open("w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "stem",
                "source_root",
                "ct_path",
                "pet_path",
                "pet_exists",
                "mrn",
                "study_date",
                "sex",
                "injected_dose_bq",
                "weight_kg",
                "reasons",
            ]
        )
        writer.writerows(excluded_rows)
    print(f"Saved JHU FDG exclusion report with {len(excluded_rows)} rows to {output_path}")


def build_jhu_fdg_records() -> List[SampleRecord]:
    clinical_map = load_jhu_fdg_clinical_map()
    records: List[SampleRecord] = []
    for root in (JHU_FDG_BATCH7_DIR, JHU_FDG_BATCH8_DIR):
        for ct_path in sorted(root.glob("*_CT.nii.gz")):
            stem = ct_path.name[: -len("_CT.nii.gz")]
            pet_path = root / f"{stem}_PET.nii.gz"
            seg_path = root / f"{stem}_CTSeg.nii.gz"
            clinical_info = clinical_map.get(stem)
            if clinical_info is None or not pet_path.exists():
                continue
            sex, pet_suv_multiplier = clinical_info
            records.append(
                SampleRecord(
                    stem=stem,
                    dataset="jhu_fdg",
                    pet_group="fdg",
                    sex=sex,
                    ct_path=str(ct_path),
                    pet_path=str(pet_path),
                    seg_path=str(seg_path) if seg_path.exists() else None,
                    pet_suv_multiplier=pet_suv_multiplier,
                )
            )
    return records


def build_jhu_psma_records() -> List[SampleRecord]:
    records: List[SampleRecord] = []
    for ct_path in sorted(JHU_PSMA_DIR.glob("*_CT.nii.gz")):
        stem = ct_path.name[: -len("_CT.nii.gz")]
        if stem.endswith("_CT_seg"):
            continue
        pet_path = JHU_PSMA_DIR / f"{stem}_SUV.nii.gz"
        if not pet_path.exists():
            continue
        records.append(
            SampleRecord(
                stem=stem,
                dataset="jhu_psma",
                    pet_group="psma",
                sex="men",
                ct_path=str(ct_path),
                pet_path=str(pet_path),
                seg_path=None,
            )
        )
    return records


def build_records() -> List[SampleRecord]:
    records = []
    records.extend(build_autopet_fdg_records())
    records.extend(build_autopet_psma_records())
    records.extend(build_jhu_fdg_records())
    records.extend(build_jhu_psma_records())
    return records


def summarize_records(records: Sequence[SampleRecord]) -> None:
    dataset_counter = Counter(record.dataset for record in records)
    ct_counter = Counter(record.ct_key for record in records)
    pet_counter = Counter(record.pet_key for record in records)
    print("Loaded sample counts by dataset:")
    for key in sorted(dataset_counter):
        print(f"  {key}: {dataset_counter[key]}")
    print("Loaded sample counts by CT atlas:")
    for key in sorted(ct_counter):
        print(f"  {key}: {ct_counter[key]}")
    print("Loaded sample counts by PET atlas:")
    for key in sorted(pet_counter):
        print(f"  {key}: {pet_counter[key]}")


def print_dataset_shape_preflight(
    records: Sequence[SampleRecord],
    dataset: WholebodyAtlasDataset,
    seed: int,
    output_dir: Path,
) -> None:
    grouped: Dict[str, List[SampleRecord]] = defaultdict(list)
    for record in records:
        grouped[record.dataset].append(record)

    rng = random.Random(seed)
    output_dir.mkdir(parents=True, exist_ok=True)
    print("Preflight random shape check by dataset:")
    for dataset_name in sorted(grouped):
        record = rng.choice(grouped[dataset_name])
        raw_ct_shape = nib.load(record.ct_path).shape
        raw_pet_shape = nib.load(record.pet_path).shape
        standardized_ct_shape = dataset.load_array(record.ct_path, record).shape
        standardized_pet_shape = dataset.load_array(record.pet_path, record).shape
        ct_tensor = dataset.load_volume(record.ct_path, is_pet=False, record=record)
        pet_tensor = dataset.load_volume(record.pet_path, is_pet=True, record=record)
        tensor_ct_shape = tuple(ct_tensor.shape)
        tensor_pet_shape = tuple(pet_tensor.shape)
        print(
            f"  {dataset_name}: stem={record.stem} "
            f"CT raw={raw_ct_shape} standardized={standardized_ct_shape} tensor={tensor_ct_shape} | "
            f"PET raw={raw_pet_shape} standardized={standardized_pet_shape} tensor={tensor_pet_shape}"
        )
        save_preflight_preview(
            ct_tensor[0].detach().cpu().numpy(),
            pet_tensor[0].detach().cpu().numpy(),
            output_dir / f"{dataset_name}__{record.stem}.png",
            title=f"{dataset_name} | {record.stem}",
        )


def _extract_middle_slices(volume: np.ndarray) -> List[np.ndarray]:
    if volume.ndim != 3:
        raise ValueError(f"Expected 3D volume for preview, got shape {volume.shape}")
    mid0 = volume.shape[0] // 2
    mid1 = volume.shape[1] // 2
    mid2 = volume.shape[2] // 2
    return [
        np.rot90(volume[mid0, :, :]),
        np.rot90(volume[:, mid1, :]),
        np.rot90(volume[:, :, mid2]),
    ]


def save_preflight_preview(ct_volume: np.ndarray, pet_volume: np.ndarray, output_path: Path, title: str) -> None:
    ct_slices = _extract_middle_slices(ct_volume)
    pet_slices = _extract_middle_slices(pet_volume)
    figure, axes = plt.subplots(2, 3, figsize=(12, 8))
    plane_titles = ["Axis 0 mid", "Axis 1 mid", "Axis 2 mid"]

    for column, plane_title in enumerate(plane_titles):
        axes[0, column].imshow(ct_slices[column], cmap="gray")
        axes[0, column].set_title(f"CT {plane_title}")
        axes[0, column].axis("off")

        axes[1, column].imshow(pet_slices[column], cmap="inferno")
        axes[1, column].set_title(f"PET {plane_title}")
        axes[1, column].axis("off")

    figure.suptitle(title)
    figure.tight_layout(rect=[0, 0, 1, 0.96])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(figure)


def save_manifest(records: Sequence[SampleRecord], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["stem", "dataset", "pet_group", "sex", "ct_path", "pet_path", "seg_path"])
        for record in records:
            writer.writerow([record.stem, record.dataset, record.pet_group, record.sex, record.ct_path, record.pet_path, record.seg_path or ""])


def split_train_val(records: Sequence[SampleRecord], val_fraction: float, seed: int) -> Tuple[List[SampleRecord], List[SampleRecord]]:
    rng = random.Random(seed)
    grouped: Dict[str, List[SampleRecord]] = defaultdict(list)
    for record in records:
        grouped[record.pet_key].append(record)
    train_records: List[SampleRecord] = []
    val_records: List[SampleRecord] = []
    for key in sorted(grouped):
        items = list(grouped[key])
        rng.shuffle(items)
        if val_fraction <= 0 or len(items) < 2:
            train_records.extend(items)
            continue
        val_count = max(1, int(round(len(items) * val_fraction)))
        val_count = min(val_count, len(items) - 1)
        val_records.extend(items[:val_count])
        train_records.extend(items[val_count:])
    return train_records, val_records


class WholebodyAtlasDataset(Dataset):
    def __init__(
        self,
        records: Sequence[SampleRecord],
        target_shape: Sequence[int],
        skip_log_path: Optional[Path] = None,
        dataset_flip_axes: Optional[Dict[str, Tuple[int, ...]]] = None,
        standardize_orientation: bool = True,
    ):
        self.records = list(records)
        self.target_shape = tuple(int(v) for v in target_shape)
        self._max_skip_attempts = max(1, len(self.records))
        self.skip_log_path = skip_log_path
        self._logged_skip_keys: set[Tuple[str, str]] = set()
        self.dataset_flip_axes = dict(dataset_flip_axes or DEFAULT_DATASET_FLIP_AXES_TO_JHU_FDG)
        self.standardize_orientation = standardize_orientation
        self._indices_by_pet_key: Dict[str, List[int]] = defaultdict(list)
        for index, record in enumerate(self.records):
            self._indices_by_pet_key[record.pet_key].append(index)

    def __len__(self) -> int:
        return len(self.records)

    @staticmethod
    def normalize_ct(image: np.ndarray) -> np.ndarray:
        image = np.clip(image, -300.0, 300.0)
        image_min = float(image.min())
        image_max = float(image.max())
        if image_max <= image_min:
            return np.zeros_like(image, dtype=np.float32)
        normalized = ((image - image_min) / (image_max - image_min)).astype(np.float32)
        return np.clip(normalized, 0.0, 1.0)

    @staticmethod
    def normalize_pet(image: np.ndarray) -> np.ndarray:
        image = np.asarray(image, dtype=np.float32)
        image = (image - 0.0) / 15.0
        return np.clip(image, 0.0, 1.0)

    @staticmethod
    def convert_pet_to_suv(image: np.ndarray, pet_suv_multiplier: float) -> np.ndarray:
        image = np.asarray(image, dtype=np.float32)
        image = np.clip(image, a_min=0.0, a_max=None)
        return (image * float(pet_suv_multiplier)).astype(np.float32)

    def load_array(self, path: str, record: SampleRecord) -> np.ndarray:
        image_nib = nib.load(path)
        if self.standardize_orientation:
            image_nib = nib.as_closest_canonical(image_nib, enforce_diag=False)
        image = image_nib.get_fdata()
        if self.standardize_orientation:
            for axis in self.dataset_flip_axes.get(record.dataset, ()):
                image = np.flip(image, axis=axis)
        return np.ascontiguousarray(image)

    def load_volume(self, path: str, is_pet: bool, record: SampleRecord) -> torch.Tensor:
        image = self.load_array(path, record)
        if is_pet:
            if record.pet_suv_multiplier is not None:
                image = self.convert_pet_to_suv(image, record.pet_suv_multiplier)
            image = self.normalize_pet(image)
        else:
            image = self.normalize_ct(image)
        tensor = torch.from_numpy(image)[None, None]
        if tuple(tensor.shape[-3:]) != self.target_shape:
            tensor = F.interpolate(tensor, size=self.target_shape, mode="trilinear", align_corners=False)
        tensor = torch.clamp(tensor, min=0.0, max=1.0)
        return tensor[0]

    def log_skip(self, record: SampleRecord, reason: str) -> None:
        if self.skip_log_path is None:
            return
        key = (record.stem, reason)
        if key in self._logged_skip_keys:
            return
        self._logged_skip_keys.add(key)
        row = [
            datetime.now().isoformat(timespec="seconds"),
            record.stem,
            record.dataset,
            record.pet_group,
            record.sex,
            record.ct_path,
            record.pet_path,
            record.seg_path or "",
            reason,
        ]
        with self.skip_log_path.open("a", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(row)

    def _get_same_pet_key_indices(self, index: int) -> List[int]:
        record = self.records[index]
        return self._indices_by_pet_key[record.pet_key]

    def _get_next_same_pet_key_index(self, index: int) -> int:
        candidate_indices = self._get_same_pet_key_indices(index)
        if len(candidate_indices) == 1:
            return index
        position = candidate_indices.index(index)
        return candidate_indices[(position + 1) % len(candidate_indices)]

    def _get_item(self, index: int, attempts: int) -> Dict[str, object]:
        record = self.records[index]
        same_key_indices = self._get_same_pet_key_indices(index)
        if attempts >= max(1, len(same_key_indices)):
            raise RuntimeError(f"Unable to load any valid sample for pet_key={record.pet_key}.")
        try:
            return {
                "ct": self.load_volume(record.ct_path, is_pet=False, record=record),
                "pet": self.load_volume(record.pet_path, is_pet=True, record=record),
                "ct_key": record.ct_key,
                "pet_key": record.pet_key,
                "stem": record.stem,
            }
        except (FileNotFoundError, OSError, ValueError) as exc:
            reason = str(exc)
            print(f"Skipping sample {record.stem}: {reason}")
            self.log_skip(record, reason)
            next_index = self._get_next_same_pet_key_index(index)
            return self._get_item(next_index, attempts + 1)

    def __getitem__(self, index: int) -> Dict[str, object]:
        start_time = time.perf_counter()
        sample = self._get_item(index, attempts=0)
        sample["sample_load_time_sec"] = time.perf_counter() - start_time
        return sample


class InterleavedGroupBatchSampler(BatchSampler):
    def __init__(self, records: Sequence[SampleRecord], group_by: str, batch_size: int, seed: int) -> None:
        if group_by not in {"dataset", "pet_key"}:
            raise ValueError(f"Unsupported group_by value: {group_by}")
        if batch_size < 1:
            raise ValueError("batch_size must be >= 1")
        self.records = list(records)
        self.group_by = group_by
        self.batch_size = batch_size
        self.seed = seed
        self.epoch = 0

    def set_epoch(self, epoch: int) -> None:
        self.epoch = epoch

    def __len__(self) -> int:
        total = 0
        grouped_indices: Dict[str, List[int]] = defaultdict(list)
        for index, record in enumerate(self.records):
            grouped_indices[getattr(record, self.group_by)].append(index)
        for indices in grouped_indices.values():
            total += math.ceil(len(indices) / self.batch_size)
        return total

    def __iter__(self) -> Iterator[List[int]]:
        rng = random.Random(self.seed + self.epoch)
        grouped_indices: Dict[str, List[int]] = defaultdict(list)
        for index, record in enumerate(self.records):
            group_key = getattr(record, self.group_by)
            grouped_indices[group_key].append(index)
        grouped_batches: Dict[str, List[List[int]]] = {}
        for group_key, indices in grouped_indices.items():
            rng.shuffle(indices)
            grouped_batches[group_key] = [
                indices[start : start + self.batch_size]
                for start in range(0, len(indices), self.batch_size)
            ]
        group_order = sorted(grouped_batches)
        rng.shuffle(group_order)

        max_len = max(len(batches) for batches in grouped_batches.values())
        interleaved: List[List[int]] = []
        for offset in range(max_len):
            current_groups = list(group_order)
            rng.shuffle(current_groups)
            for group_key in current_groups:
                batches = grouped_batches[group_key]
                if offset < len(batches):
                    interleaved.append(batches[offset])
        return iter(interleaved)


def initialize_skip_log(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["timestamp", "stem", "dataset", "pet_group", "sex", "ct_path", "pet_path", "seg_path", "reason"])


def sample_records_for_initialization(records: Sequence[SampleRecord], max_per_dataset: int, seed: int) -> List[SampleRecord]:
    if max_per_dataset <= 0:
        return list(records)
    rng = random.Random(seed)
    grouped: Dict[str, List[SampleRecord]] = defaultdict(list)
    grouped_by_pet_key: Dict[str, List[SampleRecord]] = defaultdict(list)
    for record in records:
        grouped[record.dataset].append(record)
        grouped_by_pet_key[record.pet_key].append(record)

    selected_by_stem: Dict[str, SampleRecord] = {}

    for pet_key in sorted(grouped_by_pet_key):
        pet_key_candidates = list(grouped_by_pet_key[pet_key])
        rng.shuffle(pet_key_candidates)
        selected_by_stem[pet_key_candidates[0].stem] = pet_key_candidates[0]

    print("Initial atlas estimation subset sizes:")
    for dataset in sorted(grouped):
        candidates = list(grouped[dataset])
        rng.shuffle(candidates)
        dataset_selected = [record for record in selected_by_stem.values() if record.dataset == dataset]
        target_count = max(max_per_dataset, len(dataset_selected))
        for record in candidates:
            if len(dataset_selected) >= min(target_count, len(candidates)):
                break
            if record.stem in selected_by_stem:
                continue
            selected_by_stem[record.stem] = record
            dataset_selected.append(record)
        budget_note = ""
        if len(dataset_selected) > max_per_dataset:
            budget_note = " (expanded to cover all pet keys)"
        print(f"  {dataset}: {len(dataset_selected)} / {len(candidates)}{budget_note}")
    return list(selected_by_stem.values())


def create_model(input_shape: Tuple[int, int, int]) -> SITReg:
    feature_extractor = EncoderFeatureExtractor(
        n_input_channels=1,
        activation_factory=ReLUFactory(),
        n_features_per_resolution=[12, 16, 32, 64, 128, 128],
        n_convolutions_per_resolution=[2, 2, 2, 2, 2, 2],
        input_shape=input_shape,
        normalizer_factory=GroupNormalizerFactory(2),
    ).cuda()
    forward_solver = AndersonSolver(
        MaxElementWiseAbsStopCriterion(min_iterations=2, max_iterations=50, threshold=1e-2),
        AndersonSolverArguments(memory_length=4),
    )
    backward_solver = AndersonSolver(
        RelativeL2ErrorStopCriterion(min_iterations=2, max_iterations=50, threshold=1e-2),
        AndersonSolverArguments(memory_length=4),
    )
    return SITReg(
        feature_extractor=feature_extractor,
        n_transformation_convolutions_per_resolution=[2, 2, 2, 2, 2, 2],
        n_transformation_features_per_resolution=[12, 64, 128, 256, 256, 256],
        max_control_point_multiplier=0.99,
        affine_transformation_type=None,
        input_voxel_size=(1.0, 1.0, 1.0),
        input_shape=input_shape,
        transformation_downsampling_factor=(1.0, 1.0, 1.0),
        forward_fixed_point_solver=forward_solver,
        backward_fixed_point_solver=backward_solver,
        activation_factory=ReLUFactory(),
        normalizer_factory=GroupNormalizerFactory(4),
    ).cuda()


def adjust_learning_rate(optimizer: optim.Optimizer, epoch: int, max_epochs: int, init_lr_model: float, init_lr_atlas: float, power: float = 0.9) -> None:
    for index, param_group in enumerate(optimizer.param_groups):
        base_lr = init_lr_model if index == 0 else init_lr_atlas
        param_group["lr"] = round(base_lr * np.power(1 - epoch / max_epochs, power), 8)


def make_affine_from_pixdim(pixdim: Sequence[float]) -> np.ndarray:
    affine = np.eye(4)
    affine[0, 0] = pixdim[0]
    affine[1, 1] = pixdim[1]
    affine[2, 2] = pixdim[2]
    return affine


def save_nifti(tensor: torch.Tensor, output_path: Path, pixdim: Sequence[float] = (2.8, 2.8, 3.8)) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    image = nib.Nifti1Image(tensor.detach().cpu().numpy()[0, 0], affine=make_affine_from_pixdim(pixdim))
    nib.save(image, str(output_path))


def save_preview_atlases(
    ct_atlases: nn.ParameterDict,
    pet_atlases: nn.ParameterDict,
    atlas_ct_dir: Path,
    atlas_pet_dir: Path,
) -> None:
    for key, tensor in ct_atlases.items():
        save_nifti(tensor, atlas_ct_dir / key / "preview_latest.nii.gz")
    for key, tensor in pet_atlases.items():
        save_nifti(tensor, atlas_pet_dir / key / "preview_latest.nii.gz")


def prune_old_files(pattern: str, max_keep: int = 4) -> None:
    files = natsorted(glob.glob(pattern))
    while len(files) > max_keep:
        os.remove(files[0])
        files = natsorted(glob.glob(pattern))


def save_checkpoint(state: Dict[str, object], output_path: Path, max_keep: int = 4) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    torch.save(state, str(output_path))
    prune_old_files(str(output_path.parent / "*.pth.tar"), max_keep=max_keep)


def seed_all(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def estimate_initial_atlases(dataset: WholebodyAtlasDataset) -> Tuple[Dict[str, torch.Tensor], Dict[str, torch.Tensor]]:
    ct_sums: Dict[str, torch.Tensor] = {}
    pet_sums: Dict[str, torch.Tensor] = {}
    ct_counts: Counter = Counter()
    pet_counts: Counter = Counter()
    for index in range(len(dataset)):
        print(f"Estimating atlases: processing sample {index + 1} of {len(dataset)}", end="\r")
        sample = dataset[index]
        ct_key = str(sample["ct_key"])
        pet_key = str(sample["pet_key"])
        ct_tensor = sample["ct"].float().cpu().clone()
        pet_tensor = sample["pet"].float().cpu().clone()
        if ct_key not in ct_sums:
            ct_sums[ct_key] = torch.zeros_like(ct_tensor)
        if pet_key not in pet_sums:
            pet_sums[pet_key] = torch.zeros_like(pet_tensor)
        ct_sums[ct_key] += ct_tensor
        pet_sums[pet_key] += pet_tensor
        ct_counts[ct_key] += 1
        pet_counts[pet_key] += 1
    ct_means = {key: torch.clamp((tensor / ct_counts[key]).unsqueeze(0), min=0.0, max=1.0) for key, tensor in ct_sums.items()}
    pet_means = {key: torch.clamp((tensor / pet_counts[key]).unsqueeze(0), min=0.0, max=1.0) for key, tensor in pet_sums.items()}
    print()
    return ct_means, pet_means


def _extract_uniform_key(values: object, name: str) -> str:
    if isinstance(values, (list, tuple)):
        if not values:
            raise ValueError(f"Batch field {name} is empty.")
        first = values[0]
        if any(value != first for value in values[1:]):
            raise ValueError(f"Batch contains mixed {name} values: {values}")
        return first
    return values


def to_device(batch: Dict[str, object]) -> Tuple[torch.Tensor, torch.Tensor, str, str]:
    ct = batch["ct"].cuda(non_blocking=True).float()
    pet = batch["pet"].cuda(non_blocking=True).float()
    ct_key = _extract_uniform_key(batch["ct_key"], "ct_key")
    pet_key = _extract_uniform_key(batch["pet_key"], "pet_key")
    return ct, pet, ct_key, pet_key


def average_batch_scalar(value: object) -> float:
    if isinstance(value, torch.Tensor):
        return float(value.detach().float().mean().item())
    if isinstance(value, np.ndarray):
        return float(np.asarray(value, dtype=np.float32).mean())
    if isinstance(value, (list, tuple)):
        if not value:
            return 0.0
        return float(sum(float(item) for item in value) / len(value))
    return float(value)


def cuda_sync() -> None:
    if torch.cuda.is_available():
        torch.cuda.synchronize()


def autocast_context(enabled: bool):
    return torch.cuda.amp.autocast(enabled=enabled)


def serialize_atlases(atlases: nn.ParameterDict) -> Dict[str, torch.Tensor]:
    return {key: value.detach().cpu() for key, value in atlases.items()}


def expand_atlas_to_batch(atlas: torch.Tensor, batch_size: int) -> torch.Tensor:
    if atlas.shape[0] == batch_size:
        return atlas
    if atlas.shape[0] != 1:
        raise ValueError(f"Atlas batch dimension must be 1 or match batch size, got {atlas.shape[0]} vs {batch_size}")
    return atlas.expand(batch_size, *atlas.shape[1:])


def main() -> None:
    args = parse_args()
    if args.batch_size > 1 and args.interleave_by != "pet_key":
        raise ValueError("batch_size > 1 requires --interleave-by pet_key so each batch shares a single CT/PET atlas pair.")
    if not torch.cuda.is_available():
        raise RuntimeError("CUDA is required for this training script.")

    seed_all(args.seed)
    target_shape = tuple(int(v) for v in args.target_shape)
    dataset_flip_axes = parse_dataset_flip_config(args.dataset_flips)
    standardize_orientation = not args.disable_orientation_standardization
    summarize_orientation_policy(dataset_flip_axes, standardize_orientation)

    output_root = Path.cwd() / args.output_root
    experiment_dir = output_root / "experiments"
    atlas_ct_dir = output_root / "atlas" / "ct"
    atlas_pet_dir = output_root / "atlas" / "pet"
    logs_dir = output_root / "logs"
    preflight_dir = output_root / "preflight_previews"
    skip_log_path = output_root / "skip_log.csv"
    for directory in (experiment_dir, atlas_ct_dir, atlas_pet_dir, logs_dir, preflight_dir):
        directory.mkdir(parents=True, exist_ok=True)
    initialize_skip_log(skip_log_path)

    records = build_records()
    if not records:
        raise RuntimeError("No usable PET/CT pairs were found.")
    save_jhu_fdg_exclusion_report(output_root / "jhu_fdg_excluded_suv_cases.csv")
    summarize_records(records)
    save_manifest(records, output_root / "manifest_all.csv")

    train_records, val_records = split_train_val(records, args.val_fraction, args.seed)
    if not train_records:
        raise RuntimeError("No training records remained after the train/validation split.")
    save_manifest(train_records, output_root / "manifest_train.csv")
    save_manifest(val_records, output_root / "manifest_val.csv")
    print(f"Training on {len(train_records)} scans, validating on {len(val_records)} scans")

    train_dataset = WholebodyAtlasDataset(
        train_records,
        target_shape=target_shape,
        skip_log_path=skip_log_path,
        dataset_flip_axes=dataset_flip_axes,
        standardize_orientation=standardize_orientation,
    )
    val_dataset = WholebodyAtlasDataset(
        val_records,
        target_shape=target_shape,
        skip_log_path=skip_log_path,
        dataset_flip_axes=dataset_flip_axes,
        standardize_orientation=standardize_orientation,
    )
    print_dataset_shape_preflight(records, train_dataset, seed=args.seed, output_dir=preflight_dir)
    train_sampler = InterleavedGroupBatchSampler(train_records, group_by=args.interleave_by, batch_size=args.batch_size, seed=args.seed)
    train_loader = DataLoader(train_dataset, batch_sampler=train_sampler, num_workers=args.num_workers, pin_memory=True)
    val_loader = DataLoader(val_dataset, batch_size=1, shuffle=False, num_workers=args.num_workers, pin_memory=True)

    init_records = sample_records_for_initialization(
        train_records,
        max_per_dataset=args.init_samples_per_dataset,
        seed=args.seed,
    )
    init_dataset = WholebodyAtlasDataset(
        init_records,
        target_shape=target_shape,
        skip_log_path=skip_log_path,
        dataset_flip_axes=dataset_flip_axes,
        standardize_orientation=standardize_orientation,
    )
    print("Estimating initial CT and PET atlases...")
    init_start_time = time.perf_counter()
    initial_ct_atlases, initial_pet_atlases = estimate_initial_atlases(init_dataset)
    init_duration = time.perf_counter() - init_start_time
    print(f"Initial atlas estimation took {init_duration:.2f} sec ({init_duration / 60.0:.2f} min)")

    registration_backbone = create_model(target_shape)
    model = TemplateCreation(registration_backbone, target_shape, use_sitreg=True).cuda()
    spatial_trans = SpatialTransformer(target_shape).cuda()

    ct_atlases = nn.ParameterDict(
        {key: nn.Parameter(value.cuda()) for key, value in sorted(initial_ct_atlases.items())}
    )
    pet_atlases = nn.ParameterDict(
        {key: nn.Parameter(value.cuda()) for key, value in sorted(initial_pet_atlases.items())}
    )

    optimizer = optim.AdamW(
        [
            {"params": model.parameters(), "lr": args.lr_model},
            {"params": ct_atlases.parameters(), "lr": args.lr_atlas},
            {"params": pet_atlases.parameters(), "lr": args.lr_atlas},
        ],
        lr=args.lr_model,
        weight_decay=0,
        amsgrad=True,
    )

    criterion_img = SSIM3D()
    criterion_mse = nn.MSELoss()
    criterion_reg = Grad3d(penalty="l2")
    writer = SummaryWriter(log_dir=str(logs_dir))
    best_ssim = -math.inf
    scaler = torch.cuda.amp.GradScaler(enabled=args.amp)
    print(f"AMP {'enabled' if args.amp else 'disabled'}")
    global_step = 0
    if args.profile_timing:
        writer.add_scalar("Timing/initial_atlas_estimation_sec", init_duration, 0)

    for epoch in range(args.max_epochs):
        epoch_start_time = time.perf_counter()
        train_sampler.set_epoch(epoch)
        adjust_learning_rate(optimizer, epoch, args.max_epochs, args.lr_model, args.lr_atlas)
        model.train()
        train_loss = AverageMeter()
        train_img_loss = AverageMeter()
        train_ct_loss = AverageMeter()
        train_pet_loss = AverageMeter()
        train_mean_stream_loss = AverageMeter()
        train_reg_loss = AverageMeter()
        train_data_wait_time = AverageMeter()
        train_sample_load_time = AverageMeter()
        train_h2d_time = AverageMeter()
        train_forward_time = AverageMeter()
        train_loss_time = AverageMeter()
        train_backward_time = AverageMeter()
        train_optim_time = AverageMeter()
        train_iter_time = AverageMeter()
        timing_measured_steps = 0
        previous_iter_end_time = time.perf_counter()

        for step, batch in enumerate(train_loader, start=1):
            global_step += 1
            iter_start_time = time.perf_counter()
            data_wait_time = iter_start_time - previous_iter_end_time
            sample_load_time = average_batch_scalar(batch.get("sample_load_time_sec", 0.0))

            h2d_start_time = time.perf_counter()
            moving_ct, moving_pet, ct_key, pet_key = to_device(batch)
            if args.profile_timing:
                cuda_sync()
            h2d_time = time.perf_counter() - h2d_start_time

            batch_size = moving_ct.shape[0]
            ct_atlas = expand_atlas_to_batch(ct_atlases[ct_key], batch_size)
            pet_atlas = expand_atlas_to_batch(pet_atlases[pet_key], batch_size)

            forward_start_time = time.perf_counter()
            with autocast_context(False):
                deformed_ct_atlas, deformed_ct_image, pos_flow, neg_flow, mean_stream = model((ct_atlas.float(), moving_ct.float()))
                deformed_pet_atlas = spatial_trans(pet_atlas.float(), pos_flow.float())
                deformed_pet_image = spatial_trans(moving_pet.float(), neg_flow.float())
            if args.profile_timing:
                cuda_sync()
            forward_time = time.perf_counter() - forward_start_time

            loss_start_time = time.perf_counter()
            loss_pet_1 = criterion_img(deformed_pet_atlas.float(), moving_pet.float()) * args.weights[0] / 2.0
            loss_pet_2 = criterion_img(deformed_pet_image.float(), pet_atlas.float()) * args.weights[0] / 2.0
            loss_ct_1 = criterion_img(deformed_ct_atlas.float(), moving_ct.float()) * args.weights[0] / 2.0
            loss_ct_2 = criterion_img(deformed_ct_image.float(), ct_atlas.float()) * args.weights[0] / 2.0
            loss_mean_stream = criterion_mse(mean_stream.float(), torch.zeros_like(mean_stream, dtype=torch.float32)) * args.weights[1]
            loss_reg = criterion_reg(pos_flow.float(), moving_ct.float()) * args.weights[2]
            loss_ct = loss_ct_1 + loss_ct_2
            loss_pet = loss_pet_1 + loss_pet_2
            loss_img = (loss_ct + loss_pet) / 2.0
            loss = loss_img + loss_mean_stream + loss_reg
            if args.profile_timing:
                cuda_sync()
            loss_time = time.perf_counter() - loss_start_time

            backward_start_time = time.perf_counter()
            optimizer.zero_grad()
            scaler.scale(loss).backward()
            if args.profile_timing:
                cuda_sync()
            backward_time = time.perf_counter() - backward_start_time

            optim_start_time = time.perf_counter()
            scaler.step(optimizer)
            scaler.update()
            if args.profile_timing:
                cuda_sync()
            optim_time = time.perf_counter() - optim_start_time

            iter_time = time.perf_counter() - iter_start_time
            previous_iter_end_time = time.perf_counter()

            train_loss.update(loss.item(), moving_ct.numel())
            train_img_loss.update(loss_img.item(), moving_ct.numel())
            train_ct_loss.update(loss_ct.item(), moving_ct.numel())
            train_pet_loss.update(loss_pet.item(), moving_ct.numel())
            train_mean_stream_loss.update(loss_mean_stream.item(), moving_ct.numel())
            train_reg_loss.update(loss_reg.item(), moving_ct.numel())
            if step > args.timing_warmup_steps:
                timing_measured_steps += 1
                train_data_wait_time.update(data_wait_time, 1)
                train_sample_load_time.update(sample_load_time, 1)
                train_h2d_time.update(h2d_time, 1)
                train_forward_time.update(forward_time, 1)
                train_loss_time.update(loss_time, 1)
                train_backward_time.update(backward_time, 1)
                train_optim_time.update(optim_time, 1)
                train_iter_time.update(iter_time, 1)
            print(
                f"Epoch {epoch:04d} Iter {step:04d}/{len(train_loader):04d} "
                f"loss {loss.item():.4f} "
                f"img_ssim {loss_img.item():.4f} "
                f"ct_ssim {loss_ct.item():.4f} "
                f"pet_ssim {loss_pet.item():.4f} "
                f"mean {loss_mean_stream.item():.4f} "
                f"reg {loss_reg.item():.4f} "
                f"ct_key={ct_key} pet_key={pet_key}"
            )
            if args.profile_timing and step > args.timing_warmup_steps and (step % max(1, args.timing_log_every) == 0 or step == len(train_loader)):
                print(
                    f"  timing wait={data_wait_time:.3f}s "
                    f"sample_load={sample_load_time:.3f}s "
                    f"h2d={h2d_time:.3f}s "
                    f"forward={forward_time:.3f}s "
                    f"loss={loss_time:.3f}s "
                    f"backward={backward_time:.3f}s "
                    f"optim={optim_time:.3f}s "
                    f"iter={iter_time:.3f}s"
                )
            if args.preview_save_every_steps > 0 and global_step % args.preview_save_every_steps == 0:
                preview_start_time = time.perf_counter()
                save_preview_atlases(ct_atlases, pet_atlases, atlas_ct_dir, atlas_pet_dir)
                preview_time = time.perf_counter() - preview_start_time
                print(f"Saved preview atlases at global step {global_step} in {preview_time:.2f}s")
                if args.profile_timing:
                    writer.add_scalar("Timing/preview_save_sec", preview_time, global_step)
        writer.add_scalar("Loss/train", train_loss.avg, epoch)
        writer.add_scalar("Loss/train_img", train_img_loss.avg, epoch)
        writer.add_scalar("Loss/train_ct", train_ct_loss.avg, epoch)
        writer.add_scalar("Loss/train_pet", train_pet_loss.avg, epoch)
        writer.add_scalar("Loss/train_mean_stream", train_mean_stream_loss.avg, epoch)
        writer.add_scalar("Loss/train_reg", train_reg_loss.avg, epoch)
        if args.profile_timing and timing_measured_steps > 0:
            writer.add_scalar("Timing/train_data_wait_sec", train_data_wait_time.avg, epoch)
            writer.add_scalar("Timing/train_sample_load_sec", train_sample_load_time.avg, epoch)
            writer.add_scalar("Timing/train_h2d_sec", train_h2d_time.avg, epoch)
            writer.add_scalar("Timing/train_forward_sec", train_forward_time.avg, epoch)
            writer.add_scalar("Timing/train_loss_sec", train_loss_time.avg, epoch)
            writer.add_scalar("Timing/train_backward_sec", train_backward_time.avg, epoch)
            writer.add_scalar("Timing/train_optimizer_sec", train_optim_time.avg, epoch)
            writer.add_scalar("Timing/train_iter_sec", train_iter_time.avg, epoch)
        print(
            f"Epoch {epoch:04d} training loss {train_loss.avg:.4f} "
            f"img_ssim {train_img_loss.avg:.4f} "
            f"ct_ssim {train_ct_loss.avg:.4f} "
            f"pet_ssim {train_pet_loss.avg:.4f} "
            f"mean {train_mean_stream_loss.avg:.4f} "
            f"reg {train_reg_loss.avg:.4f}"
        )
        if args.profile_timing and timing_measured_steps > 0:
            print(
                f"Epoch {epoch:04d} timing train_wait {train_data_wait_time.avg:.3f}s "
                f"sample_load {train_sample_load_time.avg:.3f}s "
                f"h2d {train_h2d_time.avg:.3f}s "
                f"forward {train_forward_time.avg:.3f}s "
                f"loss {train_loss_time.avg:.3f}s "
                f"backward {train_backward_time.avg:.3f}s "
                f"optim {train_optim_time.avg:.3f}s "
                f"iter {train_iter_time.avg:.3f}s"
            )
        eval_ssim = AverageMeter()
        val_data_wait_time = AverageMeter()
        val_h2d_time = AverageMeter()
        val_forward_time = AverageMeter()
        val_metric_time = AverageMeter()
        val_iter_time = AverageMeter()
        val_timing_measured_steps = 0
        if len(val_dataset) > 0:
            model.eval()
            with torch.no_grad():
                previous_val_iter_end_time = time.perf_counter()
                for batch in val_loader:
                    val_iter_start_time = time.perf_counter()
                    val_wait_time = val_iter_start_time - previous_val_iter_end_time
                    val_h2d_start_time = time.perf_counter()
                    moving_ct, _, ct_key, _ = to_device(batch)
                    if args.profile_timing:
                        cuda_sync()
                    val_h2d = time.perf_counter() - val_h2d_start_time
                    ct_atlas = ct_atlases[ct_key]

                    val_forward_start_time = time.perf_counter()
                    with autocast_context(False):
                        _, _, pos_flow, _, _ = model((ct_atlas.float(), moving_ct.float()))
                        deformed = spatial_trans(ct_atlas.float(), pos_flow.float())
                    if args.profile_timing:
                        cuda_sync()
                    val_forward = time.perf_counter() - val_forward_start_time

                    val_metric_start_time = time.perf_counter()
                    ssim = 1.0 - criterion_img(deformed.float(), moving_ct.float())
                    if args.profile_timing:
                        cuda_sync()
                    val_metric = time.perf_counter() - val_metric_start_time

                    val_iter_time_total = time.perf_counter() - val_iter_start_time
                    previous_val_iter_end_time = time.perf_counter()
                    eval_ssim.update(ssim.item(), moving_ct.size(0))
                    if val_timing_measured_steps >= args.timing_warmup_steps:
                        val_data_wait_time.update(val_wait_time, 1)
                        val_h2d_time.update(val_h2d, 1)
                        val_forward_time.update(val_forward, 1)
                        val_metric_time.update(val_metric, 1)
                        val_iter_time.update(val_iter_time_total, 1)
                    val_timing_measured_steps += 1
        current_ssim = eval_ssim.avg if eval_ssim.count > 0 else float("nan")
        writer.add_scalar("SSIM/validate", current_ssim, epoch)
        print(f"Epoch {epoch:04d} validation SSIM {current_ssim:.6f}")
        if args.profile_timing and val_iter_time.count > 0:
            writer.add_scalar("Timing/val_data_wait_sec", val_data_wait_time.avg, epoch)
            writer.add_scalar("Timing/val_h2d_sec", val_h2d_time.avg, epoch)
            writer.add_scalar("Timing/val_forward_sec", val_forward_time.avg, epoch)
            writer.add_scalar("Timing/val_metric_sec", val_metric_time.avg, epoch)
            writer.add_scalar("Timing/val_iter_sec", val_iter_time.avg, epoch)
            print(
                f"Epoch {epoch:04d} validation timing wait {val_data_wait_time.avg:.3f}s "
                f"h2d {val_h2d_time.avg:.3f}s "
                f"forward {val_forward_time.avg:.3f}s "
                f"metric {val_metric_time.avg:.3f}s "
                f"iter {val_iter_time.avg:.3f}s"
            )

        if eval_ssim.count > 0 and current_ssim > best_ssim:
            best_ssim = current_ssim

        should_save = (epoch + 1) % args.save_every == 0 or epoch == args.max_epochs - 1
        checkpoint_time = 0.0
        if should_save:
            checkpoint_start_time = time.perf_counter()
            checkpoint_name = f"epoch{epoch + 1:04d}_ssim{current_ssim:.4f}.pth.tar"
            save_checkpoint(
                {
                    "epoch": epoch + 1,
                    "model_state_dict": model.state_dict(),
                    "optimizer_state_dict": optimizer.state_dict(),
                    "best_ssim": best_ssim,
                    "ct_atlases": serialize_atlases(ct_atlases),
                    "pet_atlases": serialize_atlases(pet_atlases),
                    "target_shape": target_shape,
                    "train_manifest": str(output_root / "manifest_train.csv"),
                },
                experiment_dir / checkpoint_name,
            )
            for key, tensor in ct_atlases.items():
                save_nifti(tensor, atlas_ct_dir / key / f"epoch{epoch + 1:04d}_ssim{current_ssim:.4f}.nii.gz")
                prune_old_files(str(atlas_ct_dir / key / "*.nii.gz"))
            for key, tensor in pet_atlases.items():
                save_nifti(tensor, atlas_pet_dir / key / f"epoch{epoch + 1:04d}_ssim{current_ssim:.4f}.nii.gz")
                prune_old_files(str(atlas_pet_dir / key / "*.nii.gz"))
            checkpoint_time = time.perf_counter() - checkpoint_start_time
            if args.profile_timing:
                writer.add_scalar("Timing/checkpoint_sec", checkpoint_time, epoch)
                print(f"Epoch {epoch:04d} checkpoint/save time {checkpoint_time:.2f}s")

        epoch_duration = time.perf_counter() - epoch_start_time
        writer.add_scalar("Timing/epoch_total_sec", epoch_duration, epoch)
        if args.profile_timing:
            print(
                f"Epoch {epoch:04d} total time {epoch_duration:.2f}s "
                f"({epoch_duration / 60.0:.2f} min)"
                + (f" checkpoint {checkpoint_time:.2f}s" if should_save else "")
            )

    writer.close()


if __name__ == "__main__":
    main()